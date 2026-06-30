# -*- coding: utf-8 -*-
"""
دوال التبادل الودي النظيفة.

Phase 3I-a-1: نقل دوال العرض/التحليل النصي النظيفة الخاصة بالتبادل
من app.py إلى swaps.py دون أي اعتماد على Gradio أو app.py.
"""

import os
import datetime
import re
import urllib.parse

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw, ImageFont

from config import APP_DIR

from schedules import get_teacher_choices

from storage import (
    state_locked,
    teachers_db,
    swap_db,
    save_swap_db,
    get_now_oman,
    write_audit_log,
    SCHOOL_WEEK_DAYS,
    ensure_data_directories,
    IMG_DIR,
    SWAP_IMG_DIR,
)


SWAP_EMPTY_MSG = "💡 يرجى اختيار أحد المعلمين من القائمة بالأعلى لتوليد مسودة رسالة الواتساب هنا..."


def get_date_of_weekday(target_day_name):
    days_map = {"الأحد": 6, "الإثنين": 0, "الثلاثاء": 1, "الأربعاء": 2, "الخميس": 3}
    target_weekday = days_map.get(target_day_name, 6)
    now = get_now_oman()
    diff = (target_weekday - now.weekday()) % 7
    target_date = now + datetime.timedelta(days=diff)
    return target_date.strftime("%Y-%m-%d")

candidate_font_paths = [
    os.path.join(APP_DIR, "Cairo-Regular.ttf"),
    "/app/Cairo-Regular.ttf",
    "./Cairo-Regular.ttf",
]
font_path = next((p for p in candidate_font_paths if os.path.exists(p)), None)

image_font_candidate_paths = [
    os.path.join(APP_DIR, "Amiri-Regular.ttf"),
    "/app/Amiri-Regular.ttf",
    "./Amiri-Regular.ttf",
]
image_font_path = next((p for p in image_font_candidate_paths if os.path.exists(p)), None)

def get_current_day_oman():
    weekday = get_now_oman().weekday()
    days_map = {6: "الأحد", 0: "الإثنين", 1: "الثلاثاء", 2: "الأربعاء", 3: "الخميس", 4: "الأحد", 5: "الأحد"}
    return days_map.get(weekday, "الأحد")

def get_class_dna(class_string):
    s = str(class_string).strip()
    s = s.translate(str.maketrans('٠١٢٣٤٥٦٧٨٩', '0123456789')) 
    s = s.replace("ـ", "") 
    if not s: return ""
    
    nums = re.findall(r'\d+', s)
    section = nums[-1] if nums else ""
    
    grade = ""
    if any(x in s for x in ["عاشر", "10", "١٠"]): grade = "10"
    elif any(x in s for x in ["تاسع", "9", "٩"]): grade = "9"
    elif any(x in s for x in ["ثامن", "8", "٨"]): grade = "8"
    elif any(x in s for x in ["سابع", "7", "٧"]): grade = "7"
    elif any(x in s for x in ["حادي", "11", "١١"]): grade = "11"
    elif any(x in s for x in ["ثاني", "12", "١٢"]): grade = "12"
    
    if grade and section: return f"G{grade}-{section}"
    return re.sub(r'[^\w\dأ-ي]', '', s) 

def check_teacher_load(teacher_name, day_name, period_to_add):
    try:
        if teacher_name not in teachers_db: return ""
        info = teachers_db[teacher_name]
        base_p = {int(k) for k in info.get(day_name, {}).keys() if str(k).isdigit()}
        
        if str(period_to_add).isdigit():
            all_slots = sorted(list(base_p | {int(period_to_add)}))
        else:
            all_slots = sorted(list(base_p))
            
        consecutive = max_con = 1
        for i in range(len(all_slots)-1):
            if all_slots[i+1] == all_slots[i] + 1:
                consecutive += 1
                max_con = max(max_con, consecutive)  # ← داخل الحلقة
            else:
                consecutive = 1
            
        warns = []
        if max_con >= 3: warns.append("⚠️ إجهاد بدني")
        if len(all_slots) >= 6: warns.append("⚠️ كثافة عالية")
        return " | ".join(warns)
    except Exception:
        return ""

def run_radar_safe_core(t, p, d):
    """يرجع قائمة المرشحين للتبادل الودي دون أي اعتماد على Gradio."""
    t = str(t or "").split(" (")[0].strip()
    try:
        if not t or not p or "لا يوجد" in t or "اختر" in p:
            return []

        p_str_clean = extract_clean_period_number(p)
        if not p_str_clean.isdigit():
            return []
        p_int = int(p_str_clean)

        t_cls = teachers_db.get(t, {}).get(
            d, {}
        ).get(
            str(p_int),
            teachers_db.get(t, {}).get(d, {}).get(p_int, "")
        )
        if not t_cls:
            return ["❌ لا توجد حصة مسجلة لك"]

        dna = get_class_dna(t_cls)
        perf, flex = [], []

        day_weights = {"الأحد": 1, "الإثنين": 2, "الثلاثاء": 3, "الأربعاء": 4, "الخميس": 5}
        current_day_str = get_current_day_oman()
        current_weight = day_weights.get(current_day_str, 1)

        for tb, info in teachers_db.items():
            if tb == t or info.get("dept") == "الهيئة الإدارية" or info.get("role") == "إداري":
                continue
            if str(p_int) in info.get(d, {}) or p_int in info.get(d, {}):
                continue

            for db in SCHOOL_WEEK_DAYS:
                db_weight = day_weights.get(db, 1)
                db_display = f"{db} القادم" if db_weight < current_weight else db

                for pb, cb in info.get(db, {}).items():
                    if dna == get_class_dna(cb) and dna != "":
                        w_b = check_teacher_load(tb, d, p_int)
                        is_t_free = True
                        if str(pb) in teachers_db.get(t, {}).get(db, {}):
                            is_t_free = False
                        elif str(pb).isdigit() and int(str(pb)) in teachers_db.get(t, {}).get(db, {}):
                            is_t_free = False

                        if is_t_free:
                            w_a = check_teacher_load(t, db, pb)
                            warns = []
                            if w_b:
                                warns.append(f"إجهاد لـ {tb}: {w_b}")
                            if w_a:
                                warns.append(f"إجهاد لك: {w_a}")
                            w_str = f" ⚠️ ({' | '.join(warns)})" if warns else ""
                            perf.append(f"🟢 تبادل مثالي | البديل: {tb} | يغطيك ({d} ح{p_int}) وتغطيه ({db_display} ح{pb}){w_str}")
                        else:
                            w_str = f" ⚠️ (إجهاد لـ {tb}: {w_b})" if w_b else ""
                            flex.append(f"🟠 إنقاذ مرن | البديل: {tb} | يغطيك ({d} ح{p_int}) لكنك مشغول وقت حصته ({db_display} ح{pb}){w_str}")

        res = sorted(list(set(perf))) + sorted(list(set(flex)))
        if not res:
            return [f"❌ لا يوجد بديل متفرغ (بصمة: {dna})"]
        return res
    except Exception:
        return ["خطأ داخلي"]


def generate_wa_msg_core(choice, t_req, p_req, d_req):
    """يبني مسودة رسالة واتساب وزرها كقيم خام دون أي اعتماد على Gradio."""
    if not choice or "❌" in str(choice) or "خطأ" in str(choice):
        return SWAP_EMPTY_MSG, ""
    try:
        parts = str(choice).split("|")
        t_target = parts[1].split(":")[1].strip()
        details = parts[2].strip()

        p_req_clean = extract_clean_period_number(p_req)
        t_req_clean = str(t_req or "").split(" (")[0].strip()
        req_class_raw = teachers_db.get(t_req_clean, {}).get(
            d_req, {}
        ).get(
            p_req_clean,
            teachers_db.get(t_req_clean, {}).get(
                d_req, {}
            ).get(int(p_req_clean) if p_req_clean.isdigit() else p_req_clean, "")
        )
        req_class_elegant = format_elegant_class(req_class_raw)

        msg = f"السلام عليكم ورحمة الله وبركاته أستاذي العزيز ({t_target})\n\n"
        msg += f"يرغب الأستاذ ({t_req}) بالتبادل الودي معك (بعد إذنك وموافقتك طبعاً لظرف طارئ).\n"
        msg += f"ستقوم أنت مشكوراً بتغطية الصف ({req_class_elegant}) في الحصة ({p_req_clean}) يوم ({d_req}).\n"

        if "مثالي" in str(choice):
            rep_part = details.split("وتغطيه ")[1].split(")")[0].replace("(", "")
            rep_day, rep_period = rep_part.split(" ح")

            clean_rep_day = rep_day.replace(" القادم", "").strip()
            target_class_raw = teachers_db.get(t_target, {}).get(
                clean_rep_day, {}
            ).get(
                str(rep_period),
                teachers_db.get(t_target, {}).get(
                    clean_rep_day, {}
                ).get(int(rep_period) if str(rep_period).isdigit() else rep_period, "")
            )
            target_class_elegant = format_elegant_class(target_class_raw)

            msg += f"وسيقوم الأستاذ ({t_req}) بتغطية الصف ({target_class_elegant}) في الحصة ({rep_period}) يوم ({rep_day}) بدلاً عنك.\n\n"
        else:
            msg += f"ونظراً لانشغال الأستاذ ({t_req}) وقت حصتك، سيتم التنسيق لرد الحصة لاحقاً.\n\n"

        msg += "هل يناسبك هذا التبادل ليتم اعتماده؟ شاكرين ومقدرين تعاونك 🤝"

        phone = teachers_db.get(t_target, {}).get("phone", "")
        btn_color = "#25D366"

        if phone:
            phone = "".join(filter(str.isdigit, str(phone)))
            if len(phone) == 8:
                phone = "968" + phone
            btn_text = f"✅ إرسال للأستاذ {t_target}"
        else:
            phone = ""
            btn_text = "⚠️ إرسال (لا يوجد رقم)"

        encoded_msg = urllib.parse.quote(msg)
        wa_link = f"https://api.whatsapp.com/send?phone={phone}&text={encoded_msg}"

        btn_html = (
            f'<div style="margin-top: 10px; border: 2px solid {btn_color}; border-radius: 8px; padding: 2px;">'
            f'<a href="{wa_link}" target="_blank" '
            f'style="display: block; width: 100%; text-align: center; background-color: {btn_color}; color: white; '
            f'padding: 12px; border-radius: 6px; font-weight: bold; text-decoration: none; font-size: 16px;">'
            f'{btn_text}</a></div>'
        )

        return msg, btn_html

    except Exception:
        return SWAP_EMPTY_MSG, ""


def on_swap_option_selected_core(choice, t, period_value, d):
    """يرجع: (msg_value, btn_value, is_interactive) دون أي اعتماد على Gradio."""
    if not choice or "❌" in str(choice):
        return SWAP_EMPTY_MSG, "", False

    msg_value, btn_value = generate_wa_msg_core(choice, t, period_value, d)
    return msg_value, btn_value, True


def get_swap_candidates_for_period_core(t, period_value, d, confirmed_state):
    """يرجع بيانات مرشحي التبادل الخام لحصة محددة دون أي اعتماد على Gradio."""
    empty_msg = SWAP_EMPTY_MSG

    if not t or not period_value:
        return [], None, empty_msg, "", False

    candidates = run_radar_safe_core(t, period_value, d)

    if not candidates:
        candidates = ["❌ لا يوجد بديل متفرغ"]

    p_clean = extract_clean_period_number(period_value)

    saved_choice = None
    saved_message = empty_msg
    btn_value = ""
    confirm_interactive = False

    if isinstance(confirmed_state, dict) and p_clean in confirmed_state:
        saved_choice = confirmed_state[p_clean].get("choice")
        if saved_choice not in candidates:
            saved_choice = None

        if saved_choice:
            saved_message = confirmed_state[p_clean].get("message", empty_msg) or empty_msg
            _, btn_value = generate_wa_msg_core(saved_choice, t, period_value, d)
            confirm_interactive = True

    return candidates, saved_choice, saved_message, btn_value, confirm_interactive

def build_swap_button_html(candidate_name, message_text):
    phone = teachers_db.get(candidate_name, {}).get("phone", "")
    btn_color = "#25D366"

    if phone:
        phone = "".join(filter(str.isdigit, str(phone)))
        if len(phone) == 8:
            phone = "968" + phone
        btn_text = f"✅ إرسال للأستاذ {candidate_name}"
    else:
        phone = ""
        btn_text = f"⚠️ إرسال (لا يوجد رقم)"

    encoded_msg = urllib.parse.quote(message_text)
    wa_link = f"https://api.whatsapp.com/send?phone={phone}&text={encoded_msg}" if phone else f"https://api.whatsapp.com/send?text={encoded_msg}"

    return (
        f'<div style="margin-top: 10px; border: 2px solid {btn_color}; border-radius: 8px; padding: 2px;">'
        f'<a href="{wa_link}" target="_blank" '
        f'style="display: block; width: 100%; text-align: center; background-color: {btn_color}; color: white; '
        f'padding: 12px; border-radius: 6px; font-weight: bold; text-decoration: none; font-size: 16px;">'
        f'{btn_text}</a></div>'
    )


def extract_swap_choice_details(choice):
    candidate = ""
    comp_day = "يحدد لاحقاً"
    comp_period = "يحدد لاحقاً"

    try:
        parts = choice.split("|", 2)

        if len(parts) > 1 and ":" in parts[1]:
            candidate = parts[1].split(":", 1)[1].strip()
        else:
            candidate = str(choice).strip()

        details = parts[2].strip() if len(parts) > 2 else ""

        if "وتغطيه " in details:
            rep_part = details.split("وتغطيه ", 1)[1].split(")", 1)[0].replace("(", "")
            rep_day, rep_period = rep_part.split(" ح", 1)
            comp_day = rep_day.strip()
            comp_period = f"الحصة {rep_period.strip()}"

    except Exception:
        candidate = str(choice).strip()

    return candidate, comp_day, comp_period


def render_swap_table_html(state):
    if not isinstance(state, dict) or not state:
        return """
        <div style='background:#f8fafc; border:1px dashed #cbd5e1; border-radius:10px; padding:14px; text-align:center; color:#64748b; direction:rtl;'>
            لا توجد تبادلات معتمدة بعد.
        </div>
        """

    rows_html = ""
    for p, info in sorted(state.items(), key=lambda x: int(x[0])):
        rows_html += f"""
        <tr>
            <td style='padding:12px; border:1px solid #d1d5db;'>{info.get('requester', '')}</td>
            <td style='padding:12px; border:1px solid #d1d5db;'>{info.get('class', '')}</td>
            <td style='padding:12px; border:1px solid #d1d5db;'>الحصة {p}</td>
            <td style='padding:12px; border:1px solid #d1d5db;'>{info.get('candidate', '')}</td>
            <td style='padding:12px; border:1px solid #d1d5db;'>{info.get('comp_day', 'يحدد لاحقاً')}</td>
            <td style='padding:12px; border:1px solid #d1d5db;'>{info.get('comp_period', 'يحدد لاحقاً')}</td>
        </tr>
        """

    return f"""
    <div style='overflow-x:auto; direction:rtl; margin-top:12px;'>
        <table style='width:100%; min-width:900px; border-collapse:collapse; text-align:center; font-family:Cairo, Arial, sans-serif;'>
            <thead>
                <tr style='background:#e8f5e9; color:#0f5132;'>
                    <th style='padding:12px; border:1px solid #d1d5db;'>المعلم الطالب للتبادل</th>
                    <th style='padding:12px; border:1px solid #d1d5db;'>الصف</th>
                    <th style='padding:12px; border:1px solid #d1d5db;'>الحصة</th>
                    <th style='padding:12px; border:1px solid #d1d5db;'>المعلم البديل</th>
                    <th style='padding:12px; border:1px solid #d1d5db;'>يوم التعويض</th>
                    <th style='padding:12px; border:1px solid #d1d5db;'>حصة التعويض</th>
                </tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
        </table>
    </div>
    """

def extract_clean_period_number(period_value):
    raw = str(period_value).split("-")[0]
    raw = raw.replace("✅", "").replace("الحصة", "").strip()
    return raw if raw.isdigit() else ""


def format_elegant_class(raw_class):
    raw_class = str(raw_class).strip()
    if not raw_class:
        return "الصف غير محدد"
    words = raw_class.split()
    if len(words) < 2:
        return raw_class
    grade_part = ""
    subject_part = ""
    for i, word in enumerate(reversed(words)):
        if any(g in word for g in ["ثامن", "تاسع", "عاشر", "حادي", "ثاني", "1", "2", "3", "4", "5", "6", "7", "8", "9"]):
            grade_part = word
            subject_part = " ".join(words[:len(words) - 1 - i])
            break
    if grade_part and subject_part:
        return f"{grade_part} - مادة {subject_part}"
    return raw_class


def load_confirmed_swaps_for_context_core(t, d):
    """يرجع حالة التبادلات المعتمدة للمعلم/اليوم دون أي اعتماد على Gradio."""
    t = str(t or "").split(" (")[0].strip()
    state = {}

    if not t or not d:
        return state

    for _, info in swap_db.items():
        if info.get("requester") == t and info.get("day") == d:
            p = str(info.get("period", "")).strip()
            if not p:
                continue

            state[p] = {
                "requester": info.get("requester", ""),
                "class": info.get("class", ""),
                "candidate": info.get("candidate", ""),
                "choice": info.get("choice", ""),
                "message": info.get("message", ""),
                "comp_day": info.get("comp_day", "يحدد لاحقاً"),
                "comp_period": info.get("comp_period", "يحدد لاحقاً"),
            }

    return state


def clear_swap_detail_ui_core():
    """تُرجع: (choices, selected_value, message_value, button_html, confirm_interactive)."""
    return [], None, SWAP_EMPTY_MSG, "", False



def filter_swap_teachers_safe_core(dept):
    """يرجع اختيارات معلمي التبادل كقيم خام دون أي اعتماد على Gradio."""
    try:
        choices = get_teacher_choices(dept if dept != "الكل" else "الكل")
        if not choices:
            return ["لا يوجد معلمون"], None
        return choices, None
    except Exception:
        return [], None


def get_teacher_periods_safe_core(t, d):
    """يرجع حصص المعلم كقيم خام دون أي اعتماد على Gradio."""
    try:
        if t and t in teachers_db and t != "لا يوجد معلمون":
            periods_elegant = []
            for k, v in teachers_db[t].get(d, {}).items():
                if str(k).isdigit() and str(v).strip() != "" and str(v).lower() != "nan":
                    elegant_c = format_elegant_class(v)
                    display_text = f"الحصة {k} - ({elegant_c})"
                    periods_elegant.append(display_text)
            periods_elegant.sort(key=lambda x: int(x.split("-")[0].replace("الحصة", "").strip()))
            if not periods_elegant:
                return ["لا توجد حصص"], None
            return periods_elegant, None
        return ["اختر معلماً أولاً"], None
    except Exception:
        return ["خطأ داخلي"], None

def get_teacher_periods_marked_core(t, d, confirmed_state, current_value=None):
    """يرجع اختيارات حصص المعلم مع تعليم الحصص المعتمدة دون أي اعتماد على Gradio."""
    t = str(t or "").split(" (")[0].strip()
    try:
        if not t or t not in teachers_db or t == "لا يوجد معلمون":
            return ["اختر معلماً أولاً"], None

        confirmed_keys = set()
        if isinstance(confirmed_state, dict):
            confirmed_keys = {str(k) for k in confirmed_state.keys()}

        choices = []
        selected_value = None
        current_clean = extract_clean_period_number(current_value)

        for k, v in teachers_db[t].get(d, {}).items():
            if str(k).isdigit() and str(v).strip() != "" and str(v).lower() != "nan":
                elegant_c = format_elegant_class(v)
                prefix = "✅ " if str(k) in confirmed_keys else ""
                display_text = f"{prefix}الحصة {k} - ({elegant_c})"
                choices.append((int(k), display_text))

        choices.sort(key=lambda x: x[0])
        final_choices = [text for _, text in choices]

        if current_clean:
            for k, text in choices:
                if str(k) == current_clean:
                    selected_value = text
                    break

        if not final_choices:
            return ["لا توجد حصص"], None

        return final_choices, selected_value

    except Exception:
        return ["خطأ داخلي"], None


def format_period_label(period_value):
    raw = str(period_value or "").strip()
    if not raw:
        return ""
    if raw.startswith("الحصة"):
        return raw
    return f"الحصة {raw}"



def generate_swap_table_image_core(
    state,
    teacher_name,
    day_name,
    system_name,
    system_subtitle,
    theme_color,
    accent_color,
):
    """توليد صورة التبادل الودي كمسار ملف خام دون أي اعتماد على Gradio."""
    if not isinstance(state, dict) or not state:
        return None

    try:
        ensure_data_directories()
        os.makedirs(SWAP_IMG_DIR, exist_ok=True)

        target_date = get_date_of_weekday(day_name)

        rows = []
        for p, info in sorted(state.items(), key=lambda x: int(x[0])):
            rows.append({
                "المعلم الطالب": str(info.get("requester", "")),
                "الصف": str(info.get("class", "")),
                "الحصة": f"الحصة {p}",
                "المعلم البديل": str(info.get("candidate", "")),
                "يوم التعويض": str(info.get("comp_day", "يحدد لاحقاً")),
                "حصة التعويض": str(info.get("comp_period", "يحدد لاحقاً")),
            })

        pil_font_path = None
        for candidate in [
            image_font_path,
            os.path.join(APP_DIR, "Amiri-Regular.ttf"),
            "/app/Amiri-Regular.ttf",
            "./Amiri-Regular.ttf",
            font_path,
            os.path.join(APP_DIR, "Cairo-Regular.ttf"),
            "/app/Cairo-Regular.ttf",
            "./Cairo-Regular.ttf",
        ]:
            if candidate and os.path.exists(candidate):
                pil_font_path = candidate
                break

        def load_font(size, bold=False):
            try:
                if pil_font_path:
                    return ImageFont.truetype(pil_font_path, size=size)
            except Exception:
                pass
            try:
                fallback = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
                return ImageFont.truetype(fallback, size=size)
            except Exception:
                return ImageFont.load_default()

        font_title = load_font(40, bold=True)
        font_subtitle = load_font(27, bold=False)
        font_header = load_font(25, bold=True)
        font_cell = load_font(24, bold=False)
        font_footer = load_font(24, bold=True)

        temp_img = Image.new("RGB", (10, 10), "white")
        temp_draw = ImageDraw.Draw(temp_img)

        def text_size(value, font):
            text_value = "" if value is None else str(value)
            bbox = temp_draw.textbbox((0, 0), text_value, font=font)
            return bbox[2] - bbox[0], bbox[3] - bbox[1]

        def draw_text_right(draw, x_right, y_top, value, font, fill):
            text_value = "" if value is None else str(value)
            w, h = text_size(text_value, font)
            draw.text((x_right - w, y_top), text_value, font=font, fill=fill)
            return w, h

        def draw_text_center(draw, box, value, font, fill):
            x1, y1, x2, y2 = box
            text_value = "" if value is None else str(value)
            w, h = text_size(text_value, font)
            draw.text((x1 + ((x2 - x1) - w) / 2, y1 + ((y2 - y1) - h) / 2 - 2), text_value, font=font, fill=fill)

        def wrap_text_by_width(value, font, max_width):
            text_value = "" if value is None else str(value).strip()
            if not text_value:
                return [""]

            words = text_value.split()
            if len(words) <= 1:
                return [text_value]

            lines = []
            current = words[0]
            for word in words[1:]:
                trial = current + " " + word
                trial_w, _ = text_size(trial, font)
                if trial_w <= max_width:
                    current = trial
                else:
                    lines.append(current)
                    current = word
            lines.append(current)
            return lines if lines else [text_value]

        def draw_multiline_center(draw, box, value, font, fill, line_gap=5):
            x1, y1, x2, y2 = box
            max_width = max(40, int((x2 - x1) - 18))
            lines = wrap_text_by_width(value, font, max_width)
            line_heights = [text_size(line, font)[1] for line in lines]
            total_h = sum(line_heights) + max(0, len(lines) - 1) * line_gap
            y = y1 + ((y2 - y1) - total_h) / 2

            for line, h in zip(lines, line_heights):
                w, _ = text_size(line, font)
                draw.text((x1 + ((x2 - x1) - w) / 2, y), line, font=font, fill=fill)
                y += h + line_gap

        columns = [
            ("المعلم الطالب", 245),
            ("الصف", 270),
            ("الحصة", 125),
            ("المعلم البديل", 245),
            ("يوم التعويض", 165),
            ("حصة التعويض", 165),
        ]

        margin = 42
        table_width = sum(width for _, width in columns)
        image_width = table_width + margin * 2
        header_h = 135
        table_header_h = 58
        base_row_h = 64

        row_heights = []
        for row in rows:
            max_lines = 1
            for col_name, col_w in columns:
                max_lines = max(max_lines, len(wrap_text_by_width(row.get(col_name, ""), font_cell, col_w - 18)))
            row_heights.append(max(base_row_h, 44 + max_lines * 30))

        image_height = header_h + table_header_h + sum(row_heights) + 58
        image = Image.new("RGB", (image_width, image_height), "#ffffff")
        draw = ImageDraw.Draw(image)

        header_bg = theme_color
        draw.rectangle((0, 0, image_width, header_h), fill=header_bg)

        title = "جدول التبادلات الودية المعتمدة"
        subtitle = f"{teacher_name or 'الكل'} | {day_name} | {target_date}"

        title_w, title_h = text_size(title, font_title)
        subtitle_w, subtitle_h = text_size(subtitle, font_subtitle)
        draw.text(((image_width - title_w) / 2, 24), title, font=font_title, fill=accent_color)
        draw.text(((image_width - subtitle_w) / 2, 78), subtitle, font=font_subtitle, fill="#ffffff")

        y = header_h
        x_right = image_width - margin

        header_fill = "#e8f5e9"
        header_text = "#004d40"
        border = "#cbd5e1"
        row_fill_1 = "#ffffff"
        row_fill_2 = "#f8faf8"
        text_fill = "#1f2937"

        x = x_right
        for col_name, col_w in columns:
            x1 = x - col_w
            draw.rectangle((x1, y, x, y + table_header_h), fill=header_fill, outline=border)
            draw_multiline_center(draw, (x1, y, x, y + table_header_h), col_name, font_header, header_text)
            x = x1

        y += table_header_h

        for idx, row in enumerate(rows):
            row_h = row_heights[idx]
            bg = row_fill_1 if idx % 2 == 0 else row_fill_2
            x = x_right

            for col_name, col_w in columns:
                x1 = x - col_w
                draw.rectangle((x1, y, x, y + row_h), fill=bg, outline=border)
                draw_multiline_center(draw, (x1, y, x, y + row_h), row.get(col_name, ""), font_cell, text_fill)
                x = x1

            y += row_h

        footer_text = f"{system_name} {system_subtitle}"
        footer_w, footer_h = text_size(footer_text, font_footer)
        draw.text(((image_width - footer_w) / 2, image_height - 39), footer_text, font=font_footer, fill=theme_color)

        filename = os.path.join(
            SWAP_IMG_DIR,
            f"swap_table_{get_now_oman().strftime('%Y%m%d_%H%M%S_%f')}.png"
        )
        image.save(filename)
        return filename

    except Exception as e:
        print(f"generate_swap_table_image error: {e}")
        return None

def draw_schedule_image_core(df, day_name):
    target_date = get_date_of_weekday(day_name)
    absent_list = df["المعلم الغائب"].astype(str).unique().tolist() if df is not None and not df.empty else []
    absent_list = [str(name).strip() for name in absent_list if str(name).strip()]

    def chunk_absent_names(names, chunk_size=3):
        if not names:
            return ["لا يوجد"]
        return ["، ".join(names[i:i + chunk_size]) for i in range(0, len(names), chunk_size)]

    absent_lines = chunk_absent_names(absent_list, 3)

    display_df = df[["المعلم الغائب", "الصف", "الحصة", "المعلم البديل عرض"]].copy()
    display_df.columns = ["المعلم الغائب", "الصف", "الحصة", "المعلم البديل"]

    title_text = f"📅 {day_name} | {target_date}"
    absent_label_text = "المعلمون الغائبون:"

    pil_font_path = None
    for candidate in [image_font_path, os.path.join(APP_DIR, "Amiri-Regular.ttf"), "/app/Amiri-Regular.ttf", "./Amiri-Regular.ttf"]:
        if candidate and os.path.exists(candidate):
            pil_font_path = candidate
            break

    def load_font(size, bold=False):
        try:
            if pil_font_path:
                return ImageFont.truetype(pil_font_path, size=size)
        except Exception:
            pass
        try:
            fallback = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
            return ImageFont.truetype(fallback, size=size)
        except Exception:
            return ImageFont.load_default()

    font_title = load_font(40, bold=True)
    font_subtitle = load_font(28, bold=False)
    font_header = load_font(26, bold=True)
    font_cell = load_font(24, bold=False)

    temp_img = Image.new("RGB", (10, 10), "white")
    temp_draw = ImageDraw.Draw(temp_img)

    def text_size(value, font):
        bbox = temp_draw.textbbox((0, 0), str(value), font=font)
        return bbox[2] - bbox[0], bbox[3] - bbox[1]

    def draw_text_right(draw, x_right, y_top, value, font, fill):
        text = "" if value is None else str(value)
        w, h = text_size(text, font)
        draw.text((x_right - w, y_top), text, font=font, fill=fill)
        return w, h

    def wrap_text_by_width(value, font, max_width):
        text = "" if value is None else str(value).strip()
        if not text:
            return [""]
        words = text.split()
        if len(words) <= 1:
            return [text]

        lines = []
        current = words[0]
        for word in words[1:]:
            trial = current + " " + word
            trial_w, _ = text_size(trial, font)
            if trial_w <= max_width:
                current = trial
            else:
                lines.append(current)
                current = word
        lines.append(current)
        return lines if lines else [text]

    def draw_multiline_right(draw, x_right, y_top, lines, font, fill, line_gap=4):
        y = y_top
        max_w = 0
        total_h = 0
        for line in lines:
            w, h = text_size(line, font)
            draw.text((x_right - w, y), line, font=font, fill=fill)
            y += h + line_gap
            total_h += h + line_gap
            max_w = max(max_w, w)
        if total_h > 0:
            total_h -= line_gap
        return max_w, total_h

    def sanitize_image_substitute_display(value):
        """تنظيف نص المعلم البديل في الصورة فقط دون تغيير منطق التبادل أو الواجهة."""
        text = "" if value is None else str(value)
        text = text.replace(chr(0x1F91D), "")
        text = re.sub(r"\s+\)", ")", text)
        text = re.sub(r"\(\s+", "(", text)
        text = re.sub(r"\s{2,}", " ", text).strip()
        return text

    pad_x = 40
    pad_y = 30
    title_h = text_size(title_text, font_title)[1]
    label_h = text_size(absent_label_text, font_subtitle)[1]
    absent_line_h = text_size("نص", font_subtitle)[1]
    absent_line_gap = 6
    header_h = 28 + title_h + 14 + label_h + 8 + (len(absent_lines) * absent_line_h) + max(0, len(absent_lines) - 1) * absent_line_gap + 24
    header_h = max(130, header_h)
    gap_after_header = 20
    base_row_h = 58
    border_color = "#cfd8dc"
    outer_border = "#b0bec5"
    header_bg = "#004d40"
    header_fg = "#ffffff"
    alt_bg = "#f8faf8"
    white_bg = "#ffffff"
    red_bg = "#ffebee"
    teal_bg = "#e0f2f1"
    orange_bg = "#ffebee"
    text_dark = "#1f2937"
    title_fg = "#ffffff"

    columns = [
        ("المعلم الغائب", 280),
        ("الصف", 360),
        ("الحصة", 110),
        ("المعلم البديل", 340),
    ]
    col_width_map = dict(columns)

    prepared_rows = []
    row_heights = []
    line_height = text_size("نص", font_cell)[1]

    for _, row in display_df.iterrows():
        sub_display = str(row.get("المعلم البديل", ""))
        status = ""
        if "❌" in sub_display:
            status = "تقصير"
        elif "🤝" in sub_display:
            status = "تبادل"
        elif "إشراف" in sub_display:
            status = "إشراف"

        class_lines = wrap_text_by_width(
            str(row.get("الصف", "")),
            font_cell,
            max_width=col_width_map["الصف"] - 24
        )

        row_values = {
            "المعلم الغائب": str(row.get("المعلم الغائب", "")),
            "الصف": class_lines,
            "الحصة": str(row.get("الحصة", "")),
            "المعلم البديل": sanitize_image_substitute_display(sub_display),
            "_status": status,
        }
        prepared_rows.append(row_values)
        dynamic_h = max(base_row_h, (len(class_lines) * line_height) + 22)
        row_heights.append(dynamic_h)

    table_w = sum(width for _, width in columns)
    img_w = table_w + pad_x * 2
    img_h = header_h + gap_after_header + base_row_h + sum(row_heights) + pad_y * 2 + 10

    image = Image.new("RGB", (img_w, img_h), "white")
    draw = ImageDraw.Draw(image)

    draw.rounded_rectangle((pad_x, pad_y, img_w - pad_x, pad_y + header_h), radius=18, fill=header_bg)
    header_x_right = img_w - pad_x - 20
    title_y = pad_y + 18
    draw_text_right(draw, header_x_right, title_y, title_text, font_title, title_fg)

    label_y = title_y + title_h + 14
    draw_text_right(draw, header_x_right, label_y, absent_label_text, font_subtitle, title_fg)

    line_y = label_y + label_h + 8
    for line in absent_lines:
        draw_text_right(draw, header_x_right, line_y, line, font_subtitle, title_fg)
        line_y += absent_line_h + absent_line_gap

    table_top = pad_y + header_h + gap_after_header
    header_y2 = table_top + base_row_h

    x_cursor = img_w - pad_x
    for col_name, col_w in columns:
        x1 = x_cursor - col_w
        x2 = x_cursor
        draw.rectangle((x1, table_top, x2, header_y2), fill=header_bg, outline=outer_border, width=1)
        draw_text_right(draw, x2 - 16, table_top + 12, col_name, font_header, header_fg)
        x_cursor = x1

    current_y = header_y2
    for idx, row in enumerate(prepared_rows, start=1):
        row_h = row_heights[idx - 1]
        y1 = current_y
        y2 = y1 + row_h

        status = row["_status"]
        if status == "تقصير":
            row_bg = red_bg
        elif status == "تبادل":
            row_bg = teal_bg
        elif status == "إشراف":
            row_bg = orange_bg
        else:
            row_bg = alt_bg if idx % 2 == 0 else white_bg

        row_values = [
            row["المعلم الغائب"],
            row["الصف"],
            row["الحصة"],
            row["المعلم البديل"],
        ]

        x_cursor = img_w - pad_x
        for (_, col_w), value in zip(columns, row_values):
            x1 = x_cursor - col_w
            x2 = x_cursor
            draw.rectangle((x1, y1, x2, y2), fill=row_bg, outline=border_color, width=1)

            if isinstance(value, list):
                content_h = (len(value) * line_height) + ((len(value) - 1) * 4)
                text_y = y1 + max(10, (row_h - content_h) / 2)
                draw_multiline_right(draw, x2 - 12, text_y, value, font_cell, text_dark, line_gap=4)
            else:
                _, text_h = text_size(value, font_cell)
                text_y = y1 + max(10, (row_h - text_h) / 2)
                draw_text_right(draw, x2 - 12, text_y, value, font_cell, text_dark)

            x_cursor = x1

        current_y = y2

    ensure_data_directories()
    tz_oman = datetime.timezone(datetime.timedelta(hours=4))
    filename = os.path.join(IMG_DIR, f"output_{day_name}_{target_date}_{datetime.datetime.now(tz_oman).strftime('%H%M%S_%f')}.png")
    image.save(filename)
    return filename

def export_confirmed_swaps_excel_core():
    """يصدر سجل التبادلات المعتمدة إلى Excel ويُرجع اسم الملف النسبي أو None.

    يحافظ عمدًا على السلوك القديم: يكتب الملف باسم نسبي في مجلد العمل الحالي،
    ولا يغيّر مسار الحفظ ضمن هذه المرحلة المعمارية.
    """
    if not isinstance(swap_db, dict) or not swap_db:
        return None

    rows = []
    for _, info in sorted(swap_db.items(), key=lambda item: (
        str(item[1].get("updated_at", "")),
        str(item[1].get("requester", "")),
        str(item[1].get("day", "")),
        str(item[1].get("period", "")),
    )):
        updated_at = str(info.get("updated_at", "")).strip()
        approval_date = updated_at.split(" ")[0] if updated_at else ""
        rows.append({
            "المعلم الطالب للتبادل": str(info.get("requester", "")),
            "المعلم البديل": str(info.get("candidate", "")),
            "الصف": str(info.get("class", "")),
            "اليوم الأصلي": str(info.get("day", "")),
            "الحصة الأصلية": format_period_label(info.get("period", "")),
            "يوم التعويض": str(info.get("comp_day", "")),
            "حصة التعويض": str(info.get("comp_period", "")),
            "التاريخ": approval_date,
        })

    if not rows:
        return None

    df = pd.DataFrame(rows)
    filename = f"سجل_التبادلات_الودية_المعتمدة_{get_now_oman().strftime('%Y%m%d_%H%M%S')}.xlsx"

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='التبادلات المعتمدة')
        ws = writer.sheets['التبادلات المعتمدة']

        header_fill = PatternFill(fill_type='solid', fgColor='0B6E4F')
        header_font = Font(color='FFFFFF', bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = center_alignment

        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                cell_value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(cell_value))
            ws.column_dimensions[column_letter].width = min(max(max_length + 4, 14), 40)

        ws.freeze_panes = 'A2'
        ws.sheet_view.rightToLeft = True

    return filename

@state_locked
def confirm_swap_core(t, period_value, choice, d, msg_text, state, actor_name="", actor_role=""):
    """يعتمد تبادلًا وديًا ويُرجع الحالة الحالية ورسالة تحذير خام إن وجدت."""
    t = str(t or "").split(" (")[0].strip()
    current_state = dict(state) if isinstance(state, dict) else {}

    if not t or not period_value or not choice or "❌" in str(choice):
        return current_state, ""

    p_clean = extract_clean_period_number(period_value)

    req_class_raw = teachers_db.get(t, {}).get(
        d, {}
    ).get(
        p_clean,
        teachers_db.get(t, {}).get(d, {}).get(int(p_clean) if p_clean.isdigit() else p_clean, "")
    )

    elegant_class = format_elegant_class(req_class_raw)
    candidate, comp_day, comp_period = extract_swap_choice_details(choice)

    # ── فحص محلي (داخل الحالة الحالية للمعلم) ──
    for p_ex, info_ex in current_state.items():
        if (
            info_ex.get("comp_day") == comp_day
            and info_ex.get("comp_period") == comp_period
            and p_ex != p_clean
        ):
            return (
                current_state,
                f"<div style='color:red; padding:10px; text-align:center;'>⚠️ موعد التعويض ({comp_day} - {comp_period}) محجوز مسبقاً لهذا المعلم.</div>"
            )

    # ── فحص عالمي (على جميع التبادلات المعتمدة) ──
    current_key = f"{t}|{d}|{p_clean}"
    for key, info in swap_db.items():
        same_comp = (
            info.get("comp_day") == comp_day
            and info.get("comp_period") == comp_period
        )
        if not same_comp:
            continue

        if info.get("requester") == t and key != current_key:
            return (
                current_state,
                f"<div style='color:red; padding:10px; text-align:center;'>⚠️ موعد التعويض ({comp_day} - {comp_period}) محجوز مسبقاً لهذا المعلم.</div>"
            )

        if info.get("candidate") == candidate and key != current_key:
            return (
                current_state,
                f"<div style='color:red; padding:10px; text-align:center;'>⚠️ موعد التعويض ({comp_day} - {comp_period}) محجوز مسبقاً على المعلم البديل.</div>"
            )

    current_state[p_clean] = {
        "requester": t,
        "class": elegant_class,
        "candidate": candidate,
        "choice": choice,
        "message": msg_text,
        "comp_day": comp_day,
        "comp_period": comp_period,
    }

    swap_db[current_key] = {
        "requester": t,
        "day": d,
        "period": p_clean,
        "class": elegant_class,
        "candidate": candidate,
        "choice": choice,
        "message": msg_text,
        "comp_day": comp_day,
        "comp_period": comp_period,
        "updated_at": get_now_oman().strftime("%Y-%m-%d %H:%M"),
    }
    save_swap_db()

    write_audit_log(
        "اعتماد تبادل ودي",
        target_teacher=t,
        old_value="",
        new_value={
            "day": d,
            "period": p_clean,
            "class": elegant_class,
            "candidate": candidate,
            "comp_day": comp_day,
            "comp_period": comp_period,
        },
        details=f"اعتماد تبادل ودي بين {t} و {candidate}",
        actor_name=actor_name,
        actor_role=actor_role,
    )

    return current_state, ""

