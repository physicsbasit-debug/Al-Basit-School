import os
import gradio as gr
import shutil
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import datetime
import matplotlib
import random
import json
import re
import urllib.parse
import tempfile
import threading
import functools
import zipfile
import html as html_lib
import base64
import mimetypes
import secrets
matplotlib.use('Agg')  
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import arabic_reshaper
from bidi.algorithm import get_display
from PIL import Image, ImageDraw, ImageFont

from config import (
    APP_DIR,
    PAGE_SIZE,
    LOCAL_DATA_DIR,
    DEFAULT_SCHOOL_CONFIG,
    ADMIN_ROLES,
    ALL_ROLES,
)

from storage import (
    DATA_DIR,
    PERSISTENT_STORAGE_ACTIVE,
    STORAGE_MODE,
    STORAGE_ERROR,
    DB_FILE,
    DAILY_DB_FILE,
    SWAP_DB_FILE,
    AUTH_DB_FILE,
    IMG_DIR,
    SWAP_IMG_DIR,
    SCHEDULES_DIR,
    BACKUPS_DIR,
    EXPORTS_DIR,
    BRANDING_DIR,
    REFERENCE_STATUS_FILE,
    AUTH_ACCOUNTS_FILE,
    MIGRATION_STATUS_FILE,
    STATE_LOCK,
    _get_json_file_lock,
    state_locked,
    ADMIN_FILE,
    PHONES_FILE,
    EXEMPTIONS_LOG_FILE,
    AUDIT_LOG_FILE,
    SCHOOL_CONFIG_FILE,
    SCHEDULE_FILES,
    load_school_config,
    SCHOOL_CONFIG,
    MAX_PERIODS,
    SCHOOL_WEEK_DAYS,
    SCHOOL_WEEKEND_DAYS,
    OFFICIAL_DEPTS,
    ensure_data_directories,
    safe_write_json,
    teachers_db,
    daily_db,
    processed_absences,
    exemptions_log,
    swap_db,
    save_db,
    load_db,
    save_exemptions_log,
    load_exemptions_log,
    save_daily_db,
    load_daily_db,
    save_swap_db,
    load_swap_db,
    get_now_oman,
    write_audit_log,
    last_assigned_teachers,
    _queue_audit_change,
    _flush_audit_changes,
)

from auth import (
    OWNER_ROLE,
    SHARED_TEACHER_ROLE,
    ADMIN_ACCESS_ROLES,
    DEPT_LEADER_ROLES,
    OWNER_ACCOUNT_ID,
    AUTH_DB,
    AUTH_ACCOUNTS,
    load_auth_db,
    load_auth_accounts,
    save_auth_accounts,
    authenticate_login_pin,
    _auth_now_text,
    _pin_hash,
    _verify_pin_hash,
    _account_display_name,
    _validate_new_pin,
    _pin_is_used_by_another_account,
    get_auth_account_choices,
    get_permissions,
    get_permissions_from_flags,
    get_ui_visibility_updates,
)

from school_data import (
    load_reference_status_registry,
    save_reference_status_registry,
    update_reference_file_status,
    _reference_status_key,
    get_reference_file_status,
    dept_has_loaded_schedule_data,
    get_school_data_center_status,
    render_reference_file_card,
    render_admin_reference_card,
    render_phones_reference_card,
    render_schedule_reference_cards,
    save_admin_reference_file,
    save_phones_reference_file,
    save_schedule_reference_file,
    precheck_schedule_excel_template,
    render_schedule_precheck_error_html,
    validate_reference_filename,
    refresh_schedule_from_reference_core,
    refresh_admins_from_reference_core,
    refresh_phones_from_reference_core,
    delete_department_data_core,
    process_uploaded_excel_core,
    _normalize_schedule_header_text,
    _excel_column_label_zero_based,
)

from schedules import (
    get_teacher_choices,
    get_absentee_choices,
    resolve_effective_dept,
    clean_teacher_name,
    get_name_fingerprint,
    extract_class_info,
    get_day_overview,
    format_teacher_name,
    get_day_dept_style,
    render_day_department_section_html,
    render_day_all_departments_html,
    render_day_table_html,
    get_day_table_updates_core,
    DAY_DEPT_STYLE_MAP,
    DAY_DEPT_FALLBACK_STYLES,
)

from balances import (
    get_updated_balance,
    get_updated_absences,
    get_updated_shortcomings,
    render_compact_rtl_table_html,
)


from exemptions import (
    is_teacher_exempt_for_slot,
    normalize_exempt_slots,
    build_exempt_slots_from_days_periods,
    format_exempt_slots_for_display,
    render_exemptions_log_html,
    resolve_teacher_key_from_ui,
    clean_teacher_name_from_ui,
    save_teacher_rules_core,
)

from swaps import (
    build_swap_button_html,
    extract_swap_choice_details,
    render_swap_table_html,
    confirm_swap_core,
    extract_clean_period_number,
    format_elegant_class,
    get_current_day_oman,
    get_class_dna,
    check_teacher_load,
    run_radar_safe_core,
    generate_wa_msg_core,
    get_swap_candidates_for_period_core,
    on_swap_option_selected_core,
    load_confirmed_swaps_for_context_core,
    clear_swap_detail_ui_core,
    get_teacher_periods_marked_core,
    filter_swap_teachers_safe_core,
    get_teacher_periods_safe_core,
    export_confirmed_swaps_excel_core,
    generate_swap_table_image_core,
    draw_schedule_image_core,
    font_path,
    image_font_path,
    get_date_of_weekday,
    SWAP_EMPTY_MSG,
)

from distribution import (
    get_falcon_eye_candidates,
    format_sub_display,
    format_sub_display_for_image,
    normalize_absent_names,
    build_generation_signature,
    same_generation_context,
    get_empty_generation_state,
    get_existing_absents_for_context,
    detect_conflicted_absence_slots,
    detect_absence_assignment_conflicts_for_context,
    build_absence_conflict_warning_html,
    generate_styled_html_table,
    generate_whatsapp_html,
    get_teacher_schedule_choices,
    resolve_teacher_display_value,
    resolve_teacher_display_values,
    get_dynamic_header,
    get_initial_header,
    refresh_ui_on_change_core,
    update_available_subs_smart_core,
    assign_logic_core,
    rollback_auto_assignments_for_absentees_core,
    cancel_teacher_absence_core,
    process_admin_action_core,
    update_manual_count_core,
    reset_monthly_balances_core,
    add_manual_staff_core,
    delete_single_teacher_core,
)


# --- 1. الإعدادات والوقت ---
tz_oman = datetime.timezone(datetime.timedelta(hours=4))

# v1.8.5f / Phase 3E-pre
# الحالة العامة المشتركة ودوال الحفظ/التحميل انتقلت إلى storage.py.
# التحديثات تتم in-place للحفاظ على مراجع الكائنات بين الوحدات.
# v1.8.5g / Phase 3E-a
# دوال مركز البيانات النظيفة انتقلت إلى school_data.py.
# v1.8.5h / Phase 3F-a
# دوال الجداول والاختيارات النظيفة انتقلت إلى schedules.py.
# v1.8.5i / Phase 3G-a
# دوال الأرصدة والغياب والتقصير النظيفة انتقلت إلى balances.py.
# v1.8.5o / Phase 3H-a-2
# دوال الإعفاءات النظيفة انتقلت إلى exemptions.py.
# v1.8.5p / Phase 3H-a-3
# save_teacher_rules أصبحت core/wrapper، والمنطق في exemptions.py.
# v1.8.5q / Phase 3I-a-1
# دوال التبادل الودي النظيفة انتقلت إلى swaps.py.
# v1.8.5r / Phase 3I-a-3
# confirm_swap أصبحت core/wrapper، والمنطق في swaps.py.
# v1.8.5s / Phase 3I-a-4a/4b
# run_radar_safe و generate_wa_msg أصبحتا core/wrapper، والمنطق في swaps.py.


# ─────────────────────────────────────────────────────────────────────────────
# v1.8.1 — ترحيل البيانات القديمة وحالة التخزين الدائم
# ─────────────────────────────────────────────────────────────────────────────

LEGACY_ROOT_FILES = {
    os.path.join(APP_DIR, "school_balances.json"): DB_FILE,
    os.path.join(APP_DIR, "daily_assignments.json"): DAILY_DB_FILE,
    os.path.join(APP_DIR, "friendly_swaps.json"): SWAP_DB_FILE,
    os.path.join(APP_DIR, "auth_db.json"): AUTH_DB_FILE,
}

LEGACY_DATA_FILES = {
    os.path.join(LOCAL_DATA_DIR, "admin_staff.xlsx"): ADMIN_FILE,
    os.path.join(LOCAL_DATA_DIR, "teacher_phones.xlsx"): PHONES_FILE,
    os.path.join(LOCAL_DATA_DIR, "exemptions_log.json"): EXEMPTIONS_LOG_FILE,
    os.path.join(LOCAL_DATA_DIR, "audit_log.json"): AUDIT_LOG_FILE,
    os.path.join(LOCAL_DATA_DIR, "school_config.json"): SCHOOL_CONFIG_FILE,
    os.path.join(LOCAL_DATA_DIR, "reference_files_status.json"): REFERENCE_STATUS_FILE,
    os.path.join(LOCAL_DATA_DIR, "auth_accounts.json"): AUTH_ACCOUNTS_FILE,
}

LEGACY_DATA_DIRECTORIES = {
    os.path.join(LOCAL_DATA_DIR, "schedules"): SCHEDULES_DIR,
    os.path.join(LOCAL_DATA_DIR, "backups"): BACKUPS_DIR,
    os.path.join(LOCAL_DATA_DIR, "branding"): BRANDING_DIR,
    os.path.join(LOCAL_DATA_DIR, "exports"): EXPORTS_DIR,
    os.path.join(LOCAL_DATA_DIR, "generated_images"): IMG_DIR,
    os.path.join(APP_DIR, "generated_swap_tables"): SWAP_IMG_DIR,
}

def _atomic_copy_file_if_missing(source_path, destination_path):
    """نسخ الملف القديم فقط إذا لم يوجد في التخزين الدائم."""
    source = os.path.abspath(str(source_path))
    destination = os.path.abspath(str(destination_path))

    if source == destination:
        return "same"
    if not os.path.isfile(source):
        return "missing"
    if os.path.exists(destination):
        return "skipped_existing"

    os.makedirs(os.path.dirname(destination), exist_ok=True)
    temp_destination = f"{destination}.migration_{os.getpid()}.tmp"
    try:
        shutil.copy2(source, temp_destination)
        os.replace(temp_destination, destination)
        return "copied"
    finally:
        try:
            if os.path.exists(temp_destination):
                os.remove(temp_destination)
        except Exception:
            pass

def _copy_directory_missing_only(source_dir, destination_dir):
    copied = 0
    skipped = 0
    errors = []

    source = os.path.abspath(str(source_dir))
    destination = os.path.abspath(str(destination_dir))

    if source == destination or not os.path.isdir(source):
        return copied, skipped, errors

    for root, _dirs, files in os.walk(source):
        rel = os.path.relpath(root, source)
        target_root = destination if rel == "." else os.path.join(destination, rel)
        os.makedirs(target_root, exist_ok=True)

        for filename in files:
            source_file = os.path.join(root, filename)
            target_file = os.path.join(target_root, filename)
            try:
                result = _atomic_copy_file_if_missing(source_file, target_file)
                if result == "copied":
                    copied += 1
                elif result == "skipped_existing":
                    skipped += 1
            except Exception as exc:
                errors.append(f"{source_file}: {exc}")

    return copied, skipped, errors

def migrate_legacy_data_once():
    """ترحيل البيانات السابقة مرة واحدة دون استبدال بيانات الـBucket."""
    ensure_data_directories()
    report = {
        "version": "1.8.1",
        "created_at": datetime.datetime.now(tz_oman).strftime("%Y-%m-%d %H:%M:%S"),
        "persistent_storage_active": bool(PERSISTENT_STORAGE_ACTIVE),
        "data_dir": DATA_DIR,
        "copied_files": [],
        "skipped_existing": [],
        "missing_sources": [],
        "errors": [],
    }

    if not PERSISTENT_STORAGE_ACTIVE:
        report["note"] = "يعمل التطبيق على التخزين المحلي الاحتياطي."
        return report

    if os.path.exists(MIGRATION_STATUS_FILE):
        try:
            with open(MIGRATION_STATUS_FILE, "r", encoding="utf-8") as status_file:
                previous = json.load(status_file)
            if isinstance(previous, dict):
                previous["already_completed"] = True
                return previous
        except Exception:
            pass

    mappings = {}
    mappings.update(LEGACY_ROOT_FILES)
    mappings.update(LEGACY_DATA_FILES)

    for source_path, destination_path in mappings.items():
        try:
            result = _atomic_copy_file_if_missing(source_path, destination_path)
            if result == "copied":
                report["copied_files"].append(destination_path)
            elif result == "skipped_existing":
                report["skipped_existing"].append(destination_path)
            elif result == "missing":
                report["missing_sources"].append(source_path)
        except Exception as exc:
            report["errors"].append(f"{source_path}: {exc}")

    for source_dir, destination_dir in LEGACY_DATA_DIRECTORIES.items():
        try:
            copied, skipped, errors = _copy_directory_missing_only(
                source_dir, destination_dir
            )
            if copied:
                report["copied_files"].append(f"{destination_dir} ({copied} ملف)")
            if skipped:
                report["skipped_existing"].append(
                    f"{destination_dir} ({skipped} ملف موجود)"
                )
            report["errors"].extend(errors)
        except Exception as exc:
            report["errors"].append(f"{source_dir}: {exc}")

    safe_write_json(MIGRATION_STATUS_FILE, report, make_backup=False)
    return report

MIGRATION_REPORT = migrate_legacy_data_once()

def get_persistent_storage_health():
    ensure_data_directories()
    writable = False
    error = ""
    probe_path = os.path.join(DATA_DIR, f".health_probe_{os.getpid()}")

    try:
        with open(probe_path, "w", encoding="utf-8") as probe:
            probe.write("ok")
            probe.flush()
            os.fsync(probe.fileno())
        writable = True
    except Exception as exc:
        error = str(exc)
    finally:
        try:
            if os.path.exists(probe_path):
                os.remove(probe_path)
        except Exception:
            pass

    return {
        "persistent": bool(PERSISTENT_STORAGE_ACTIVE),
        "writable": bool(writable),
        "path": DATA_DIR,
        "mode": STORAGE_MODE,
        "error": error or STORAGE_ERROR,
        "migration": MIGRATION_REPORT,
    }

def render_persistent_storage_status_html():
    health = get_persistent_storage_health()

    if health["persistent"] and health["writable"]:
        title = "التخزين الدائم متصل ويعمل"
        bg, fg, border, icon = "#dcfce7", "#166534", "#16a34a", "✅"
    elif health["writable"]:
        title = "التخزين المحلي يعمل، لكن الـBucket غير متصل"
        bg, fg, border, icon = "#fff7ed", "#9a3412", "#ea580c", "⚠️"
    else:
        title = "مسار التخزين غير قابل للكتابة"
        bg, fg, border, icon = "#fee2e2", "#991b1b", "#dc2626", "❌"

    migration = health.get("migration") or {}
    copied_count = len(migration.get("copied_files", []) or [])
    skipped_count = len(migration.get("skipped_existing", []) or [])
    error_text = html_lib.escape(str(health.get("error", "")))
    error_html = (
        f"<div style='margin-top:6px;'>التفاصيل: {error_text}</div>"
        if error_text else ""
    )

    return f"""
    <div style='background:{bg};color:{fg};padding:13px;border-radius:10px;
                border-right:5px solid {border};margin-bottom:12px;
                font-weight:800;line-height:1.8;direction:rtl;'>
        <div style='font-size:17px;'>{icon} {title}</div>
        <div>المسار الفعلي: <code>{html_lib.escape(DATA_DIR)}</code></div>
        <div>وضع التشغيل: <b>{html_lib.escape(STORAGE_MODE)}</b></div>
        <div>الترحيل: نُسخ {copied_count} بند، وتُرك {skipped_count} بند موجود دون استبدال.</div>
        {error_html}
    </div>
    """


# Phase 3C: load_school_config وSCHOOL_CONFIG انتقلت إلى storage.py.

MINISTRY_NAME = str(DEFAULT_SCHOOL_CONFIG["ministry_name"])
DIRECTORATE_PREFIX = str(DEFAULT_SCHOOL_CONFIG["directorate_prefix"])
DIRECTORATE_REGION = str(SCHOOL_CONFIG.get("directorate_region", DEFAULT_SCHOOL_CONFIG["directorate_region"]))
DIRECTORATE_FULL_NAME = f"{DIRECTORATE_PREFIX} {DIRECTORATE_REGION}".strip()
SYSTEM_NAME = str(DEFAULT_SCHOOL_CONFIG["system_name"])
SYSTEM_SUBTITLE = str(DEFAULT_SCHOOL_CONFIG["system_subtitle"])
SCHOOL_NAME = str(SCHOOL_CONFIG.get("school_name", DEFAULT_SCHOOL_CONFIG["school_name"]))
DEVELOPER_CREDIT = str(DEFAULT_SCHOOL_CONFIG["developer_credit"])
SCHOOL_LOGO_URL = str(SCHOOL_CONFIG.get("logo_url", DEFAULT_SCHOOL_CONFIG["logo_url"]))
THEME_COLOR = str(SCHOOL_CONFIG.get("theme_color", DEFAULT_SCHOOL_CONFIG["theme_color"]))
THEME_COLOR_2 = str(SCHOOL_CONFIG.get("theme_color_2", DEFAULT_SCHOOL_CONFIG["theme_color_2"]))
ACCENT_COLOR = str(SCHOOL_CONFIG.get("accent_color", DEFAULT_SCHOOL_CONFIG["accent_color"]))

# Phase 3C: MAX_PERIODS / SCHOOL_WEEK_DAYS / SCHOOL_WEEKEND_DAYS / OFFICIAL_DEPTS مستوردة من storage.py.


# v1.8.5e / Phase 3D
# طبقة الحسابات والصلاحيات الأساسية انتقلت إلى auth.py.
# تبقى دوال واجهة الحسابات وتخصيص الترحيب داخل app.py مؤقتًا لحماية ربط Gradio.


ACCOUNT_WELCOME_DEFAULT_TEMPLATE = "{welcome_title} ({display_name}) {welcome_phrase}"
ACCOUNT_WELCOME_PLACEHOLDERS = [
    "{name}",
    "{display_name}",
    "{welcome_title}",
    "{official_title}",
    "{whatsapp_title}",
    "{department}",
    "{department_label}",
    "{school_name}",
    "{role}",
]


def _clean_account_profile_value(value):
    return str(value or "").strip()


def _default_department_label(record):
    dept = _clean_account_profile_value(record.get("dept", ""))
    if dept and dept not in {"الكل", "المعلمون", "الهيئة الإدارية"}:
        return f"قسم {dept}"
    return dept


def _account_profile_context(record):
    safe_record = record if isinstance(record, dict) else {}
    name = _clean_account_profile_value(safe_record.get("name", ""))
    role = _clean_account_profile_value(safe_record.get("role", ""))
    dept = _clean_account_profile_value(safe_record.get("dept", ""))
    display_name = _clean_account_profile_value(
        safe_record.get("display_name") or name
    )
    official_title = _clean_account_profile_value(
        safe_record.get("official_title") or role
    )
    whatsapp_title = _clean_account_profile_value(
        safe_record.get("whatsapp_title") or official_title or role
    )
    welcome_title = _clean_account_profile_value(
        safe_record.get("welcome_title", "")
    )
    department_label = _clean_account_profile_value(
        safe_record.get("department_label") or _default_department_label(safe_record)
    )
    welcome_phrase = _clean_account_profile_value(
        safe_record.get("welcome_phrase", "")
    )
    return {
        "name": name,
        "display_name": display_name,
        "official_title": official_title,
        "whatsapp_title": whatsapp_title,
        "welcome_title": welcome_title,
        "department": dept,
        "department_label": department_label,
        "school_name": SCHOOL_NAME,
        "role": role,
        "welcome_phrase": welcome_phrase,
    }


def _safe_format_account_template(template, context):
    template_text = _clean_account_profile_value(template)
    if not template_text:
        return ""
    try:
        return template_text.format(**context)
    except Exception:
        # لا نكسر الدخول بسبب قوس ناقص في عبارة ترحيب. الواجهة ليست مكانًا لتأملات الترايسباك.
        return template_text


def build_account_welcome_text(record):
    context = _account_profile_context(record)
    template = _clean_account_profile_value(record.get("welcome_template", "")) if isinstance(record, dict) else ""

    has_custom_profile = any([
        context.get("welcome_title"),
        context.get("welcome_phrase"),
        template,
        _clean_account_profile_value(record.get("display_name", "")) if isinstance(record, dict) else "",
    ])

    if has_custom_profile:
        if not template:
            template = ACCOUNT_WELCOME_DEFAULT_TEMPLATE
        text_value = _safe_format_account_template(template, context)
        text_value = re.sub(r"\s+", " ", str(text_value or "")).strip()
        # تنظيف الأقواس الفارغة إذا لم يضبط المستخدم بعض الحقول.
        text_value = text_value.replace("()", "").replace("( )", "").strip()
        if text_value:
            return text_value

    role = context.get("role", "")
    dept = context.get("department", "")
    raw_msg = WELCOME_MESSAGES.get(
        role,
        WELCOME_MESSAGES.get(dept, "مرحباً بك ({name}) في النظام."),
    )
    try:
        return raw_msg.format(name=context.get("display_name") or context.get("name"))
    except Exception:
        return raw_msg


def render_account_welcome_html(record, temporary_note=""):
    welcome_text = html_lib.escape(build_account_welcome_text(record))
    return (
        "<div style='background:#004d40;color:#ffca28;padding:15px;"
        "border-radius:10px;text-align:center;font-size:18px;"
        "font-weight:bold;margin-bottom:15px;line-height:1.8;'>"
        f"{welcome_text}{temporary_note}</div>"
    )


def get_account_session_display_name(record):
    context = _account_profile_context(record)
    return context.get("display_name") or context.get("name") or context.get("official_title") or ""


def render_account_profile_preview_html(
    display_name,
    official_title,
    welcome_title,
    department_label,
    welcome_phrase,
    welcome_template,
    whatsapp_title,
):
    fake_record = {
        "name": _clean_account_profile_value(display_name).replace("أ. ", ""),
        "display_name": display_name,
        "official_title": official_title,
        "welcome_title": welcome_title,
        "department_label": department_label,
        "welcome_phrase": welcome_phrase,
        "welcome_template": welcome_template,
        "whatsapp_title": whatsapp_title,
        "dept": department_label,
        "role": official_title,
    }
    welcome_text = html_lib.escape(build_account_welcome_text(fake_record))
    whatsapp = html_lib.escape(_clean_account_profile_value(whatsapp_title) or "غير محدد")
    official = html_lib.escape(_clean_account_profile_value(official_title) or "غير محدد")
    return f"""
    <div style='background:linear-gradient(135deg,#004d40,#0f766e);color:#ffca28;
                padding:16px;border-radius:14px;text-align:center;font-size:18px;
                font-weight:900;line-height:1.9;margin-top:8px;'>
        {welcome_text}
    </div>
    <div style='background:#f8fafc;border:1px solid #dbe3e8;border-radius:12px;
                padding:12px;margin-top:10px;line-height:1.8;color:#334155;'>
        <b>المسمى الرسمي:</b> {official}<br>
        <b>المسمى المختصر:</b> {whatsapp}
    </div>
    """


def preview_account_profile_settings(
    display_name,
    official_title,
    welcome_title,
    department_label,
    welcome_phrase,
    welcome_template,
    whatsapp_title,
    is_owner=False,
):
    if not bool(is_owner):
        return (
            gr.update(),
            "<div style='color:#b91c1c;font-weight:800;'>هذه المعاينة مخصصة لمالك النظام فقط.</div>",
        )
    return (
        gr.update(
            value=render_account_profile_preview_html(
                display_name,
                official_title,
                welcome_title,
                department_label,
                welcome_phrase,
                welcome_template,
                whatsapp_title,
            )
        ),
        gr.update(value=""),
    )


def load_auth_account_profile_for_editor(account_id, is_owner=False):
    if not bool(is_owner):
        return (
            gr.update(value=""),
            gr.update(value=""),
            gr.update(value=""),
            gr.update(value=""),
            gr.update(value=""),
            gr.update(value=""),
            gr.update(value=""),
            gr.update(value=""),
            "<div style='color:#b91c1c;font-weight:800;'>هذه اللوحة مخصصة لمالك النظام فقط.</div>",
        )

    payload = load_auth_accounts()
    record = payload.get("accounts", {}).get(str(account_id or "").strip())
    if not isinstance(record, dict):
        return (
            gr.update(value=""),
            gr.update(value=""),
            gr.update(value=""),
            gr.update(value=""),
            gr.update(value=""),
            gr.update(value=""),
            gr.update(value=""),
            gr.update(value=""),
            "<div style='color:#a16207;font-weight:800;'>اختر حسابًا لعرض تخصيص الترحيب.</div>",
        )

    context = _account_profile_context(record)
    template = _clean_account_profile_value(record.get("welcome_template", "")) or ACCOUNT_WELCOME_DEFAULT_TEMPLATE
    preview = render_account_profile_preview_html(
        context["display_name"],
        context["official_title"],
        context["welcome_title"],
        context["department_label"],
        context["welcome_phrase"],
        template,
        context["whatsapp_title"],
    )
    return (
        gr.update(value=context["display_name"]),
        gr.update(value=context["official_title"]),
        gr.update(value=context["welcome_title"]),
        gr.update(value=context["department_label"]),
        gr.update(value=context["welcome_phrase"]),
        gr.update(value=template),
        gr.update(value=context["whatsapp_title"]),
        gr.update(value=preview),
        gr.update(value=""),
    )


@state_locked
def save_auth_account_profile(
    account_id,
    display_name,
    official_title,
    welcome_title,
    department_label,
    welcome_phrase,
    welcome_template,
    whatsapp_title,
    is_owner=False,
    actor_name="",
    actor_role="",
):
    if not bool(is_owner):
        return (
            gr.update(),
            gr.update(),
            gr.update(),
            "<div style='color:#b91c1c;font-weight:800;'>هذه العملية مخصصة لمالك النظام فقط.</div>",
        )

    account_id = str(account_id or "").strip()
    payload = load_auth_accounts()
    account = payload.get("accounts", {}).get(account_id)
    if not isinstance(account, dict):
        return (
            gr.update(),
            gr.update(),
            gr.update(),
            "<div style='color:#b91c1c;font-weight:800;'>اختر حسابًا صالحًا.</div>",
        )

    old_profile = {
        "display_name": account.get("display_name", ""),
        "official_title": account.get("official_title", ""),
        "welcome_title": account.get("welcome_title", ""),
        "department_label": account.get("department_label", ""),
        "welcome_phrase": account.get("welcome_phrase", ""),
        "welcome_template": account.get("welcome_template", ""),
        "whatsapp_title": account.get("whatsapp_title", ""),
    }

    account["display_name"] = _clean_account_profile_value(display_name)
    account["official_title"] = _clean_account_profile_value(official_title)
    account["welcome_title"] = _clean_account_profile_value(welcome_title)
    account["department_label"] = _clean_account_profile_value(department_label)
    account["welcome_phrase"] = _clean_account_profile_value(welcome_phrase)
    account["welcome_template"] = _clean_account_profile_value(welcome_template)
    account["whatsapp_title"] = _clean_account_profile_value(whatsapp_title)
    account["updated_at"] = _auth_now_text()
    account["profile_updated_at"] = account["updated_at"]

    if not save_auth_accounts(payload):
        return (
            gr.update(),
            gr.update(),
            gr.update(),
            "<div style='color:#b91c1c;font-weight:800;'>تعذر حفظ تخصيص الحساب.</div>",
        )

    new_profile = {
        "display_name": account.get("display_name", ""),
        "official_title": account.get("official_title", ""),
        "welcome_title": account.get("welcome_title", ""),
        "department_label": account.get("department_label", ""),
        "welcome_phrase": account.get("welcome_phrase", ""),
        "welcome_template": account.get("welcome_template", ""),
        "whatsapp_title": account.get("whatsapp_title", ""),
    }

    write_audit_log(
        "تعديل تخصيص حساب دخول",
        target_teacher="",
        old_value=old_profile,
        new_value=new_profile,
        details=f"تعديل هيدر ومسمى حساب: {_account_display_name(account)}",
        actor_name=actor_name,
        actor_role=actor_role,
    )

    choices = get_auth_account_choices()
    preview = render_account_profile_preview_html(
        account.get("display_name", ""),
        account.get("official_title", ""),
        account.get("welcome_title", ""),
        account.get("department_label", ""),
        account.get("welcome_phrase", ""),
        account.get("welcome_template", "") or ACCOUNT_WELCOME_DEFAULT_TEMPLATE,
        account.get("whatsapp_title", ""),
    )
    return (
        gr.update(value=render_auth_accounts_html(True)),
        gr.update(choices=choices, value=account_id),
        gr.update(value=preview),
        (
            "<div style='color:#166534;background:#dcfce7;padding:10px;"
            "border-radius:8px;font-weight:800;'>"
            "تم حفظ تخصيص الترحيب والمسميات بنجاح. سيظهر الهيدر الجديد في تسجيل الدخول القادم."
            "</div>"
        ),
    )

def render_auth_accounts_html(is_owner=False):
    if not bool(is_owner):
        return (
            "<div style='color:#64748b;text-align:center;padding:12px;'>"
            "تظهر الحسابات لمالك النظام بعد الدخول."
            "</div>"
        )

    payload = load_auth_accounts()
    rows = []

    for account_id, record in sorted(
        payload.get("accounts", {}).items(),
        key=lambda item: _account_display_name(item[1]),
    ):
        enabled = bool(record.get("enabled", True))
        status_text = "مفعل" if enabled else "معطل"
        status_color = "#166534" if enabled else "#b91c1c"
        temporary_text = (
            "نعم"
            if bool(record.get("must_change_pin", False))
            else "لا"
        )
        context = _account_profile_context(record)

        rows.append(f"""
        <tr>
            <td>{html_lib.escape(_account_display_name(record))}</td>
            <td>{html_lib.escape(context.get("display_name", "—"))}</td>
            <td>{html_lib.escape(context.get("official_title", "—"))}</td>
            <td>{html_lib.escape(context.get("whatsapp_title", "—"))}</td>
            <td>{html_lib.escape(str(record.get("role", "—")))}</td>
            <td>{html_lib.escape(str(record.get("dept", "—")))}</td>
            <td style='color:{status_color};font-weight:900;'>{status_text}</td>
            <td>{temporary_text}</td>
            <td>{html_lib.escape(str(record.get("updated_at", "—")))}</td>
        </tr>
        """)

    if not rows:
        return (
            "<div style='background:#fff7ed;color:#9a3412;padding:12px;"
            "border-radius:10px;font-weight:800;'>"
            "لا توجد حسابات غير حساب المالك. تحقق من AUTH_DB_JSON ثم أعد التشغيل."
            "</div>"
        )

    return f"""
    <div style='overflow-x:auto;direction:rtl;border:1px solid #dbe3e8;
                border-radius:12px;'>
        <table style='width:100%;min-width:1150px;border-collapse:collapse;
                      text-align:center;font-size:13px;'>
            <thead>
                <tr style='background:#0f766e;color:#fff;'>
                    <th>اسم الحساب</th>
                    <th>اسم العرض</th>
                    <th>المسمى الرسمي</th>
                    <th>المسمى المختصر</th>
                    <th>الدور الداخلي</th>
                    <th>القسم</th>
                    <th>الحالة</th>
                    <th>رمز مؤقت</th>
                    <th>آخر تحديث</th>
                </tr>
            </thead>
            <tbody>{''.join(rows)}</tbody>
        </table>
    </div>
    """


def refresh_owner_accounts_panel(is_owner=False):
    if not bool(is_owner):
        return (
            gr.update(),
            gr.update(choices=[], value=None),
            gr.update(value=""),
            gr.update(
                value=(
                    "<div style='color:#b91c1c;font-weight:800;'>"
                    "هذه اللوحة مخصصة لمالك النظام فقط."
                    "</div>"
                )
            ),
        )

    choices = get_auth_account_choices()
    return (
        gr.update(value=render_auth_accounts_html(True)),
        gr.update(choices=choices, value=None),
        gr.update(value=""),
        gr.update(value=""),
    )

@state_locked
def change_own_account_pin(
    account_id,
    current_pin,
    new_pin,
    confirm_pin,
    actor_name="",
    actor_role="",
    is_owner=False,
):
    if bool(is_owner) or str(account_id) == OWNER_ACCOUNT_ID:
        return (
            gr.update(value=""),
            gr.update(value=""),
            gr.update(value=""),
            (
                "<div style='color:#9a3412;background:#fff7ed;padding:10px;"
                "border-radius:8px;font-weight:800;'>"
                "رمز مالك النظام يُغيّر من Secret الاستضافة، وليس من داخل المنظومة."
                "</div>"
            ),
        )

    account_id = str(account_id or "").strip()
    payload = load_auth_accounts()
    account = payload.get("accounts", {}).get(account_id)

    if not isinstance(account, dict):
        return (
            gr.update(value=""),
            gr.update(value=""),
            gr.update(value=""),
            "<div style='color:#b91c1c;font-weight:800;'>تعذر تحديد حساب الجلسة الحالية.</div>",
        )

    if not bool(account.get("enabled", True)):
        return (
            gr.update(value=""),
            gr.update(value=""),
            gr.update(value=""),
            "<div style='color:#b91c1c;font-weight:800;'>الحساب معطل.</div>",
        )

    if not _verify_pin_hash(current_pin, account.get("pin_hash", "")):
        return (
            gr.update(value=""),
            gr.update(value=""),
            gr.update(value=""),
            "<div style='color:#b91c1c;font-weight:800;'>الرمز الحالي غير صحيح.</div>",
        )

    valid, validation_message = _validate_new_pin(new_pin)
    if not valid:
        return (
            gr.update(),
            gr.update(),
            gr.update(),
            f"<div style='color:#b91c1c;font-weight:800;'>{html_lib.escape(validation_message)}</div>",
        )

    if str(new_pin) != str(confirm_pin):
        return (
            gr.update(),
            gr.update(),
            gr.update(),
            "<div style='color:#b91c1c;font-weight:800;'>تأكيد الرمز الجديد غير مطابق.</div>",
        )

    if _verify_pin_hash(new_pin, account.get("pin_hash", "")):
        return (
            gr.update(),
            gr.update(),
            gr.update(),
            "<div style='color:#a16207;font-weight:800;'>الرمز الجديد مطابق للرمز الحالي.</div>",
        )

    if _pin_is_used_by_another_account(
        new_pin,
        exclude_account_id=account_id,
    ):
        return (
            gr.update(),
            gr.update(),
            gr.update(),
            "<div style='color:#b91c1c;font-weight:800;'>هذا الرمز مستخدم لحساب آخر.</div>",
        )

    account["pin_hash"] = _pin_hash(new_pin)
    account["must_change_pin"] = False
    account["updated_at"] = _auth_now_text()
    account["pin_changed_at"] = account["updated_at"]

    if not save_auth_accounts(payload):
        return (
            gr.update(),
            gr.update(),
            gr.update(),
            "<div style='color:#b91c1c;font-weight:800;'>تعذر حفظ الرمز الجديد.</div>",
        )

    write_audit_log(
        "تغيير رمز دخول",
        target_teacher="",
        old_value="رمز مشفر",
        new_value="رمز مشفر",
        details=f"غيّر المستخدم رمز حساب: {_account_display_name(account)}",
        actor_name=actor_name,
        actor_role=actor_role,
    )

    return (
        gr.update(value=""),
        gr.update(value=""),
        gr.update(value=""),
        (
            "<div style='color:#166534;background:#dcfce7;padding:10px;"
            "border-radius:8px;font-weight:800;'>"
            "تم تغيير رمز الدخول بنجاح. استخدم الرمز الجديد في الدخول القادم."
            "</div>"
        ),
    )

@state_locked
def owner_reset_account_pin(
    account_id,
    requested_pin,
    is_owner=False,
    actor_name="",
    actor_role="",
):
    if not bool(is_owner):
        return (
            gr.update(),
            gr.update(),
            gr.update(value=""),
            gr.update(value=""),
            "<div style='color:#b91c1c;font-weight:800;'>هذه العملية للمالك فقط.</div>",
        )

    account_id = str(account_id or "").strip()
    payload = load_auth_accounts()
    account = payload.get("accounts", {}).get(account_id)

    if not isinstance(account, dict):
        return (
            gr.update(),
            gr.update(),
            gr.update(value=""),
            gr.update(value=""),
            "<div style='color:#b91c1c;font-weight:800;'>اختر حسابًا صالحًا.</div>",
        )

    new_pin = str(requested_pin or "").strip()
    if not new_pin:
        new_pin = "".join(secrets.choice("0123456789") for _ in range(6))

    valid, validation_message = _validate_new_pin(new_pin)
    if not valid:
        return (
            gr.update(),
            gr.update(),
            gr.update(value=""),
            gr.update(),
            f"<div style='color:#b91c1c;font-weight:800;'>{html_lib.escape(validation_message)}</div>",
        )

    if _pin_is_used_by_another_account(
        new_pin,
        exclude_account_id=account_id,
    ):
        return (
            gr.update(),
            gr.update(),
            gr.update(value=""),
            gr.update(),
            "<div style='color:#b91c1c;font-weight:800;'>هذا الرمز مستخدم لحساب آخر.</div>",
        )

    account["pin_hash"] = _pin_hash(new_pin)
    account["must_change_pin"] = True
    account["updated_at"] = _auth_now_text()
    account["pin_reset_at"] = account["updated_at"]
    account["pin_reset_by"] = str(actor_name or "مالك النظام")

    if not save_auth_accounts(payload):
        return (
            gr.update(),
            gr.update(),
            gr.update(value=""),
            gr.update(),
            "<div style='color:#b91c1c;font-weight:800;'>تعذر حفظ إعادة التعيين.</div>",
        )

    write_audit_log(
        "إعادة تعيين رمز دخول",
        target_teacher="",
        old_value="رمز مشفر",
        new_value="رمز مؤقت مشفر",
        details=f"إعادة تعيين حساب: {_account_display_name(account)}",
        actor_name=actor_name,
        actor_role=actor_role,
    )

    choices = get_auth_account_choices()
    return (
        gr.update(value=render_auth_accounts_html(True)),
        gr.update(choices=choices, value=account_id),
        gr.update(value=""),
        gr.update(value=new_pin),
        (
            "<div style='color:#166534;background:#dcfce7;padding:10px;"
            "border-radius:8px;font-weight:800;'>"
            "تمت إعادة التعيين. يظهر الرمز الجديد في خانة «الرمز الجديد لمرة واحدة»."
            "</div>"
        ),
    )

@state_locked
def owner_toggle_account_status(
    account_id,
    is_owner=False,
    actor_name="",
    actor_role="",
):
    if not bool(is_owner):
        return (
            gr.update(),
            gr.update(),
            gr.update(value=""),
            "<div style='color:#b91c1c;font-weight:800;'>هذه العملية للمالك فقط.</div>",
        )

    account_id = str(account_id or "").strip()
    payload = load_auth_accounts()
    account = payload.get("accounts", {}).get(account_id)

    if not isinstance(account, dict):
        return (
            gr.update(),
            gr.update(),
            gr.update(value=""),
            "<div style='color:#b91c1c;font-weight:800;'>اختر حسابًا صالحًا.</div>",
        )

    new_enabled = not bool(account.get("enabled", True))
    account["enabled"] = new_enabled
    account["updated_at"] = _auth_now_text()

    if not save_auth_accounts(payload):
        return (
            gr.update(),
            gr.update(),
            gr.update(value=""),
            "<div style='color:#b91c1c;font-weight:800;'>تعذر تحديث حالة الحساب.</div>",
        )

    action_name = "تفعيل حساب دخول" if new_enabled else "تعطيل حساب دخول"
    write_audit_log(
        action_name,
        target_teacher="",
        old_value="معطل" if new_enabled else "مفعل",
        new_value="مفعل" if new_enabled else "معطل",
        details=f"{action_name}: {_account_display_name(account)}",
        actor_name=actor_name,
        actor_role=actor_role,
    )

    choices = get_auth_account_choices()
    status_word = "تفعيل" if new_enabled else "تعطيل"
    return (
        gr.update(value=render_auth_accounts_html(True)),
        gr.update(choices=choices, value=account_id),
        gr.update(value=""),
        (
            "<div style='color:#166534;background:#dcfce7;padding:10px;"
            "border-radius:8px;font-weight:800;'>"
            f"تم {status_word} الحساب بنجاح."
            "</div>"
        ),
    )


WELCOME_MESSAGES = {
    "صاحب النظام": " أهلاً بك يا صاحب النظام ({name}).. جميع الصلاحيات العليا أصبحت بين يديك.",
    "مستخدم عام": " أهلاً بكم في منظومة الباسط ونورتونا.. تم تجهيز الوصول إلى التبادل الودي الأسبوعي، جدول اليوم، وجدول المعلم الأسبوعي.",
    "مدير المدرسة": " أهلاً بك يا قائد المدرسة وربان سفينتها ({name}).. الرادار الإداري وغرفة العمليات رهن إشارتك.",
    "المدير المساعد": " أهلاً بالذراع الأيمن للقيادة والسند الإداري ({name}).. صلاحيات التدخل المفتوحة مفعلة.",
    "العلوم": " مرحباً بقائد الملحمة والمعلم الأول ({name}).. تم تجهيز شاشة قسم العلوم بدقة.",
    "الرياضيات": " نورتنا مهندس الأرقام ({name}) شاشة قسم الرياضيات جاهزة لك.",
    "التربية الإسلامية": " سُعدنا بانضمامك ({name}) شاشة قسم التربية الإسلامية جاهزة لك.",
    "اللغة العربية": " مايسترو البيان ({name}) نورتنا وقسم اللغة العربية جاهز لك.",
    "اللغة الإنجليزية": " أهلا بك سفير اللغة ({name}) شاشة قسم اللغة الإنجليزية جاهزة لك.",
    "الدراسات الإجتماعية": " مرحبا بك ({name}) قسم الدراسات الإجتماعية جاهز.",
    "المهارات الفردية": " سُعدنا بانضمامك ({name}) هذه مساحة للتنسيق وتنظيم العمل."
}



def refresh_admins_from_reference(dept_filter, is_owner=False):
    (
        message,
        abs_update,
        teacher_update_1,
        teacher_update_2,
        staff_names_update,
        balance_update,
        card_update,
        file_update,
    ) = refresh_admins_from_reference_core(dept_filter, is_owner=is_owner)
    return (
        message,
        gr.update(**abs_update),
        gr.update(**teacher_update_1),
        gr.update(**teacher_update_2),
        gr.update(**staff_names_update),
        gr.update(**balance_update),
        gr.update(**card_update),
        gr.update(**file_update),
    )
def refresh_phones_from_reference(dept_filter, is_owner=False):
    message, balance_update, card_update, file_update = refresh_phones_from_reference_core(
        dept_filter, is_owner=is_owner
    )
    return (
        message,
        gr.update(**balance_update),
        gr.update(**card_update),
        gr.update(**file_update),
    )
def refresh_schedule_from_reference(dept_name, current_day, is_owner=False):
    msg, abs_c, choices_all, balance_h, absences_h, day_ov, cards, _reset = refresh_schedule_from_reference_core(
        dept_name,
        current_day,
        is_owner,
    )
    return (
        msg,
        gr.update() if abs_c is None else gr.update(choices=abs_c),
        gr.update() if choices_all is None else gr.update(choices=choices_all, value=None),
        gr.update() if choices_all is None else gr.update(choices=choices_all, value=None),
        gr.update() if balance_h is None else gr.update(value=balance_h),
        gr.update() if absences_h is None else gr.update(value=absences_h),
        gr.update() if day_ov is None else gr.update(value=day_ov),
        gr.update(value=cards),
        gr.update(value=None),
    )

    


    
if font_path:
    print(f"font ok: {font_path}")
    arabic_font = fm.FontProperties(fname=font_path)
else:
    print("font warning: Cairo-Regular.ttf not found")
    arabic_font = fm.FontProperties()

if image_font_path:
    print(f"image font ok: {image_font_path}")
    image_font = fm.FontProperties(fname=image_font_path)
else:
    print("image font warning: Amiri-Regular.ttf not found; falling back to Cairo/default")
    image_font = arabic_font
reshaper_config = {'support_ligatures': False}
reshaper = arabic_reshaper.ArabicReshaper(configuration=reshaper_config)

def fix_arabic(text):
    reshaped = reshaper.reshape(str(text))
    bidi = get_display(reshaped)
    for c in ['\u202a', '\u202b', '\u202c', '\u200e', '\u200f']: bidi = bidi.replace(c, '')
    return bidi



# ─────────────────────────────────────────────────────────────────────────────
# v1.7 — لوحة سجل العمليات والنسخ الاحتياطية (مالك النظام فقط)
# ─────────────────────────────────────────────────────────────────────────────

def load_audit_log_records():
    """قراءة سجل العمليات الحساسة بصورة آمنة وإرجاع قائمة مرتبة من الأحدث."""
    lock = _get_json_file_lock(AUDIT_LOG_FILE)
    with lock:
        if not os.path.exists(AUDIT_LOG_FILE):
            return []
        try:
            with open(AUDIT_LOG_FILE, "r", encoding="utf-8") as f:
                loaded = json.load(f)
            if not isinstance(loaded, list):
                return []
            records = [item for item in loaded if isinstance(item, dict)]
            records.sort(key=lambda item: str(item.get("timestamp", "")), reverse=True)
            return records
        except Exception as e:
            print(f"load_audit_log_records error: {e}")
            return []


def _audit_value_to_text(value):
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        try:
            return json.dumps(value, ensure_ascii=False, sort_keys=True)
        except Exception:
            return str(value)
    return str(value)


def _parse_audit_date(value, label):
    if value is None or value == "":
        return None, ""

    if isinstance(value, datetime.datetime):
        return value.date(), ""

    if isinstance(value, datetime.date):
        return value, ""

    if isinstance(value, (int, float)):
        try:
            return datetime.datetime.fromtimestamp(
                float(value),
                tz=tz_oman,
            ).date(), ""
        except Exception:
            return None, f"قيمة {label} غير صالحة."

    raw = str(value or "").strip()
    if not raw:
        return None, ""

    # يدعم YYYY-MM-DD وكذلك قيم ISO التي قد يعيدها مكوّن Gradio DateTime.
    date_part = raw[:10]
    try:
        return datetime.datetime.strptime(date_part, "%Y-%m-%d").date(), ""
    except Exception:
        return None, f"صيغة {label} غير صحيحة. اختر التاريخ من التقويم."


def get_audit_date_range(preset):
    """إرجاع نطاق تاريخ جاهز وفق توقيت سلطنة عمان."""
    today = datetime.datetime.now(tz_oman).date()
    preset_value = str(preset or "").strip()

    if preset_value == "today":
        start_date = today
        end_date = today
    elif preset_value == "last_7_days":
        start_date = today - datetime.timedelta(days=6)
        end_date = today
    elif preset_value == "this_month":
        start_date = today.replace(day=1)
        end_date = today
    elif preset_value == "clear":
        return None, None
    else:
        return None, None

    return start_date.isoformat(), end_date.isoformat()


def filter_audit_records(records, action_filter="الكل", actor_filter="الكل", teacher_filter="الكل", date_from="", date_to=""):
    start_date, start_error = _parse_audit_date(date_from, "تاريخ البداية")
    end_date, end_error = _parse_audit_date(date_to, "تاريخ النهاية")
    error = start_error or end_error
    if error:
        return [], error
    if start_date and end_date and start_date > end_date:
        return [], "تاريخ البداية يجب ألا يكون بعد تاريخ النهاية."

    action_filter = str(action_filter or "الكل").strip()
    actor_filter = str(actor_filter or "الكل").strip()
    teacher_filter = str(teacher_filter or "الكل").strip()

    filtered = []
    for record in records:
        if action_filter != "الكل" and str(record.get("action", "")).strip() != action_filter:
            continue
        if actor_filter != "الكل" and str(record.get("actor_name", "")).strip() != actor_filter:
            continue
        if teacher_filter != "الكل" and str(record.get("target_teacher", "")).strip() != teacher_filter:
            continue

        timestamp = str(record.get("timestamp", "")).strip()
        record_date = None
        try:
            record_date = datetime.datetime.strptime(timestamp[:10], "%Y-%m-%d").date()
        except Exception:
            pass

        if start_date and (record_date is None or record_date < start_date):
            continue
        if end_date and (record_date is None or record_date > end_date):
            continue
        filtered.append(record)

    return filtered, ""


def _audit_filter_choices(records):
    actions = sorted({str(r.get("action", "")).strip() for r in records if str(r.get("action", "")).strip()})
    actors = sorted({str(r.get("actor_name", "")).strip() for r in records if str(r.get("actor_name", "")).strip()})
    teachers = sorted({str(r.get("target_teacher", "")).strip() for r in records if str(r.get("target_teacher", "")).strip()})
    return ["الكل"] + actions, ["الكل"] + actors, ["الكل"] + teachers


def render_audit_log_html(records, total_records=0, error_message=""):
    if error_message:
        return f"<div style='background:#ffebee;color:#b91c1c;border-right:5px solid #c62828;padding:12px;border-radius:10px;font-weight:800;'>{html_lib.escape(error_message)}</div>"

    if not records:
        return "<div style='background:#f8fafc;border:1px dashed #cbd5e1;border-radius:10px;padding:18px;text-align:center;color:#64748b;'>لا توجد عمليات مطابقة للعرض.</div>"

    display_records = records[:200]
    rows = []
    for record in display_records:
        values = {
            "timestamp": str(record.get("timestamp", "")),
            "actor_name": str(record.get("actor_name", "")),
            "actor_role": str(record.get("actor_role", "")),
            "action": str(record.get("action", "")),
            "target_teacher": str(record.get("target_teacher", "")),
            "old_value": _audit_value_to_text(record.get("old_value")),
            "new_value": _audit_value_to_text(record.get("new_value")),
            "details": str(record.get("details", "")),
        }
        safe = {key: html_lib.escape(value) for key, value in values.items()}
        rows.append(f"""
        <tr>
            <td>{safe['timestamp']}</td>
            <td>{safe['actor_name']}</td>
            <td>{safe['actor_role']}</td>
            <td><b>{safe['action']}</b></td>
            <td>{safe['target_teacher']}</td>
            <td class='audit-wide'>{safe['old_value']}</td>
            <td class='audit-wide'>{safe['new_value']}</td>
            <td class='audit-wide'>{safe['details']}</td>
        </tr>
        """)

    shown_note = f"يعرض آخر {len(display_records)} سجل من أصل {len(records)} سجل مطابق" if len(records) > 200 else f"عدد السجلات المطابقة: {len(records)}"
    return f"""
    <div style='direction:rtl;'>
        <div style='background:#e8f5e9;color:#004d40;border-right:5px solid #2e7d32;padding:11px 14px;border-radius:10px;margin-bottom:10px;font-weight:800;'>
            {shown_note} | إجمالي السجلات المحفوظة: {int(total_records)}
        </div>
        <div style='overflow-x:auto;border:1px solid #dbe3e8;border-radius:12px;'>
            <table style='width:100%;min-width:1450px;border-collapse:collapse;text-align:center;font-family:Cairo,Arial,sans-serif;font-size:13px;'>
                <thead>
                    <tr style='background:#004d40;color:white;'>
                        <th>التاريخ والوقت</th><th>المنفذ</th><th>الدور</th><th>العملية</th>
                        <th>المعلم المتأثر</th><th>القيمة القديمة</th><th>القيمة الجديدة</th><th>التفاصيل</th>
                    </tr>
                </thead>
                <tbody>{''.join(rows)}</tbody>
            </table>
        </div>
    </div>
    """


def refresh_audit_dashboard(action_filter, actor_filter, teacher_filter, date_from, date_to, is_owner=False):
    if not bool(is_owner):
        denied = "<div style='color:#b91c1c;background:#ffebee;padding:12px;border-radius:10px;font-weight:800;'>هذه الأداة متاحة لمالك النظام فقط.</div>"
        return (
            gr.update(choices=["الكل"], value="الكل"),
            gr.update(choices=["الكل"], value="الكل"),
            gr.update(choices=["الكل"], value="الكل"),
            gr.update(value=denied),
        )

    records = load_audit_log_records()
    action_choices, actor_choices, teacher_choices = _audit_filter_choices(records)

    action_value = action_filter if action_filter in action_choices else "الكل"
    actor_value = actor_filter if actor_filter in actor_choices else "الكل"
    teacher_value = teacher_filter if teacher_filter in teacher_choices else "الكل"

    filtered, error = filter_audit_records(records, action_value, actor_value, teacher_value, date_from, date_to)
    return (
        gr.update(choices=action_choices, value=action_value),
        gr.update(choices=actor_choices, value=actor_value),
        gr.update(choices=teacher_choices, value=teacher_value),
        gr.update(value=render_audit_log_html(filtered, len(records), error)),
    )


def export_audit_log_excel(action_filter, actor_filter, teacher_filter, date_from, date_to, is_owner=False):
    if not bool(is_owner):
        return (
            gr.update(value=None),
            "<div style='color:#b91c1c;font-weight:800;'>هذه العملية متاحة لمالك النظام فقط.</div>",
        )

    try:
        records = load_audit_log_records()
        filtered, error = filter_audit_records(
            records,
            action_filter,
            actor_filter,
            teacher_filter,
            date_from,
            date_to,
        )

        if error:
            return (
                gr.update(value=None),
                f"<div style='color:#b91c1c;font-weight:800;'>{html_lib.escape(error)}</div>",
            )

        if not filtered:
            return (
                gr.update(value=None),
                "<div style='color:#a16207;font-weight:800;'>لا توجد سجلات مطابقة لتصديرها.</div>",
            )

        ensure_data_directories()
        rows = []
        for record in filtered:
            rows.append({
                "التاريخ والوقت": str(record.get("timestamp", "")),
                "اسم المنفذ": str(record.get("actor_name", "")),
                "دور المنفذ": str(record.get("actor_role", "")),
                "نوع العملية": str(record.get("action", "")),
                "المعلم المتأثر": str(record.get("target_teacher", "")),
                "القيمة القديمة": _audit_value_to_text(record.get("old_value")),
                "القيمة الجديدة": _audit_value_to_text(record.get("new_value")),
                "التفاصيل": str(record.get("details", "")),
                "المصدر": str(record.get("source", "")),
            })

        filename = os.path.join(
            EXPORTS_DIR,
            f"سجل_العمليات_الحساسة_{get_now_oman().strftime('%Y%m%d_%H%M%S_%f')}.xlsx",
        )

        df = pd.DataFrame(rows)
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="سجل العمليات")
            ws = writer.sheets["سجل العمليات"]

            header_fill = PatternFill(fill_type="solid", fgColor="004D40")
            header_font = Font(color="FFFFFF", bold=True)
            center = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = center

            for column_cells in ws.columns:
                max_length = max(
                    len(str(cell.value or ""))
                    for cell in column_cells
                )
                ws.column_dimensions[
                    column_cells[0].column_letter
                ].width = min(max(max_length + 3, 14), 55)

            ws.freeze_panes = "A2"
            ws.sheet_view.rightToLeft = True

        absolute_filename = os.path.abspath(filename)
        if not os.path.isfile(absolute_filename):
            raise FileNotFoundError(
                "تم إنشاء طلب التصدير لكن ملف Excel غير موجود في مسار المخرجات."
            )

        return (
            gr.update(value=absolute_filename),
            (
                "<div style='color:#166534;background:#dcfce7;padding:10px;"
                "border-radius:8px;font-weight:800;'>"
                f"تم تجهيز ملف Excel ويحتوي على {len(filtered)} سجل."
                "</div>"
            ),
        )

    except Exception as exc:
        print(f"export_audit_log_excel error: {exc}")
        return (
            gr.update(value=None),
            (
                "<div style='color:#b91c1c;background:#fee2e2;padding:10px;"
                "border-radius:8px;font-weight:800;'>"
                "تعذر تصدير السجل. التفاصيل التقنية: "
                f"{html_lib.escape(str(exc))}"
                "</div>"
            ),
        )


def _backup_files_for_target(file_path):
    ensure_data_directories()
    base = os.path.basename(str(file_path))
    stem, ext = os.path.splitext(base)
    if not ext:
        ext = ".json"
    prefix = f"{stem}_"
    candidates = []
    for name in os.listdir(BACKUPS_DIR):
        full_path = os.path.join(BACKUPS_DIR, name)
        if os.path.isfile(full_path) and name.startswith(prefix) and name.endswith(ext):
            candidates.append(full_path)
    candidates.sort(key=lambda path: os.path.getmtime(path), reverse=True)
    return candidates


def _format_file_size(size_bytes):
    try:
        size = float(size_bytes)
    except Exception:
        return "—"
    if size < 1024:
        return f"{int(size)} B"
    if size < 1024 * 1024:
        return f"{size / 1024:.1f} KB"
    return f"{size / (1024 * 1024):.2f} MB"


def get_monitored_backup_files():
    return [
        ("قاعدة بيانات المعلمين والأرصدة", DB_FILE),
        ("التوزيع اليومي", DAILY_DB_FILE),
        ("التبادلات الودية", SWAP_DB_FILE),
        ("سجل الإعفاءات", EXEMPTIONS_LOG_FILE),
        ("سجل العمليات الحساسة", AUDIT_LOG_FILE),
        ("إعدادات المدرسة", SCHOOL_CONFIG_FILE),
        ("حسابات الدخول المشفرة", AUTH_ACCOUNTS_FILE),
    ]


def render_backup_status_html(is_owner=False):
    if not bool(is_owner):
        return "<div style='color:#b91c1c;background:#ffebee;padding:12px;border-radius:10px;font-weight:800;'>هذه الأداة متاحة لمالك النظام فقط.</div>"

    rows = []
    total_backups = 0
    for label, source_path in get_monitored_backup_files():
        backups = _backup_files_for_target(source_path)
        total_backups += len(backups)
        source_exists = os.path.exists(source_path)
        latest = backups[0] if backups else None
        latest_name = os.path.basename(latest) if latest else "—"
        latest_at = datetime.datetime.fromtimestamp(os.path.getmtime(latest), tz=tz_oman).strftime("%Y-%m-%d %H:%M:%S") if latest else "—"
        latest_size = _format_file_size(os.path.getsize(latest)) if latest else "—"
        source_status = "موجود" if source_exists else "غير موجود"
        source_color = "#166534" if source_exists else "#b91c1c"
        rows.append(f"""
        <tr>
            <td><b>{html_lib.escape(label)}</b></td>
            <td style='color:{source_color};font-weight:800;'>{source_status}</td>
            <td>{len(backups)}</td>
            <td>{html_lib.escape(latest_name)}</td>
            <td>{latest_at}</td>
            <td>{latest_size}</td>
        </tr>
        """)

    if PERSISTENT_STORAGE_ACTIVE:
        storage_note = (
            "<div style='margin-top:10px;background:#dcfce7;color:#166534;padding:10px;"
            "border-radius:8px;border-right:4px solid #16a34a;'>"
            "✅ النسخ محفوظة داخل التخزين الدائم: "
            + html_lib.escape(DATA_DIR)
            + "</div>"
        )
    else:
        storage_note = (
            "<div style='margin-top:10px;background:#fff7ed;color:#9a3412;padding:10px;"
            "border-radius:8px;border-right:4px solid #ea580c;'>"
            "⚠️ التطبيق يعمل على التخزين المحلي الاحتياطي؛ تحقّق من تركيب الـBucket."
            "</div>"
        )

    return f"""
    <div style='direction:rtl;'>
        <div style='background:#e0f2fe;color:#0c4a6e;border-right:5px solid #0284c7;padding:11px 14px;border-radius:10px;margin-bottom:10px;font-weight:800;'>
            إجمالي النسخ الاحتياطية المحلية: {total_backups}
        </div>
        <div style='overflow-x:auto;border:1px solid #dbe3e8;border-radius:12px;'>
            <table style='width:100%;min-width:1050px;border-collapse:collapse;text-align:center;font-family:Cairo,Arial,sans-serif;font-size:13px;'>
                <thead><tr style='background:#0f766e;color:white;'>
                    <th>الملف</th><th>حالة الأصل</th><th>عدد النسخ</th><th>أحدث نسخة</th><th>تاريخها</th><th>حجمها</th>
                </tr></thead>
                <tbody>{''.join(rows)}</tbody>
            </table>
        </div>
        {storage_note}
    </div>
    """


def refresh_backup_status(is_owner=False):
    return gr.update(value=render_backup_status_html(is_owner))


def _latest_backup_for(file_path):
    candidates = _backup_files_for_target(file_path)
    return candidates[0] if candidates else None


def create_backup_bundle(is_owner=False):
    if not bool(is_owner):
        return gr.update(value=None), "<div style='color:#b91c1c;font-weight:800;'>هذه العملية متاحة لمالك النظام فقط.</div>"

    ensure_data_directories()
    filename = os.path.join(EXPORTS_DIR, f"نسخة_احتياطية_منظومة_مسار_{get_now_oman().strftime('%Y%m%d_%H%M%S_%f')}.zip")
    manifest = {
        "created_at": get_now_oman().strftime("%Y-%m-%d %H:%M:%S"),
        "school_name": SCHOOL_NAME,
        "system_name": SYSTEM_NAME,
        "contents": [],
    }

    current_files = [path for _label, path in get_monitored_backup_files()]
    reference_files = [
        ADMIN_FILE,
        PHONES_FILE,
        REFERENCE_STATUS_FILE,
    ] + list(SCHEDULE_FILES.values())

    with STATE_LOCK:
        try:
            with zipfile.ZipFile(filename, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                for path in current_files:
                    if os.path.exists(path) and os.path.isfile(path):
                        arcname = os.path.join("current", os.path.basename(path))
                        zf.write(path, arcname=arcname)
                        manifest["contents"].append(arcname)

                    latest_backup = _latest_backup_for(path)
                    if latest_backup:
                        arcname = os.path.join("latest_backups", os.path.basename(latest_backup))
                        zf.write(latest_backup, arcname=arcname)
                        manifest["contents"].append(arcname)

                for path in reference_files:
                    if os.path.exists(path) and os.path.isfile(path):
                        arcname = os.path.join("reference_files", os.path.basename(path))
                        zf.write(path, arcname=arcname)
                        manifest["contents"].append(arcname)

                zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))

            return gr.update(value=os.path.abspath(filename)), "<div style='color:#166534;background:#dcfce7;padding:10px;border-radius:8px;font-weight:800;'>تم تجهيز النسخة الاحتياطية المضغوطة بنجاح.</div>"
        except Exception as e:
            print(f"create_backup_bundle error: {e}")
            try:
                if os.path.exists(filename):
                    os.remove(filename)
            except Exception:
                pass
            return gr.update(value=None), f"<div style='color:#b91c1c;font-weight:800;'>تعذر تجهيز النسخة الاحتياطية: {html_lib.escape(str(e))}</div>"


def refresh_owner_tools_dashboard(action_filter, actor_filter, teacher_filter, date_from, date_to, is_owner=False):
    action_upd, actor_upd, teacher_upd, audit_upd = refresh_audit_dashboard(
        action_filter, actor_filter, teacher_filter, date_from, date_to, is_owner
    )
    backup_upd = refresh_backup_status(is_owner)
    return action_upd, actor_upd, teacher_upd, audit_upd, backup_upd


def refresh_school_data_center_cards(is_owner=False):
    """
    إعادة قراءة حالة التخزين والملفات المرجعية عند كل دخول إلى مركز البيانات.
    يمنع رجوع البطاقات إلى القيم الحمراء الابتدائية بعد الخروج أو تحديث الصفحة.
    """
    if not bool(is_owner):
        denied = gr.update()
        return denied, denied, denied, denied, denied

    return (
        gr.update(value=render_persistent_storage_status_html()),
        gr.update(value=render_school_config_summary_html()),
        gr.update(value=render_admin_reference_card()),
        gr.update(value=render_phones_reference_card()),
        gr.update(value=render_schedule_reference_cards()),
    )


# ─────────────────────────────────────────────────────────────────────────────
# v1.8 — إعداد هوية المدرسة من الواجهة (مالك النظام فقط)
# ─────────────────────────────────────────────────────────────────────────────

IDENTITY_CONFIG_KEYS = (
    "school_name",
    "directorate_region",
    "logo_url",
    "theme_color",
    "theme_color_2",
    "accent_color",
)

FIXED_IDENTITY_KEYS = (
    "ministry_name",
    "directorate_prefix",
    "system_name",
    "system_subtitle",
    "developer_credit",
)

def _normalize_identity_text(value, fallback="", max_length=220):
    cleaned = re.sub(r"\s+", " ", str(value or "").strip())
    if not cleaned:
        cleaned = str(fallback or "").strip()
    return cleaned[:max_length]

def _normalize_hex_color(value, fallback):
    raw = str(value or "").strip()
    if re.fullmatch(r"#[0-9a-fA-F]{6}", raw):
        return raw.lower()
    fallback_raw = str(fallback or "#004d40").strip()
    return fallback_raw.lower() if re.fullmatch(r"#[0-9a-fA-F]{6}", fallback_raw) else "#004d40"

def _is_valid_identity_logo_value(value):
    raw = str(value or "").strip()
    if not raw:
        return False
    if raw.startswith("data:image/"):
        return True
    parsed = urllib.parse.urlparse(raw)
    if parsed.scheme in {"http", "https"} and parsed.netloc:
        return True
    path = os.path.abspath(raw)
    return os.path.isfile(path)

def _resolve_identity_logo_source(value=None):
    raw = str(value if value is not None else SCHOOL_LOGO_URL).strip()
    if not raw:
        raw = str(DEFAULT_SCHOOL_CONFIG["logo_url"])

    if raw.startswith(("http://", "https://", "data:image/")):
        return raw

    candidate = os.path.abspath(raw)
    if not os.path.isfile(candidate):
        return str(DEFAULT_SCHOOL_CONFIG["logo_url"])

    mime_type = mimetypes.guess_type(candidate)[0] or "image/png"
    try:
        encoded = base64.b64encode(Path(candidate).read_bytes()).decode("ascii")
        return f"data:{mime_type};base64,{encoded}"
    except Exception:
        return str(DEFAULT_SCHOOL_CONFIG["logo_url"])

def _save_uploaded_identity_logo(uploaded_file):
    if uploaded_file is None:
        return None

    source_path = getattr(uploaded_file, "name", uploaded_file)
    source_path = str(source_path or "").strip()
    if not source_path or not os.path.isfile(source_path):
        raise ValueError("ملف الشعار المرفوع غير صالح.")

    try:
        with Image.open(source_path) as image:
            image.verify()
    except Exception as exc:
        raise ValueError("الملف المرفوع ليس صورة صالحة.") from exc

    ext = Path(source_path).suffix.lower()
    if ext not in {".png", ".jpg", ".jpeg", ".webp"}:
        ext = ".png"

    ensure_data_directories()
    destination = os.path.join(BRANDING_DIR, f"school_logo{ext}")
    shutil.copy2(source_path, destination)
    return os.path.relpath(destination, os.getcwd())

def _identity_directorate_full_name(region=None):
    region_clean = _normalize_identity_text(
        region,
        DEFAULT_SCHOOL_CONFIG["directorate_region"],
        80,
    )
    return f"{DEFAULT_SCHOOL_CONFIG['directorate_prefix']} {region_clean}".strip()

def _apply_school_identity_globals(config):
    global SCHOOL_CONFIG
    global MINISTRY_NAME, DIRECTORATE_PREFIX, DIRECTORATE_REGION, DIRECTORATE_FULL_NAME
    global SYSTEM_NAME, SYSTEM_SUBTITLE, SCHOOL_NAME, DEVELOPER_CREDIT
    global SCHOOL_LOGO_URL, THEME_COLOR, THEME_COLOR_2, ACCENT_COLOR

    SCHOOL_CONFIG = dict(config)

    MINISTRY_NAME = str(DEFAULT_SCHOOL_CONFIG["ministry_name"])
    DIRECTORATE_PREFIX = str(DEFAULT_SCHOOL_CONFIG["directorate_prefix"])
    SYSTEM_NAME = str(DEFAULT_SCHOOL_CONFIG["system_name"])
    SYSTEM_SUBTITLE = str(DEFAULT_SCHOOL_CONFIG["system_subtitle"])
    DEVELOPER_CREDIT = str(DEFAULT_SCHOOL_CONFIG["developer_credit"])

    DIRECTORATE_REGION = _normalize_identity_text(
        SCHOOL_CONFIG.get("directorate_region"),
        DEFAULT_SCHOOL_CONFIG["directorate_region"],
        80,
    )
    DIRECTORATE_FULL_NAME = _identity_directorate_full_name(DIRECTORATE_REGION)

    SCHOOL_NAME = _normalize_identity_text(
        SCHOOL_CONFIG.get("school_name"),
        DEFAULT_SCHOOL_CONFIG["school_name"],
        140,
    )
    SCHOOL_LOGO_URL = str(
        SCHOOL_CONFIG.get("logo_url") or DEFAULT_SCHOOL_CONFIG["logo_url"]
    ).strip()
    THEME_COLOR = _normalize_hex_color(
        SCHOOL_CONFIG.get("theme_color"),
        DEFAULT_SCHOOL_CONFIG["theme_color"],
    )
    THEME_COLOR_2 = _normalize_hex_color(
        SCHOOL_CONFIG.get("theme_color_2"),
        DEFAULT_SCHOOL_CONFIG["theme_color_2"],
    )
    ACCENT_COLOR = _normalize_hex_color(
        SCHOOL_CONFIG.get("accent_color"),
        DEFAULT_SCHOOL_CONFIG["accent_color"],
    )

def _current_identity_config():
    return {
        "ministry_name": MINISTRY_NAME,
        "directorate_prefix": DIRECTORATE_PREFIX,
        "directorate_region": DIRECTORATE_REGION,
        "directorate_full_name": DIRECTORATE_FULL_NAME,
        "system_name": SYSTEM_NAME,
        "system_subtitle": SYSTEM_SUBTITLE,
        "school_name": SCHOOL_NAME,
        "developer_credit": DEVELOPER_CREDIT,
        "logo_url": SCHOOL_LOGO_URL,
        "theme_color": THEME_COLOR,
        "theme_color_2": THEME_COLOR_2,
        "accent_color": ACCENT_COLOR,
    }

def build_login_branding_html(config=None):
    cfg = dict(_current_identity_config())
    if isinstance(config, dict):
        cfg.update(config)

    system_name = html_lib.escape(str(cfg.get("system_name", SYSTEM_NAME)))
    school_name = html_lib.escape(str(cfg.get("school_name", SCHOOL_NAME)))
    logo_src = html_lib.escape(
        _resolve_identity_logo_source(cfg.get("logo_url")),
        quote=True,
    )
    theme = _normalize_hex_color(cfg.get("theme_color"), THEME_COLOR)
    theme_2 = _normalize_hex_color(cfg.get("theme_color_2"), THEME_COLOR_2)
    accent = _normalize_hex_color(cfg.get("accent_color"), ACCENT_COLOR)

    return f"""
<div style="
    background:linear-gradient(145deg,#003d33 0%,{theme} 40%,{theme_2} 80%,{theme} 100%);
    margin:0px 0px 20px 0px;
    padding:30px 20px 25px;
    padding-bottom:0 !important;
    overflow:hidden;
    border-radius:16px 16px 0 0;
    text-align:center;
    border-bottom:none;
">
    <img id="main-logo" src="{logo_src}" alt="{system_name}" style="
        width:115px;height:115px;
        border-radius:50%;
        border:3px solid {accent};
        background:white;
        padding:3px;
        display:inline-block;
        margin-bottom:14px;
        object-fit:contain;
        box-shadow:
            0 15px 40px rgba(0,0,0,0.6),
            0 6px 15px rgba(0,0,0,0.4),
            0 0 0 5px rgba(255,202,40,0.3),
            0 0 0 10px rgba(0,77,64,0.15),
            4px -4px 15px rgba(255,255,255,0.2),
            -4px 4px 15px rgba(0,0,0,0.3);
        animation:logo4d 4s ease-in-out infinite;
        cursor:pointer;
    ">
    <div style="font-size:26px;font-weight:900;color:{accent};text-shadow:0 2px 8px rgba(0,0,0,0.4);margin-bottom:6px;">
        بوابة الدخول
    </div>
    <div style="font-size:13px;color:rgba(255,255,255,0.92);font-weight:700;">
        {school_name}
    </div>
    <div style="font-size:11px;color:rgba(255,255,255,0.78);font-weight:600;margin-top:4px;">
        {system_name}
    </div>

    <div class="masar-login-wave-safe" style="--masar-wave-gold:{accent};--masar-wave-green:{theme};">
        <div class="masar-login-wave-safe-track">
            <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 1440 60" preserveAspectRatio="none" aria-hidden="true">
                <path fill="{theme}" fill-opacity="1" d="M0,0 L0,34 C120,8 240,58 360,34 C480,8 600,58 720,34 C840,8 960,58 1080,34 C1200,8 1320,58 1440,34 L1440,0 Z"/>
            </svg>
            <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 1440 60" preserveAspectRatio="none" aria-hidden="true">
                <path fill="{theme}" fill-opacity="1" d="M0,0 L0,34 C120,8 240,58 360,34 C480,8 600,58 720,34 C840,8 960,58 1080,34 C1200,8 1320,58 1440,34 L1440,0 Z"/>
            </svg>
        </div>
    </div>
</div>
"""

def build_login_credits_html(config=None):
    cfg = dict(_current_identity_config())
    if isinstance(config, dict):
        cfg.update(config)
    credit = html_lib.escape(str(cfg.get("developer_credit", DEVELOPER_CREDIT)))
    return (
        "<div style='text-align:center;'>"
        f"<div class='credits-box' style='font-size:10px;padding:5px 10px;'>{credit}</div>"
        "</div>"
    )

def build_header_html(config=None):
    cfg = dict(_current_identity_config())
    if isinstance(config, dict):
        cfg.update(config)

    system_name = html_lib.escape(str(DEFAULT_SCHOOL_CONFIG["system_name"]))
    subtitle = html_lib.escape(str(DEFAULT_SCHOOL_CONFIG["system_subtitle"]))
    school_name = html_lib.escape(str(cfg.get("school_name", SCHOOL_NAME)))
    credit = html_lib.escape(str(DEFAULT_SCHOOL_CONFIG["developer_credit"]))
    ministry = html_lib.escape(str(DEFAULT_SCHOOL_CONFIG["ministry_name"]))
    directorate = html_lib.escape(
        _identity_directorate_full_name(
            cfg.get("directorate_region", DIRECTORATE_REGION)
        )
    )
    logo_src = html_lib.escape(
        _resolve_identity_logo_source(cfg.get("logo_url")),
        quote=True,
    )
    accent = _normalize_hex_color(cfg.get("accent_color"), ACCENT_COLOR)

    return f"""
<div class='main-header'>
    <div class='header-grid'>
        <div class='h-logo'><img src='{logo_src}' alt='Logo'></div>
        <div class='h-ministry'>{ministry}<br>{directorate}</div>
        <div class='h-title'>
            <div class='h-title-main'>{system_name}</div>
            <div class='h-title-sub'>{subtitle}</div>
        </div>
        <div class='h-school' style='color:{accent} !important;-webkit-text-fill-color:{accent} !important;white-space:nowrap;'>{school_name}</div>
        <div class='h-credits'><div class='credits-box'>{credit}</div></div>
    </div>
</div>
"""

def _home_hero_account_note(record):
    context = _account_profile_context(record)
    role = context.get("role", "")
    official = context.get("official_title", "")
    department_label = context.get("department_label", "")

    if bool(record.get("is_owner", False)) or role == OWNER_ROLE:
        return "لوحة التحكم العليا جاهزة لإدارة المنظومة."

    if official and department_label:
        return f"{official} — {department_label}"

    if official:
        return f"صلاحية الدخول: {official}"

    if department_label:
        return f"{department_label} جاهز لك."

    return "تم تجهيز لوحة العمل حسب صلاحيتك."


def _resolve_home_hero_account_record(
    account_id="",
    user_name="",
    user_role="",
    is_owner=False,
):
    account_id = str(account_id or "").strip()

    if bool(is_owner) or account_id == OWNER_ACCOUNT_ID:
        owner_name = (
            str(user_name or "").strip()
            or os.getenv("SYSTEM_OWNER_NAME", "صاحب النظام").strip()
            or "صاحب النظام"
        )
        return {
            "account_id": OWNER_ACCOUNT_ID,
            "name": owner_name,
            "display_name": owner_name,
            "role": OWNER_ROLE,
            "official_title": OWNER_ROLE,
            "whatsapp_title": OWNER_ROLE,
            "department_label": "غرفة القيادة",
            "welcome_title": "صاحب النظام",
            "welcome_phrase": "لوحة التحكم العليا جاهزة لإدارة المنظومة",
            "welcome_template": "مرحبًا بك يا {welcome_title} ({display_name})",
            "is_owner": True,
        }

    if account_id:
        payload = load_auth_accounts()
        record = payload.get("accounts", {}).get(account_id)
        if isinstance(record, dict):
            return dict(record)

    fallback_name = str(user_name or "").strip()
    fallback_role = str(user_role or "").strip()
    if fallback_name or fallback_role:
        return {
            "account_id": account_id,
            "name": fallback_name,
            "display_name": fallback_name,
            "role": fallback_role,
            "official_title": fallback_role,
            "whatsapp_title": fallback_role,
            "department_label": "",
            "welcome_title": "",
            "welcome_phrase": "",
            "welcome_template": "",
            "is_owner": False,
        }

    return {}


def build_home_hero_html(config=None, account_record=None):
    cfg = dict(_current_identity_config())
    if isinstance(config, dict):
        cfg.update(config)

    system_name = html_lib.escape(str(cfg.get("system_name", SYSTEM_NAME)))
    school_name = html_lib.escape(str(cfg.get("school_name", SCHOOL_NAME)))

    if isinstance(account_record, dict) and account_record:
        title_text = build_account_welcome_text(account_record)
        note_text = _home_hero_account_note(account_record)
    else:
        title_text = f"مرحبًا بك في {system_name}"
        note_text = "تم تجهيز لوحة العمل حسب صلاحيتك. اختر القسم المناسب للبدء."

    title_html = html_lib.escape(str(title_text or f"مرحبًا بك في {system_name}"))
    note_html = html_lib.escape(str(note_text or ""))

    return f"""
<div class='masar-home-hero'>
    <div class='masar-home-title'>{title_html}</div>
    <div class='masar-home-subtitle'>{school_name}</div>
    <div class='masar-home-note'>{note_html}</div>
</div>
"""


def update_home_hero_after_login(
    account_id,
    user_name,
    user_role,
    is_owner=False,
):
    record = _resolve_home_hero_account_record(
        account_id=account_id,
        user_name=user_name,
        user_role=user_role,
        is_owner=is_owner,
    )
    return gr.update(value=build_home_hero_html(account_record=record))


def _coerce_periods_per_day(value, default=None):
    try:
        parsed = int(str(value).strip())
    except Exception:
        parsed = default if default is not None else int(DEFAULT_SCHOOL_CONFIG["periods_per_day"])
    if parsed not in (7, 8):
        parsed = default if default in (7, 8) else int(DEFAULT_SCHOOL_CONFIG["periods_per_day"])
    return int(parsed)


def get_config_periods_per_day(config=None):
    cfg = config if isinstance(config, dict) else load_school_config()
    return _coerce_periods_per_day(
        cfg.get("periods_per_day", DEFAULT_SCHOOL_CONFIG["periods_per_day"]),
        int(DEFAULT_SCHOOL_CONFIG["periods_per_day"]),
    )


def render_operational_settings_status_html(config=None):
    cfg = config if isinstance(config, dict) else load_school_config()
    saved_periods = get_config_periods_per_day(cfg)
    running_periods = int(MAX_PERIODS)
    if saved_periods == running_periods:
        status_line = "الإعداد المحفوظ مطابق لوضع التشغيل الحالي."
        status_bg = "#dcfce7"
        status_fg = "#166534"
        status_border = "#16a34a"
    else:
        status_line = "تم حفظ إعداد مختلف عن وضع التشغيل الحالي. يلزم إعادة تشغيل المنظومة لتطبيقه."
        status_bg = "#fff7ed"
        status_fg = "#9a3412"
        status_border = "#f59e0b"

    return f"""
<div style='background:{status_bg};color:{status_fg};padding:12px;border-radius:10px;border-right:5px solid {status_border};font-weight:800;line-height:1.9;text-align:right;margin-bottom:12px;'>
    <div>عدد الحصص المحفوظ في ملف الإعدادات: <b>{saved_periods}</b></div>
    <div>عدد الحصص في التشغيل الحالي: <b>{running_periods}</b></div>
    <div>{status_line}</div>
</div>
"""


def render_school_config_summary_html(config=None):
    cfg = dict(_current_identity_config())
    loaded_config = load_school_config()
    cfg.update(loaded_config)
    if isinstance(config, dict):
        cfg.update(config)

    directorate = _identity_directorate_full_name(
        cfg.get("directorate_region", DIRECTORATE_REGION)
    )
    saved_periods = get_config_periods_per_day(cfg)
    if saved_periods == int(MAX_PERIODS):
        periods_text = f"{saved_periods}"
    else:
        periods_text = f"المحفوظ: {saved_periods} | التشغيل الحالي: {MAX_PERIODS} — يلزم إعادة التشغيل للتطبيق"

    return f"""
<div style='background:#fffde7;color:#4d3b00;padding:12px;border-radius:10px;border-right:5px solid {html_lib.escape(str(cfg.get("accent_color", ACCENT_COLOR)))};margin-bottom:12px;font-weight:800;line-height:1.8;'>
    ملف إعدادات المدرسة: <b>{html_lib.escape(SCHOOL_CONFIG_FILE)}</b><br>
    المدرسة الحالية: <b>{html_lib.escape(str(cfg.get("school_name", SCHOOL_NAME)))}</b><br>
    الوزارة: <b>{html_lib.escape(DEFAULT_SCHOOL_CONFIG["ministry_name"])}</b><br>
    المديرية: <b>{html_lib.escape(directorate)}</b><br>
    اسم النظام: <b>{html_lib.escape(DEFAULT_SCHOOL_CONFIG["system_name"])} - {html_lib.escape(DEFAULT_SCHOOL_CONFIG["system_subtitle"])}</b><br>
    عدد الحصص اليومية: <b>{html_lib.escape(str(periods_text))}</b>
</div>
"""


@state_locked
def save_school_operational_settings(periods_per_day, is_owner=False, actor_name="", actor_role=""):
    current_config = load_school_config()
    current_saved = get_config_periods_per_day(current_config)

    if not bool(is_owner):
        return (
            gr.update(value=current_saved),
            "<div style='color:#b91c1c;font-weight:800;'>رفض الحفظ: إعدادات التشغيل مخصصة لمالك النظام فقط.</div>",
            gr.update(value=render_school_config_summary_html(current_config)),
            gr.update(value=render_operational_settings_status_html(current_config)),
        )

    new_periods = _coerce_periods_per_day(periods_per_day, current_saved)
    if new_periods not in (7, 8):
        return (
            gr.update(value=current_saved),
            "<div style='color:#b91c1c;font-weight:800;'>عدد الحصص يجب أن يكون 7 أو 8 فقط.</div>",
            gr.update(value=render_school_config_summary_html(current_config)),
            gr.update(value=render_operational_settings_status_html(current_config)),
        )

    old_periods = current_saved
    current_config["periods_per_day"] = int(new_periods)

    if not safe_write_json(SCHOOL_CONFIG_FILE, current_config):
        return (
            gr.update(value=old_periods),
            "<div style='color:#b91c1c;font-weight:800;'>تعذر حفظ إعداد عدد الحصص في ملف المدرسة.</div>",
            gr.update(value=render_school_config_summary_html(current_config)),
            gr.update(value=render_operational_settings_status_html(current_config)),
        )

    if old_periods != new_periods:
        write_audit_log(
            "تعديل إعداد عدد الحصص اليومية",
            target_teacher="",
            old_value=old_periods,
            new_value=new_periods,
            details="تحديث عدد الحصص اليومية من إعدادات التشغيل المدرسية. يلزم إعادة تشغيل المنظومة للتطبيق.",
            actor_name=actor_name,
            actor_role=actor_role,
        )

    saved_config = load_school_config()
    reboot_note = ""
    if int(MAX_PERIODS) != int(new_periods):
        reboot_note = "<br>⚠️ يلزم عمل Restart / Factory reboot حتى تعمل المنظومة بعدد الحصص الجديد."

    return (
        gr.update(value=int(new_periods)),
        (
            "<div style='color:#166534;background:#dcfce7;padding:10px;"
            "border-radius:8px;font-weight:800;line-height:1.8;'>"
            f"تم حفظ عدد الحصص اليومية: {int(new_periods)}.{reboot_note}"
            "</div>"
        ),
        gr.update(value=render_school_config_summary_html(saved_config)),
        gr.update(value=render_operational_settings_status_html(saved_config)),
    )

def render_school_identity_preview_html(
    system_name,
    system_subtitle,
    school_name,
    developer_credit,
    logo_url,
    theme_color,
    theme_color_2,
    accent_color,
    directorate_region=None,
):
    preview_cfg = {
        "system_name": DEFAULT_SCHOOL_CONFIG["system_name"],
        "system_subtitle": DEFAULT_SCHOOL_CONFIG["system_subtitle"],
        "ministry_name": DEFAULT_SCHOOL_CONFIG["ministry_name"],
        "directorate_region": _normalize_identity_text(
            directorate_region if directorate_region is not None else DIRECTORATE_REGION,
            DEFAULT_SCHOOL_CONFIG["directorate_region"],
            80,
        ),
        "school_name": _normalize_identity_text(
            school_name, DEFAULT_SCHOOL_CONFIG["school_name"], 140
        ),
        "developer_credit": DEFAULT_SCHOOL_CONFIG["developer_credit"],
        "logo_url": str(logo_url or SCHOOL_LOGO_URL).strip(),
        "theme_color": _normalize_hex_color(
            theme_color, DEFAULT_SCHOOL_CONFIG["theme_color"]
        ),
        "theme_color_2": _normalize_hex_color(
            theme_color_2, DEFAULT_SCHOOL_CONFIG["theme_color_2"]
        ),
        "accent_color": _normalize_hex_color(
            accent_color, DEFAULT_SCHOOL_CONFIG["accent_color"]
        ),
    }

    directorate_full = _identity_directorate_full_name(
        preview_cfg["directorate_region"]
    )
    logo_src = html_lib.escape(
        _resolve_identity_logo_source(preview_cfg["logo_url"]),
        quote=True,
    )
    return f"""
<div style='direction:rtl;border:1px solid #d1d5db;border-radius:16px;overflow:hidden;background:#ffffff;box-shadow:0 8px 22px rgba(0,0,0,0.08);'>
    <div style='background:linear-gradient(145deg,#003d33 0%,{preview_cfg["theme_color"]} 45%,{preview_cfg["theme_color_2"]} 100%);padding:22px;text-align:center;'>
        <div style='font-size:13px;font-weight:800;color:rgba(255,255,255,0.92);margin-bottom:6px;'>{html_lib.escape(preview_cfg["ministry_name"])}</div>
        <div style='font-size:12px;font-weight:700;color:rgba(255,255,255,0.86);margin-bottom:12px;'>{html_lib.escape(directorate_full)}</div>
        <img src='{logo_src}' style='width:92px;height:92px;object-fit:contain;border-radius:50%;background:#fff;padding:4px;border:3px solid {preview_cfg["accent_color"]};'>
        <div style='font-size:24px;font-weight:900;color:{preview_cfg["accent_color"]};margin-top:12px;'>{html_lib.escape(preview_cfg["system_name"])}</div>
        <div style='font-size:14px;font-weight:700;color:#fff;margin-top:4px;'>{html_lib.escape(preview_cfg["system_subtitle"])}</div>
        <div style='font-size:13px;font-weight:700;color:rgba(255,255,255,0.9);margin-top:8px;'>{html_lib.escape(preview_cfg["school_name"])}</div>
    </div>
    <div style='padding:12px;text-align:center;color:#334155;font-weight:700;'>{html_lib.escape(preview_cfg["developer_credit"])}</div>
</div>
"""

def preview_school_identity_settings(
    system_name,
    system_subtitle,
    school_name,
    developer_credit,
    directorate_region,
    logo_url,
    theme_color,
    theme_color_2,
    accent_color,
    is_owner=False,
):
    if not bool(is_owner):
        return (
            gr.update(),
            "<div style='color:#b91c1c;font-weight:800;'>هذه الأداة مخصصة لمالك النظام فقط.</div>",
        )

    preview = render_school_identity_preview_html(
        DEFAULT_SCHOOL_CONFIG["system_name"],
        DEFAULT_SCHOOL_CONFIG["system_subtitle"],
        school_name,
        DEFAULT_SCHOOL_CONFIG["developer_credit"],
        logo_url,
        theme_color,
        theme_color_2,
        accent_color,
        directorate_region,
    )
    return (
        gr.update(value=preview),
        "<div style='color:#0f766e;font-weight:800;'>تم تحديث المعاينة فقط، ولم تُحفظ التغييرات بعد.</div>",
    )

def _identity_full_output(config, status_html):
    preview = render_school_identity_preview_html(
        DEFAULT_SCHOOL_CONFIG["system_name"],
        DEFAULT_SCHOOL_CONFIG["system_subtitle"],
        config["school_name"],
        DEFAULT_SCHOOL_CONFIG["developer_credit"],
        config["logo_url"],
        config["theme_color"],
        config["theme_color_2"],
        config["accent_color"],
        config.get("directorate_region", DEFAULT_SCHOOL_CONFIG["directorate_region"]),
    )
    return (
        gr.update(value=DEFAULT_SCHOOL_CONFIG["system_name"]),
        gr.update(value=DEFAULT_SCHOOL_CONFIG["system_subtitle"]),
        gr.update(value=config["school_name"]),
        gr.update(value=config.get("directorate_region", DEFAULT_SCHOOL_CONFIG["directorate_region"])),
        gr.update(value=DEFAULT_SCHOOL_CONFIG["developer_credit"]),
        gr.update(value=config["logo_url"]),
        gr.update(value=None),
        gr.update(value=config["theme_color"]),
        gr.update(value=config["theme_color_2"]),
        gr.update(value=config["accent_color"]),
        gr.update(value=status_html),
        gr.update(value=preview),
        gr.update(value=build_login_branding_html(config)),
        gr.update(value=build_login_credits_html(config)),
        gr.update(value=build_header_html(config)),
        gr.update(value=build_home_hero_html(config)),
        gr.update(value=render_school_config_summary_html(config)),
    )

@state_locked
def save_school_identity_settings(
    system_name,
    system_subtitle,
    school_name,
    developer_credit,
    directorate_region,
    logo_url,
    logo_upload,
    theme_color,
    theme_color_2,
    accent_color,
    is_owner=False,
):
    if not bool(is_owner):
        return _identity_full_output(
            _current_identity_config(),
            "<div style='color:#b91c1c;font-weight:800;'>رفض الحفظ: إعدادات الهوية مخصصة لمالك النظام فقط.</div>",
        )

    school_name_clean = _normalize_identity_text(school_name, "", 140)
    directorate_region_clean = _normalize_identity_text(
        directorate_region,
        DEFAULT_SCHOOL_CONFIG["directorate_region"],
        80,
    )
    if not school_name_clean:
        return _identity_full_output(
            _current_identity_config(),
            "<div style='color:#b91c1c;font-weight:800;'>اسم المدرسة حقل إلزامي.</div>",
        )

    colors = {
        "theme_color": str(theme_color or "").strip(),
        "theme_color_2": str(theme_color_2 or "").strip(),
        "accent_color": str(accent_color or "").strip(),
    }
    invalid_colors = [
        key for key, value in colors.items()
        if not re.fullmatch(r"#[0-9a-fA-F]{6}", value)
    ]
    if invalid_colors:
        return _identity_full_output(
            _current_identity_config(),
            "<div style='color:#b91c1c;font-weight:800;'>ألوان الهوية يجب أن تكون بصيغة HEX مثل #004d40.</div>",
        )

    saved_logo_value = str(logo_url or "").strip()
    try:
        uploaded_logo = _save_uploaded_identity_logo(logo_upload)
        if uploaded_logo:
            saved_logo_value = uploaded_logo
    except Exception as exc:
        return _identity_full_output(
            _current_identity_config(),
            f"<div style='color:#b91c1c;font-weight:800;'>{html_lib.escape(str(exc))}</div>",
        )

    if not saved_logo_value:
        saved_logo_value = str(DEFAULT_SCHOOL_CONFIG["logo_url"])

    if not _is_valid_identity_logo_value(saved_logo_value):
        return _identity_full_output(
            _current_identity_config(),
            "<div style='color:#b91c1c;font-weight:800;'>رابط أو ملف الشعار غير صالح.</div>",
        )

    new_config = load_school_config()
    for fixed_key in FIXED_IDENTITY_KEYS:
        new_config[fixed_key] = DEFAULT_SCHOOL_CONFIG[fixed_key]

    new_config.update({
        "school_name": school_name_clean,
        "directorate_region": directorate_region_clean,
        "logo_url": saved_logo_value,
        "theme_color": colors["theme_color"].lower(),
        "theme_color_2": colors["theme_color_2"].lower(),
        "accent_color": colors["accent_color"].lower(),
    })

    if not safe_write_json(SCHOOL_CONFIG_FILE, new_config):
        return _identity_full_output(
            _current_identity_config(),
            "<div style='color:#b91c1c;font-weight:800;'>تعذر حفظ ملف إعدادات المدرسة.</div>",
        )

    _apply_school_identity_globals(new_config)
    return _identity_full_output(
        _current_identity_config(),
        "<div style='color:#166534;background:#dcfce7;padding:10px;border-radius:8px;font-weight:800;'>تم حفظ هوية المدرسة بنجاح. العناصر الثابتة بقيت كما هي، وتغيرت المدرسة والمحافظة والشعار والألوان فقط.</div>",
    )

@state_locked
def reset_school_identity_settings(is_owner=False):
    if not bool(is_owner):
        return _identity_full_output(
            _current_identity_config(),
            "<div style='color:#b91c1c;font-weight:800;'>رفض الاستعادة: هذه الأداة مخصصة لمالك النظام فقط.</div>",
        )

    config = load_school_config()

    for key in FIXED_IDENTITY_KEYS:
        config[key] = DEFAULT_SCHOOL_CONFIG[key]
    for key in IDENTITY_CONFIG_KEYS:
        config[key] = DEFAULT_SCHOOL_CONFIG[key]

    if not safe_write_json(SCHOOL_CONFIG_FILE, config):
        return _identity_full_output(
            _current_identity_config(),
            "<div style='color:#b91c1c;font-weight:800;'>تعذر استعادة الهوية الافتراضية.</div>",
        )

    _apply_school_identity_globals(config)
    return _identity_full_output(
        _current_identity_config(),
        "<div style='color:#166534;background:#dcfce7;padding:10px;border-radius:8px;font-weight:800;'>تمت استعادة الهوية الافتراضية. تُطبق الألوان العامة بالكامل بعد إعادة تشغيل التطبيق.</div>",
    )


        

os.makedirs(SWAP_IMG_DIR, exist_ok=True)

def sync_current_school_days():
    current_day = get_current_day_oman()
    return gr.update(value=current_day), gr.update(value=current_day)


def generate_swap_table_image(state, teacher_name, day_name):
    filename = generate_swap_table_image_core(
        state,
        teacher_name,
        day_name,
        SYSTEM_NAME,
        SYSTEM_SUBTITLE,
        THEME_COLOR,
        ACCENT_COLOR,
    )
    return gr.update(value=filename)

def export_confirmed_swaps_excel():
    filename = export_confirmed_swaps_excel_core()
    return gr.update(value=filename)

def load_confirmed_swaps_for_context(t, d):
    state = load_confirmed_swaps_for_context_core(t, d)
    return state, gr.update(value=render_swap_table_html(state))

def clear_swap_detail_ui():
    choices, selected_value, message_value, button_html, confirm_interactive = clear_swap_detail_ui_core()
    return (
        gr.update(choices=choices, value=selected_value, visible=True),
        gr.update(value=message_value, visible=True),
        gr.update(value=button_html, visible=True),
        gr.update(visible=True, interactive=confirm_interactive)
    )

def confirm_swap(t, period_value, choice, d, msg_text, state, actor_name="", actor_role=""):
    current_state, warning = confirm_swap_core(
        t, period_value, choice, d, msg_text, state, actor_name, actor_role
    )
    return current_state, gr.update(value=render_swap_table_html(current_state) + warning)







def add_manual_staff(name, dept, phone, role, dept_filter, is_owner=False):
    raw = add_manual_staff_core(name, dept, phone, role, dept_filter, is_owner=is_owner)
    return (
        raw["message"],
        gr.update(**raw["abs_update"]),
        gr.update(**raw["teacher_update_1"]),
        gr.update(**raw["teacher_update_2"]),
        gr.update(**raw["staff_names_update"]),
        gr.update(**raw["name_input_update"]),
        gr.update(**raw["phone_input_update"]),
    )
def process_admin_excel(file, dept_filter):
    if file is None:
        return (
            "<div style='color:red; font-weight:bold;'>❌ الرجاء رفع ملف الإداريين أولاً.</div>",
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update()
        )

    try:
        df = pd.read_excel(file.name, header=None) if not file.name.endswith('.csv') else pd.read_csv(file.name, header=None)
        df = df.fillna("")

        added_or_updated = 0
        found_names = []

        for r in range(len(df)):
            raw_phone = str(df.iloc[r, 0]).strip() if df.shape[1] > 0 else ""
            raw_role = str(df.iloc[r, 1]).strip() if df.shape[1] > 1 else ""
            raw_name = str(df.iloc[r, 2]).strip() if df.shape[1] > 2 else ""

            if not raw_name or raw_name == "nan":
                continue

            t_name = clean_teacher_name(raw_name)
            if not t_name or len(t_name) < 3:
                continue

            if raw_phone.endswith(".0"):
                raw_phone = raw_phone[:-2]

            phone_digits = re.sub(r"\D", "", raw_phone)
            if len(phone_digits) == 8:
                phone_digits = "968" + phone_digits

            role_val = raw_role if raw_role else "أخصائي اجتماعي"

            if t_name not in teachers_db:
                teachers_db[t_name] = {
                    "dept": "الهيئة الإدارية",
                    "cover_count": 0,
                    "absent_count": 0,
                    "shortcoming_count": 0,
                    "phone": "",
                    "specialty": "",
                    "role": role_val,
                    "exempt_days": [],
                    "exempt_periods": [],
                    "exempt_slots": [],
                    "absence_dates": [],
                    "الأحد": {},
                    "الإثنين": {},
                    "الثلاثاء": {},
                    "الأربعاء": {},
                    "الخميس": {}
                }
            else:
                teachers_db[t_name]["dept"] = "الهيئة الإدارية"
                teachers_db[t_name]["role"] = role_val

            teachers_db[t_name]["phone"] = phone_digits if phone_digits else teachers_db[t_name].get("phone", "")
            found_names.append(t_name)
            added_or_updated += 1

        save_db()

        dept_filter = resolve_effective_dept(dept_filter)
        choices_all = get_teacher_choices(dept_filter)
        abs_choices = get_absentee_choices(dept_filter)
        t_names_filtered = sorted([t for t, d in teachers_db.items() if dept_filter == "الكل" or d.get("dept") == dept_filter])

        names_list_str = "، ".join(found_names) if found_names else "لا توجد أسماء صالحة"
        msg = (
            f"<div style='color:#2e7d32; font-weight:bold; background:#e8f5e9; padding:10px; border-radius:5px;'>"
            f"✅ تم استيراد/تحديث ({added_or_updated}) من الإداريين بنجاح."
            f"<br>👥 الأسماء: {names_list_str}"
            f"</div>"
        )

        return (
            msg,
            gr.update(choices=abs_choices),
            gr.update(choices=choices_all, value=None),
            gr.update(choices=choices_all, value=None),
            gr.update(choices=t_names_filtered, value=None),
            gr.update(value=None),
            gr.update(value=get_updated_balance(dept_filter))
        )

    except Exception as e:
        return (
            f"<div style='color:red; font-weight:bold;'>❌ خطأ أثناء رفع ملف الإداريين: {str(e)}</div>",
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update()
        )
def process_phone_excel(file):
    if file is None: return "<div style='color:red; font-weight:bold;'>❌ الرجاء رفع ملف أرقام الهواتف.</div>", gr.update()
    try:
        df = pd.read_excel(file.name, header=None) if not file.name.endswith('.csv') else pd.read_csv(file.name, header=None)
        updated = 0
        db_fingerprints = {k: get_name_fingerprint(k) for k in teachers_db.keys()}
        for r in range(len(df)):
            raw_name = str(df.iloc[r, 0]).strip()
            raw_phone = str(df.iloc[r, 1]).strip()
            if not raw_name or raw_name == 'nan': continue
            
            if raw_phone.endswith('.0'):
                raw_phone = raw_phone[:-2]
                
            phone_digits = re.sub(r'\D', '', raw_phone) 
            if len(phone_digits) == 8: phone_digits = "968" + phone_digits
            if not phone_digits: continue
            
            phone_first_name, phone_name_fingerprint = get_name_fingerprint(raw_name)
            if not phone_first_name: continue
            
            for db_key, (db_first_name, db_words) in db_fingerprints.items():
                if db_first_name == phone_first_name and len(db_words) > 0 and db_words.issubset(phone_name_fingerprint):
                    teachers_db[db_key]["phone"] = phone_digits
                    updated += 1
                    break
                    
        save_db()
        return f"<div style='color:#2e7d32; font-weight:bold; background:#e8f5e9; padding:10px; border-radius:5px;'>✅ تم بنجاح ربط أرقام ({updated}) معلماً بفضل الرادار الذكي!</div>", gr.update(value=None)
    except Exception as e: return f"<div style='color:red;'>❌ خطأ: {str(e)}</div>", gr.update()

def process_uploaded_excel(file, selected_dept, current_day):
    (
        dept_choices,
        abs_choices,
        teacher_choices_a,
        teacher_choices_b,
        balance_html,
        absences_html,
        day_overview,
        message_html,
        teacher_names_all,
        reset_upload,
    ) = process_uploaded_excel_core(file, selected_dept, current_day)

    return (
        gr.update() if dept_choices is None else gr.update(choices=dept_choices),
        gr.update() if abs_choices is None else gr.update(choices=abs_choices),
        gr.update() if teacher_choices_a is None else gr.update(choices=teacher_choices_a, value=None),
        gr.update() if teacher_choices_b is None else gr.update(choices=teacher_choices_b, value=None),
        gr.update() if balance_html is None else gr.update(value=balance_html),
        gr.update() if absences_html is None else gr.update(value=absences_html),
        gr.update() if day_overview is None else gr.update(value=day_overview),
        message_html,
        gr.update() if teacher_names_all is None else gr.update(choices=teacher_names_all),
        gr.update(value=None) if reset_upload else gr.update(),
    )

def delete_department_data(dept_to_delete, current_day):
    (
        dept_choices,
        abs_choices,
        teacher_choices_a,
        teacher_choices_b,
        balance_html,
        absences_html,
        day_overview,
        message_html,
        teacher_names_all,
        reset_upload,
    ) = delete_department_data_core(dept_to_delete, current_day)

    return (
        gr.update() if dept_choices is None else gr.update(choices=dept_choices),
        gr.update() if abs_choices is None else gr.update(choices=abs_choices),
        gr.update() if teacher_choices_a is None else gr.update(choices=teacher_choices_a, value=None),
        gr.update() if teacher_choices_b is None else gr.update(choices=teacher_choices_b, value=None),
        gr.update() if balance_html is None else gr.update(value=balance_html),
        gr.update() if absences_html is None else gr.update(value=absences_html),
        gr.update() if day_overview is None else gr.update(value=day_overview),
        message_html,
        gr.update() if teacher_names_all is None else gr.update(choices=teacher_names_all, value=None),
        gr.update(value=None) if reset_upload else gr.update(),
    )


def get_day_table_updates(day_name, dept_filter, page=0):
    (
        df,
        table_html,
        pager_visible,
        prev_interactive,
        next_interactive,
        page_html,
        safe_page,
    ) = get_day_table_updates_core(day_name, dept_filter, page)

    return (
        gr.update(value=df, visible=False),
        gr.update(value=table_html, visible=True),
        gr.update(visible=pager_visible),
        gr.update(interactive=prev_interactive),
        gr.update(interactive=next_interactive),
        gr.update(value=page_html, visible=True),
        safe_page,
    )

def change_day_page(delta, day_name, dept_filter, current_page):
    effective_dept = resolve_effective_dept(dept_filter)
    df = get_day_overview(day_name, effective_dept)
    total_pages = max(1, (len(df) + PAGE_SIZE - 1) // PAGE_SIZE)

    try:
        safe_page = int(current_page or 0)
    except Exception:
        safe_page = 0

    new_page = max(0, min(safe_page + delta, total_pages - 1))
    return get_day_table_updates(day_name, effective_dept, new_page)

def get_teacher_weekly_schedule(teacher_name):
    teacher_name = str(teacher_name or "").split(" (")[0].strip()

    if (
        not teacher_name
        or teacher_name not in teachers_db
        or teachers_db[teacher_name].get("dept") == "الهيئة الإدارية"
        or teachers_db[teacher_name].get("role", "معلم") in ADMIN_ROLES
    ):
        return pd.DataFrame(columns=["اليوم"] + [f"ح {p}" for p in range(1, MAX_PERIODS + 1)])

    rows = [
        {
            "اليوم": day,
            **{f"ح {p}": teachers_db[teacher_name].get(day, {}).get(p, "-") for p in range(1, MAX_PERIODS + 1)}
        }
        for day in SCHOOL_WEEK_DAYS
    ]
    return pd.DataFrame(rows)


def render_weekly_schedule_html(df):
    if df is None or df.empty:
        return """
        <div style='text-align:center; color:#64748b; padding:18px; background:#f8fafc; border:1px dashed #cbd5e1; border-radius:12px; direction:rtl;'>
            لا توجد بيانات لعرضها في جدول المعلم.
        </div>
        """

    safe_df = df.fillna("-").copy()

    headers_html = "".join(
        f"<th style='padding:10px 12px; background:#0f766e; color:#ffffff; border:1px solid #d1d5db; white-space:nowrap; font-size:13px;'>{col}</th>"
        for col in safe_df.columns
    )

    rows_html = ""
    for _, row in safe_df.iterrows():
        row_cells = "".join(
            f"<td style='padding:10px 12px; border:1px solid #d1d5db; white-space:nowrap; font-size:13px; color:#0f172a; font-weight:600;'>{row[col]}</td>"
            for col in safe_df.columns
        )
        rows_html += f"<tr>{row_cells}</tr>"

    return f"""
    <div style='background:#ffffff; border:1px solid #dbeafe; border-radius:12px; overflow:hidden; box-shadow:0 1px 2px rgba(15,23,42,0.05); direction:rtl;'>
        <div style='overflow-x:auto; width:100%; -webkit-overflow-scrolling:touch;'>
            <table style='width:100%; min-width:760px; border-collapse:collapse; text-align:center; direction:rtl; font-family:Cairo, Arial, sans-serif;'>
                <thead>
                    <tr>{headers_html}</tr>
                </thead>
                <tbody>{rows_html}</tbody>
            </table>
        </div>
    </div>
    """


def get_teacher_weekly_schedule_html(teacher_name):
    return render_weekly_schedule_html(get_teacher_weekly_schedule(teacher_name))



def draw_schedule_image(df, day_name):
    return draw_schedule_image_core(df, day_name)




def generate_image_only(dept, day_name):
    effective_dept = resolve_effective_dept(dept)
    target_date = get_date_of_weekday(day_name)
    display_records = [r for r in daily_db if r["date"] == target_date and (effective_dept == "الكل" or r["dept"] == effective_dept)]
    df = pd.DataFrame(display_records, columns=["المعلم الغائب", "الصف", "الحصة", "المعلم البديل", "dept", "date", "حالة_التكليف"]).sort_values(["المعلم الغائب", "الحصة"])
    if not df.empty:
        df["المعلم البديل عرض"] = df.apply(format_sub_display_for_image, axis=1)
        df["المعلم الغائب"] = df["المعلم الغائب"].apply(format_teacher_name)
        img_path = draw_schedule_image(df, day_name)
        return gr.update(value=img_path)
    return gr.update(value=None)

# ✂️ المقص الرياضي الحاسم


def force_refresh_data(dept, day_name, is_admin_logged_in, current_abs):
    load_db()         
    load_daily_db()   
    return refresh_ui_on_change(dept, day_name, is_admin_logged_in, current_abs)




    
def get_generation_button_updates(absent_list, day_name, dept_filter, generation_state):
    cleaned = normalize_absent_names(absent_list)
    has_selection = bool(cleaned)
    target_date = get_date_of_weekday(day_name)

    has_existing = any(
        r.get("date") == target_date and (dept_filter == "الكل" or r.get("dept") == dept_filter)
        for r in daily_db
    )

    same_context = same_generation_context(generation_state, day_name, dept_filter)

    prev_absents = normalize_absent_names(
        generation_state.get("absents", [])
    ) if same_context else []

    if has_existing and not prev_absents:
        prev_absents = get_existing_absents_for_context(day_name, dept_filter)

    selection_changed = set(cleaned) != set(prev_absents)

    # الأصفر: يعمل فقط إذا لا يوجد توليد سابق أو إذا تغيرت القائمة
    generate_enabled = has_selection and (not has_existing or selection_changed)

    # البرتقالي: ظاهر بعد أول توليد، لكنه لا يتفعل إلا عند تغير القائمة
    regen_visible = has_existing
    regen_enabled = has_existing and has_selection and selection_changed

    return (
        gr.update(interactive=generate_enabled),
        gr.update(visible=regen_visible, interactive=regen_enabled),
    )


def clear_generated_image():
    return gr.update(value=None)


def school_data_panel_js(panel_name):
    """
    v1.8.3 Fix 13b — direct DOM visibility for school data panels.
    لا يكتفي بتلوين البطاقة؛ يفتح اللوحة بصريًا من النقرة الأولى،
    ثم تُثبت دالة show_school_data_panel الحالة من جهة Gradio.
    """
    panel = str(panel_name or "overview").strip()
    allowed = {"overview", "references", "identity", "periods", "accounts", "audit"}
    if panel not in allowed:
        panel = "overview"

    classes = [
        "masar-sd-panel-references",
        "masar-sd-panel-identity",
        "masar-sd-panel-periods",
        "masar-sd-panel-accounts",
        "masar-sd-panel-audit",
    ]
    class_to_add = "" if panel == "overview" else f"masar-sd-panel-{panel}"
    class_args = ", ".join(repr(c) for c in classes)

    panel_map = {
        "references": "school_data_panel_references",
        "identity": "school_data_panel_identity",
        "periods": "school_data_panel_periods",
        "accounts": "school_data_panel_accounts",
        "audit": "school_data_panel_audit",
    }
    panel_map_js = json.dumps(panel_map, ensure_ascii=False)

    return f"""() => {{
        const selectedPanel = {panel!r};
        const panelMap = {panel_map_js};

        document.body.classList.remove({class_args});
        if ({class_to_add!r}) {{
            document.body.classList.add({class_to_add!r});
        }}

        const setVisible = (id, visible) => {{
            const el = document.getElementById(id);
            if (!el) return;
            if (visible) {{
                el.style.setProperty('display', 'block', 'important');
                el.style.setProperty('visibility', 'visible', 'important');
                el.removeAttribute('hidden');
                el.setAttribute('aria-hidden', 'false');
            }} else {{
                el.style.setProperty('display', 'none', 'important');
                el.style.setProperty('visibility', 'hidden', 'important');
                el.setAttribute('aria-hidden', 'true');
            }}
        }};

        const showOverview = selectedPanel === 'overview';
        setVisible('school_data_overview_storage', showOverview);
        setVisible('school_data_overview_config', showOverview);

        Object.entries(panelMap).forEach(([key, id]) => {{
            setVisible(id, key === selectedPanel);
        }});

        const status = document.getElementById('school_data_section_status');
        if (status) {{
            status.style.setProperty('display', 'block', 'important');
            status.style.setProperty('visibility', 'visible', 'important');
        }}
    }}"""


def show_school_data_panel(panel_name="overview"):
    panel = str(panel_name or "overview").strip()
    labels = {
        "overview": "اختر بطاقة من الأعلى لعرض أدواتها. حالة التخزين وإعدادات المدرسة تظهر هنا دائمًا.",
        "references": "رفع وتحديث الجداول المرجعية، أرقام المعلمين، وملف الإداريين.",
        "identity": "تعديل اسم المدرسة، المحافظة، الشعار، والألوان العامة.",
        "periods": "إعدادات التشغيل المدرسية: اختيار عدد الحصص اليومية 7 أو 8.",
        "accounts": "إدارة الحسابات والرموز، ثم ضبط الترحيب والمسميات من صفحة واحدة.",
        "audit": "متابعة سجل العمليات والنسخ الاحتياطية في أقسام واضحة.",
    }

    show_overview = panel == "overview"

    return (
        gr.update(value=f"<div class='school-data-panel-title'>{labels.get(panel, labels['overview'])}</div>"),
        gr.update(visible=show_overview),
        gr.update(visible=show_overview),
        gr.update(visible=(panel == "references")),
        gr.update(visible=(panel == "identity")),
        gr.update(visible=(panel == "periods")),
        gr.update(visible=(panel == "accounts")),
        gr.update(visible=(panel == "audit")),
    )







def run_main_generation(absent_list, day_name, dept_filter, max_reserves, is_admin_logged_in, generation_state, actor_name="", actor_role=""):
    cleaned = normalize_absent_names(absent_list)

    if not cleaned or not day_name:
        ui = refresh_ui_on_change(dept_filter, day_name, is_admin_logged_in, current_abs=cleaned)
        btn_upd, regen_upd = get_generation_button_updates(cleaned, day_name, dept_filter, generation_state)
        return tuple(ui) + (btn_upd, regen_upd, generation_state)

    same_context = same_generation_context(generation_state, day_name, dept_filter)
    prev_absents = normalize_absent_names(
        generation_state.get("absents", [])
    ) if same_context else []

    if not prev_absents:
        prev_absents = get_existing_absents_for_context(day_name, dept_filter)

    if not prev_absents:
        run_list = cleaned
    else:
        run_list = [name for name in cleaned if name not in prev_absents]

    if run_list:
        assign_logic_core(
            run_list,
            day_name,
            dept_filter,
            max_reserves,
            False,
            is_admin_logged_in,
            actor_name=actor_name,
            actor_role=actor_role,
        )

    new_state = {
        "day": day_name,
        "dept": dept_filter,
        "absents": cleaned,
        "signature": build_generation_signature(cleaned, day_name, dept_filter),
        "generated": True,
        "has_results": True,
    }

    ui = list(refresh_ui_on_change(dept_filter, day_name, is_admin_logged_in, current_abs=cleaned))
    btn_upd, regen_upd = get_generation_button_updates(cleaned, day_name, dept_filter, new_state)

    return tuple(ui) + (btn_upd, regen_upd, new_state)

def run_full_regeneration(absent_list, day_name, dept_filter, max_reserves, is_admin_logged_in, generation_state, actor_name="", actor_role=""):
    cleaned = normalize_absent_names(absent_list)

    if not cleaned or not day_name:
        ui = refresh_ui_on_change(dept_filter, day_name, is_admin_logged_in, current_abs=cleaned)
        btn_upd, regen_upd = get_generation_button_updates(cleaned, day_name, dept_filter, generation_state)
        return tuple(ui) + (btn_upd, regen_upd, generation_state)

    rollback_auto_assignments_for_absentees_core(cleaned, day_name, actor_name, actor_role)
    assign_logic_core(
        cleaned,
        day_name,
        dept_filter,
        max_reserves,
        False,
        is_admin_logged_in,
        actor_name=actor_name,
        actor_role=actor_role,
    )

    new_state = {
        "day": day_name,
        "dept": dept_filter,
        "absents": cleaned,
        "signature": build_generation_signature(cleaned, day_name, dept_filter),
        "generated": True,
        "has_results": True,
    }

    ui = refresh_ui_on_change(dept_filter, day_name, is_admin_logged_in, current_abs=cleaned)
    btn_upd, regen_upd = get_generation_button_updates(cleaned, day_name, dept_filter, new_state)

    return tuple(ui) + (btn_upd, regen_upd, new_state)
    
def refresh_ui_on_change(dept, day_name, is_admin_logged_in, current_abs=None):
    refresh_values = refresh_ui_on_change_core(dept, day_name, is_admin_logged_in, current_abs=current_abs)
    if len(refresh_values) != 27:
        raise ValueError(f"refresh_ui_on_change_core returned {len(refresh_values)} outputs, expected 27")

    (
        abs_update_raw,
        balance_html,
        absences_html,
        shortcomings_html,
        day_df,
        day_table_html,
        pager_visible,
        prev_interactive,
        next_interactive,
        page_html,
        safe_page,
        t_names_filtered,
        teacher_schedule_choices,
        choices,
        warning_html,
        styled_table_html,
        opts_abs,
        df,
        summary_txt,
        html_cards,
        dynamic_header,
        admin_title_val,
        admin_help_val,
        period_update_raw,
        cb_cross_update_raw,
        first_action_interactive,
        second_action_interactive,
    ) = refresh_values

    return (
        gr.update(**abs_update_raw),
        gr.update(value=balance_html),
        gr.update(value=absences_html),
        gr.update(value=shortcomings_html),
        gr.update(value=day_df, visible=False),
        gr.update(value=day_table_html, visible=True),
        gr.update(visible=pager_visible),
        gr.update(interactive=prev_interactive),
        gr.update(interactive=next_interactive),
        gr.update(value=page_html, visible=True),
        safe_page,
        gr.update(choices=t_names_filtered, value=None),
        gr.update(choices=teacher_schedule_choices, value=None),
        gr.update(choices=choices, value=None),
        warning_html,
        gr.update(value=styled_table_html),
        gr.update(choices=opts_abs, value=None),
        df,
        summary_txt,
        html_cards,
        dynamic_header,
        admin_title_val,
        gr.update(value=admin_help_val),
        gr.update(**period_update_raw),
        gr.update(**cb_cross_update_raw),
        gr.update(interactive=first_action_interactive),
        gr.update(interactive=second_action_interactive),
    )

def assign_logic(absent_list, day_name, dept_filter, max_reserves, is_alt, is_admin_logged_in, actor_name="", actor_role=""):
    result = assign_logic_core(
        absent_list,
        day_name,
        dept_filter,
        max_reserves,
        is_alt,
        is_admin_logged_in,
        actor_name=actor_name,
        actor_role=actor_role,
    )
    return refresh_ui_on_change(
        result["refresh_dept"],
        result["refresh_day"],
        result["refresh_is_admin"],
        current_abs=result.get("refresh_current_abs"),
    )
    
def cancel_teacher_absence(abs_t, day_name, dept_filter, is_admin_logged_in, current_abs, actor_name="", actor_role=""):
    result = cancel_teacher_absence_core(
        abs_t,
        day_name,
        dept_filter,
        is_admin_logged_in,
        current_abs,
        actor_name=actor_name,
        actor_role=actor_role,
    )
    return refresh_ui_on_change(
        result["refresh_dept"],
        result["refresh_day"],
        result["refresh_is_admin"],
        current_abs=result.get("refresh_current_abs"),
    )
def cancel_teacher_absence_with_generation_state(abs_t, day_name, dept_filter, is_admin_logged_in, current_abs, actor_name="", actor_role=""):
    ui = cancel_teacher_absence(abs_t, day_name, dept_filter, is_admin_logged_in, current_abs, actor_name, actor_role)

    remaining_absents = get_existing_absents_for_context(day_name, dept_filter)

    if remaining_absents:
        new_state = {
            "day": day_name,
            "dept": dept_filter,
            "absents": remaining_absents,
            "signature": build_generation_signature(remaining_absents, day_name, dept_filter),
            "generated": True,
            "has_results": True,
        }
    else:
        new_state = get_empty_generation_state()

    btn_upd, regen_upd = get_generation_button_updates(
        remaining_absents,
        day_name,
        dept_filter,
        new_state
    )

    return tuple(ui) + (btn_upd, regen_upd, new_state)
    
def on_abs_t_change(df_state, abs_t, is_admin_logged_in):
    if not abs_t or df_state is None or df_state.empty:
        cb_update = gr.update(visible=False, value=False) if is_admin_logged_in else gr.update(visible=True, value=False)
        return gr.update(choices=[], value=None), gr.update(choices=[], value=None), cb_update
    
    abs_t_clean = clean_teacher_name_from_ui(abs_t)
    periods_elegant = []
    df_filtered = df_state[
        df_state["المعلم الغائب"].apply(lambda x: str(x).split(" (")[0].strip()) == abs_t_clean
    ]
    _, conflicted_slots = detect_conflicted_absence_slots(df_state.to_dict("records"))

    for _, row in df_filtered.iterrows():
        elegant_class = format_elegant_class(row['الصف'])
        is_admin_sup = str(row.get("المعلم البديل", "")) == "إشراف إداري"
        slot_key = (abs_t_clean, str(row['الحصة']).strip())

        if is_admin_sup:
            radar_icon = " 🚨 "
        elif slot_key in conflicted_slots:
            radar_icon = " 🔷 "
        else:
            radar_icon = " ✅ "

        display_text = f"الحصة {row['الحصة']} - ({elegant_class}){radar_icon}"
        periods_elegant.append(display_text)
        
    abs_dept = teachers_db.get(abs_t_clean, {}).get("dept", "عام")
    base_choice = f"نفس القسم ({abs_dept})"
    if is_admin_logged_in:
        choices = [base_choice, "التربية الإسلامية", "اللغة العربية", "الرياضيات", "العلوم", "اللغة الإنجليزية", "الدراسات الإجتماعية", "المهارات الفردية", "معلمو الصف", "الهيئة التدريسية", "الهيئة الإدارية"]
        filtered = [c for c in choices if not (c in OFFICIAL_DEPTS and c == abs_dept)]
        return gr.update(choices=periods_elegant, value=None), gr.update(choices=filtered, value=base_choice, interactive=True), gr.update(visible=False, value=False)
    else:
        return gr.update(choices=periods_elegant, value=None), gr.update(choices=[base_choice], value=base_choice, interactive=False), gr.update(visible=True, value=False)
        
def toggle_cross_dept(is_checked, abs_t):
    if not abs_t: return gr.update()
    abs_t_clean = clean_teacher_name_from_ui(abs_t)
    abs_dept = teachers_db.get(abs_t_clean, {}).get("dept", "عام")
    base_choice = f"نفس القسم ({abs_dept})"
    if is_checked:
        choices = [base_choice, "التربية الإسلامية", "اللغة العربية", "الرياضيات", "العلوم", "اللغة الإنجليزية", "الدراسات الإجتماعية", "المهارات الفردية", "معلمو الصف", "الهيئة التدريسية"]
        filtered = [c for c in choices if not (c in OFFICIAL_DEPTS and c == abs_dept)]
        return gr.update(choices=filtered, value=base_choice, interactive=True)
    else:
        return gr.update(choices=[base_choice], value=base_choice, interactive=False)


def update_available_subs_smart(abs_t, period, intervention_type, day_name, df_state, is_admin):
    choices, value, interactive = update_available_subs_smart_core(
        abs_t, period, intervention_type, day_name, df_state, is_admin
    )
    return gr.update(choices=choices, value=value, interactive=interactive)

def process_admin_action(df_state, abs_t, period, new_sub, day_name, dept_filter, is_admin_logged_in, current_abs, action_type, actor_name="", actor_role=""):
    result = process_admin_action_core(
        df_state,
        abs_t,
        period,
        new_sub,
        day_name,
        dept_filter,
        is_admin_logged_in,
        current_abs,
        action_type,
        actor_name=actor_name,
        actor_role=actor_role,
    )
    return refresh_ui_on_change(
        result["refresh_dept"],
        result["refresh_day"],
        result["refresh_is_admin"],
        current_abs=result.get("refresh_current_abs"),
    )
    
def load_teacher_data_for_edit(selected_teacher, is_admin=False, is_owner=False):
    permissions = get_permissions_from_flags(is_admin=is_admin, is_owner=is_owner)
    can_edit_vault = permissions["can_edit_vault_basic"]
    owner_mode = permissions["can_edit_sensitive_teacher_data"]

    if selected_teacher and selected_teacher in teachers_db: 
        dept = teachers_db[selected_teacher].get("dept", "عام")
        spec = teachers_db[selected_teacher].get("specialty", "")
        role = teachers_db[selected_teacher].get("role", "معلم")
        is_admin_staff = dept == "الهيئة الإدارية"
        is_spec_visible = dept in ["العلوم", "المهارات الفردية"]
        return (
            gr.update(value=dept, visible=not is_admin_staff),
            gr.update(value=teachers_db[selected_teacher].get("cover_count", 0), interactive=can_edit_vault),
            gr.update(value=teachers_db[selected_teacher].get("absent_count", 0), interactive=can_edit_vault),
            gr.update(value=teachers_db[selected_teacher].get("shortcoming_count", 0), interactive=can_edit_vault),
            gr.update(value=teachers_db[selected_teacher].get("phone", ""), interactive=owner_mode),
            gr.update(value=spec, visible=is_spec_visible and not is_admin_staff, interactive=owner_mode),
            gr.update(value=role, interactive=owner_mode)
        )

    return (
        gr.update(value="", visible=True),
        gr.update(value=0, interactive=can_edit_vault),
        gr.update(value=0, interactive=can_edit_vault),
        gr.update(value=0, interactive=can_edit_vault),
        gr.update(value="", interactive=owner_mode),
        gr.update(value="", interactive=owner_mode),
        gr.update(value="معلم", interactive=owner_mode)
    )
    
def toggle_specialty_visibility(dept): return gr.update(visible=dept in ["العلوم", "المهارات الفردية"])

def update_manual_count(name, new_val, new_abs_val, new_short_val, new_phone, new_specialty, new_role, dept_filter, day_val, df_state, abs_in_list, is_admin=False, is_owner=False, actor_name="", actor_role=""):
    raw = update_manual_count_core(
        name, new_val, new_abs_val, new_short_val, new_phone, new_specialty, new_role,
        dept_filter, day_val, df_state, abs_in_list,
        is_admin=is_admin,
        is_owner=is_owner,
        actor_name=actor_name,
        actor_role=actor_role,
    )
    return (
        gr.update(value=raw["balance"]),
        gr.update(value=raw["absences"]),
        gr.update(value=raw["shortcomings"]),
        gr.update(value=raw["day_overview"]),
        raw["message"],
        gr.update(**raw["abs_update"]),
        gr.update(**raw["teacher_update_1"]),
        gr.update(**raw["teacher_update_2"]),
    )

def delete_single_teacher(name, dept_filter, day_val, is_owner=False):
    raw = delete_single_teacher_core(name, dept_filter, day_val, is_owner=is_owner)
    return (
        gr.update(**raw["balance_update"]),
        gr.update(**raw["absences_update"]),
        gr.update(**raw["shortcomings_update"]),
        gr.update(**raw["day_overview_update"]),
        raw["message"],
        gr.update(**raw["abs_update"]),
        gr.update(**raw["teacher_update_1"]),
        gr.update(**raw["teacher_update_2"]),
        gr.update(**raw["delete_choices_update"]),
    )


def load_teacher_rules(t_name):
    t_key = resolve_teacher_key_from_ui(t_name)
    if t_key and t_key in teachers_db:
        info = teachers_db[t_key]
        slots = normalize_exempt_slots(info.get("exempt_slots", []))
        if slots and not info.get("exempt_days") and not info.get("exempt_periods"):
            # عرض الإعفاءات المحددة في الواجهة الحالية كأيام وحصص فريدة.
            # عند الحفظ ستُعاد كتابتها كضرب تلقائي للأيام × الحصص المختارة.
            slot_days = []
            slot_periods = []
            for slot in slots:
                if slot["day"] not in slot_days:
                    slot_days.append(slot["day"])
                if int(slot["period"]) not in slot_periods:
                    slot_periods.append(int(slot["period"]))
            return (gr.update(value=slot_days), gr.update(value=slot_periods))
        return (
            gr.update(value=info.get("exempt_days", [])),
            gr.update(value=info.get("exempt_periods", []))
        )
    return gr.update(value=[]), gr.update(value=[])

def save_teacher_rules(t_name, days, periods, actor_name="", actor_role="", is_admin=False, is_owner=False):
    message = save_teacher_rules_core(t_name, days, periods, actor_name, actor_role, is_admin, is_owner)
    return message, gr.update(value=render_exemptions_log_html())

def export_excel_report(dept_filter):
    data = []
    for t, d in teachers_db.items():
        effective_dept = resolve_effective_dept(dept_filter)
        if effective_dept == "الكل" or d.get("dept") == effective_dept:
            absence_dates_str = " \n ".join(d.get("absence_dates", [])) if d.get("absence_dates") else "-"
            data.append({
                "المعلم": format_teacher_name(t), 
                "المنصب": d.get("role", "معلم"), 
                "القسم": d.get("dept", "عام"),
                "التخصص الدقيق": d.get("specialty", "-"), 
                "رصيد الاحتياط": d.get("cover_count", 0),
                "مرات الغياب": d.get("absent_count", 0),
                "أيام وتواريخ الغياب": absence_dates_str,
                "حالات التقصير في الاحتياط": d.get("shortcoming_count", 0),
                "رقم الهاتف": d.get("phone", "")
            })
            
    df = pd.DataFrame(data).sort_values("رصيد الاحتياط", ascending=False) if data else pd.DataFrame(columns=["المعلم", "المنصب", "القسم", "التخصص الدقيق", "رصيد الاحتياط", "مرات الغياب", "أيام وتواريخ الغياب", "حالات التقصير في الاحتياط", "رقم الهاتف"])
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"تقرير_العدالة_والغياب_{timestamp}.xlsx"
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='تقرير المدرسة')
        worksheet = writer.sheets['تقرير المدرسة']
        worksheet.sheet_view.rightToLeft = True
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="004D40", end_color="004D40", fill_type="solid")
        for col in worksheet.columns:       # ← داخل الكتلة ✅
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill
            adjusted_width = min(max_length + 4, 40)
            worksheet.column_dimensions[column].width = adjusted_width
    return gr.update(value=filename)

def reset_monthly_balances(dept_filter, day_val, is_admin=False, is_owner=False, actor_name="", actor_role=""):
    raw = reset_monthly_balances_core(
        dept_filter,
        day_val,
        is_admin=is_admin,
        is_owner=is_owner,
        actor_name=actor_name,
        actor_role=actor_role,
    )
    return (
        gr.update(value=raw["balance"]),
        gr.update(value=raw["absences"]),
        gr.update(value=raw["shortcomings"]),
        gr.update(value=raw["day_overview"]),
        raw["message"],
    )
    
@state_locked
def clear_all_data(is_owner_logged_in):
    empty_balance_df = pd.DataFrame(columns=["المعلم", "الرصيد"])
    empty_absence_df = pd.DataFrame(columns=["المعلم", "مرات الغياب"])
    empty_shortcomings_html = render_compact_rtl_table_html(
        pd.DataFrame(columns=["المعلم", "حالات التقصير"]),
        "لا توجد حالات تقصير مسجلة للعرض."
    )
    empty_day_df = pd.DataFrame(columns=["المعلم"] + [f"ح {p}" for p in range(1, MAX_PERIODS + 1)])
    empty_day_html, _, _, _ = render_day_table_html(empty_day_df, 0, PAGE_SIZE)
    empty_generation_state = get_empty_generation_state()

    if not is_owner_logged_in:
        return (
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            "<div style='color:red; font-weight:bold;'>❌ هذه العملية متاحة لمالك النظام فقط.</div>",
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update(),
            gr.update()
        )

    teachers_db.clear()
    daily_db.clear()
    processed_absences.clear()
    last_assigned_teachers.clear()

    # الحفظ الآمن ينشئ نسخة احتياطية قبل كتابة الحالة الفارغة.
    save_db()
    save_daily_db()

    return (
        gr.update(choices=["الكل"] + OFFICIAL_DEPTS, value="الكل"),
        gr.update(choices=[], value=[]),
        gr.update(choices=[], value=None),
        gr.update(choices=[], value=None),
        gr.update(value=empty_balance_df),
        gr.update(value=empty_absence_df),
        gr.update(value=empty_shortcomings_html),
        gr.update(value=empty_day_df),
        gr.update(value=empty_day_html, visible=True),
        gr.update(visible=False),
        gr.update(interactive=False),
        gr.update(interactive=False),
        gr.update(value="", visible=False),
        0,
        "<div style='color:orange; font-weight:bold;'>⚠️ تم تصفير المنظومة بالكامل ومسح بقايا الواجهة السابقة.</div>",
        gr.update(choices=[], value=None),
        gr.update(value=""),
        gr.update(value="", visible=False),
        gr.update(value=None),
        pd.DataFrame(),
        empty_generation_state,
        gr.update(value=""),
        gr.update(value="<div style='text-align:center; color:gray; padding:20px;'>لا توجد تكليفات لعرضها</div>"),
        get_initial_header(),
        "",
        gr.update(choices=[], value=None),
        gr.update(choices=[], value=None),
        gr.update(interactive=False),
        gr.update(visible=False, interactive=False),
        gr.update(interactive=False),
        gr.update(interactive=False)
    )  

# Phase 3D: دوال الصلاحيات المركزية get_permissions انتقلت إلى auth.py.


def get_day_dept_filter_update(role="", dept_value="", is_owner=False, is_admin=False, is_shared_teacher=False):
    """
    فلتر مستقل خاص بتبويب جدول اليوم فقط.
    لا يوسع dept_in المشترك حتى لا تتسرب صلاحية "الكل" إلى التوزيع أو الأرصدة أو الإعفاءات.
    """
    effective_dept = resolve_effective_dept(dept_value)

    if bool(is_shared_teacher):
        return gr.update(choices=["الكل"], value="الكل", interactive=False)

    if bool(is_owner) or bool(is_admin):
        choices = ["الكل"] + OFFICIAL_DEPTS
        value = effective_dept if effective_dept in choices else "الكل"
        return gr.update(choices=choices, value=value, interactive=True)

    dept_clean = str(effective_dept or "").strip()
    if dept_clean and dept_clean != "الكل":
        return gr.update(choices=[dept_clean, "الكل"], value=dept_clean, interactive=True)

    return gr.update(choices=["الكل"], value="الكل", interactive=True)


def attempt_login(pin, day_val):
    load_db()
    load_daily_db()

    pin = str(pin or "").strip()
    login_account_id, user_info, auth_error = authenticate_login_pin(pin)

    if user_info:
        role = user_info.get("role", "")
        dept = user_info.get("dept", "الكل")
        if role == "مستخدم عام":
            dept = "المعلمون"

        name = user_info.get("name", "")
        session_display_name = get_account_session_display_name(user_info)
        is_owner = bool(
            user_info.get("is_owner", False)
            or role == "صاحب النظام"
        )

        ui_vis = get_ui_visibility_updates(pin, role, is_owner)
        is_shared_teacher = ui_vis["is_shared_teacher"]

        effective_dept = resolve_effective_dept(dept)
        dept_for_ui = effective_dept
        is_admin = bool(ui_vis["is_admin"])

        temporary_note = ""
        if bool(user_info.get("must_change_pin", False)) and not is_shared_teacher:
            temporary_note = (
                "<div style='margin-top:8px;color:#fff;background:#b45309;"
                "padding:7px;border-radius:7px;font-size:14px;'>"
                "تنبيه: رمز الدخول مؤقت. غيّره من قسم «تغيير رمز دخولي»."
                "</div>"
            )

        welcome_msg = render_account_welcome_html(user_info, temporary_note)

        if is_admin:
            up_dept_update = gr.update(interactive=True)
            manual_entry_visibility = gr.update(visible=is_owner)
        else:
            up_dept_update = gr.update(value=None, interactive=False)
            manual_entry_visibility = gr.update(visible=False)

        updates = refresh_ui_on_change(
            dept_for_ui,
            day_val,
            is_admin,
        )

        return [
            gr.update(visible=False),
            gr.update(visible=True),
            welcome_msg,
            gr.update(visible=not is_shared_teacher),
            gr.update(
                choices=["الكل"] + OFFICIAL_DEPTS,
                value=dept_for_ui,
                interactive=is_admin,
            ),
            get_day_dept_filter_update(role, dept, is_owner, is_admin, is_shared_teacher),
            gr.update(value=""),
            up_dept_update,
            manual_entry_visibility,
            is_admin,
            is_owner,
            session_display_name,
            role,
            login_account_id,
        ] + list(updates) + [
            gr.update(
                visible=(
                    dept_for_ui in ["العلوم", "المهارات الفردية"]
                    and not is_shared_teacher
                )
            ),
            gr.update(visible=ui_vis["can_clear_system"]),
            gr.update(visible=ui_vis["school_data_tab"]),
            gr.update(visible=ui_vis["controls_row"]),
            gr.update(visible=ui_vis["exemptions_tab"]),
            gr.update(visible=ui_vis["distribution_tab"]),
            gr.update(visible=ui_vis["balances_tab"]),
            gr.update(visible=ui_vis["swap_tab"]),
            gr.update(visible=ui_vis["day_tab"]),
            gr.update(visible=ui_vis["teacher_tab"]),
            gr.update(visible=ui_vis["swap_excel_btn"]),
        ]

    if auth_error == "disabled":
        error_text = "هذا الحساب معطل. راجع مالك النظام."
    else:
        error_text = "رمز الدخول غير صحيح! حاول مرة أخرى."

    gr.Warning(f"❌ {error_text}")
    error_updates = [gr.update()] * 27

    return [
        gr.update(),
        gr.update(),
        (
            "<div style='color:red;text-align:center;font-weight:bold;"
            "margin-top:10px;'>"
            f"❌ {html_lib.escape(error_text)}</div>"
        ),
        gr.update(),
        gr.update(),
        gr.update(),
        gr.update(),
        gr.update(),
        gr.update(),
        False,
        False,
        "",
        "",
        "",
    ] + error_updates + [
        gr.update(),
        gr.update(),
        gr.update(),
        gr.update(),
        gr.update(),
        gr.update(),
        gr.update(),
        gr.update(),
        gr.update(),
        gr.update(),
        gr.update(),
    ]


def do_logout(): 
    return (
        gr.update(visible=True),
        gr.update(visible=False),
        "",
        gr.update(visible=True),
        gr.update(choices=["الكل"] + OFFICIAL_DEPTS, value="الكل"),
        gr.update(choices=["الكل"] + OFFICIAL_DEPTS, value="الكل", interactive=True),
        False,
        False,
        "",
        "",
        "",
        None,
        None,
        gr.update(visible=False, value=False),
        gr.update(visible=False),
        gr.update(visible=True),
        gr.update(visible=False),
        gr.update(visible=False),
        gr.update(visible=False),
        gr.update(visible=True),
        gr.update(visible=True),
        gr.update(visible=True),
        gr.update(visible=True),
        get_empty_generation_state(),
        {}
    )
    
# CSS extracted in v1.8.5a-css-extraction.
# Keep visual/RTL/Gradio fixes in masar_styles.css, not inside app.py.
def load_masar_css():
    css_path = os.path.join(APP_DIR, "masar_styles.css")
    try:
        with open(css_path, "r", encoding="utf-8") as css_file:
            return css_file.read()
    except FileNotFoundError:
        print(f"WARNING: masar_styles.css not found at {css_path}; running without custom CSS.")
        return ""

css = load_masar_css()


def apply_school_theme_to_css(css_text):
    """تطبيق ألوان الهوية المحفوظة عند تشغيل التطبيق."""
    themed = str(css_text)
    replacements = {
        "#004d40": THEME_COLOR,
        "#00695c": THEME_COLOR_2,
        "#ffca28": ACCENT_COLOR,
    }
    for old_color, new_color in replacements.items():
        themed = themed.replace(old_color, new_color)
        themed = themed.replace(old_color.upper(), new_color)
    return themed

css = apply_school_theme_to_css(css)

js_code = """
function() {
    let isRunning = false;

    const removeDark = () => {
        if (isRunning) return;

        const html = document.documentElement;
        const body = document.body;
        if (!html || !body) return;

        const needsFix =
            html.classList.contains('dark') ||
            body.classList.contains('dark') ||
            html.getAttribute('data-theme') === 'dark';

        if (!needsFix) return;

        isRunning = true;
        html.classList.remove('dark');
        body.classList.remove('dark');
        html.setAttribute('data-theme', 'light');
        body.style.backgroundColor = '#ffffff';
        setTimeout(() => { isRunning = false; }, 100);
    };

    removeDark();

    const observer = new MutationObserver(removeDark);
    observer.observe(document.documentElement, {
        attributes: true,
        attributeFilter: ['class', 'data-theme']
    });

    if (document.body) {
        observer.observe(document.body, {
            attributes: true,
        });
    }
}


/* v44-AV — تفاعل لمس شعار الدخول */
function setupMainLogoTouchInteraction() {
    const logo = document.getElementById('main-logo');
    if (logo && !logo.dataset.touchReady) {
        logo.dataset.touchReady = '1';
        logo.addEventListener('touchmove', function(e) {
            const touch = e.touches[0];
            const rect = logo.getBoundingClientRect();
            const x = touch.clientX - rect.left - rect.width/2;
            const y = touch.clientY - rect.top - rect.height/2;
            const rotX = -(y / rect.height) * 30;
            const rotY = (x / rect.width) * 30;
            logo.style.transform = `perspective(400px) rotateX(${rotX}deg) rotateY(${rotY}deg) scale(1.12)`;
            logo.style.animation = 'none';
        }, { passive: true });
        logo.addEventListener('touchend', function() {
            logo.style.transform = '';
            logo.style.animation = 'logo4d 4s ease-in-out infinite';
        });
    }
}

setTimeout(setupMainLogoTouchInteraction, 300);
setTimeout(setupMainLogoTouchInteraction, 1000);
setTimeout(setupMainLogoTouchInteraction, 2500);
document.addEventListener('DOMContentLoaded', setupMainLogoTouchInteraction);


"""

header_html = build_header_html()

def filter_swap_teachers_safe(dept):
    choices, value = filter_swap_teachers_safe_core(dept)
    return gr.update(choices=choices, value=value)

def get_teacher_periods_safe(t, d):
    choices, value = get_teacher_periods_safe_core(t, d)
    return gr.update(choices=choices, value=value)

def get_teacher_periods_marked(t, d, confirmed_state, current_value=None):
    choices, selected_value = get_teacher_periods_marked_core(t, d, confirmed_state, current_value)
    return gr.update(choices=choices, value=selected_value)

def run_radar_safe(t, p, d):
    default_msg = "💡 يرجى اختيار أحد المعلمين من القائمة بالأعلى لتوليد مسودة رسالة الواتساب هنا..."
    candidates = run_radar_safe_core(t, p, d)
    return gr.update(choices=candidates, value=None), gr.update(value=default_msg), gr.update(value="")

def generate_wa_msg(choice, t_req, p_req, d_req):
    msg, btn_html = generate_wa_msg_core(choice, t_req, p_req, d_req)
    return gr.update(value=msg), gr.update(value=btn_html)

def get_swap_candidates_for_period(t, period_value, d, confirmed_state):
    candidates, saved_choice, saved_message, btn_value, confirm_interactive = get_swap_candidates_for_period_core(
        t, period_value, d, confirmed_state
    )
    return (
        gr.update(choices=candidates, value=saved_choice, visible=True),
        gr.update(value=saved_message, visible=True),
        gr.update(value=btn_value, visible=True),
        gr.update(visible=True, interactive=confirm_interactive)
    )

def on_swap_option_selected(choice, t, period_value, d):
    msg_value, btn_value, is_interactive = on_swap_option_selected_core(choice, t, period_value, d)
    return (
        gr.update(value=msg_value, visible=True),
        gr.update(value=btn_value, visible=True),
        gr.update(visible=True, interactive=is_interactive)
    )

def on_swap_option_selected_from_event(choice_current, t, period_value, d, evt: gr.SelectData):
    choice = None

    if evt is not None:
        try:
            choice = evt.value
        except Exception:
            choice = None

    if not choice:
        choice = choice_current

    if not choice:
        try:
            choice = str(evt) if evt else None
        except Exception:
            choice = None

    return on_swap_option_selected(choice, t, period_value, d)

def get_leader_action_button_updates(abs_teacher, period_value=None, substitute_teacher=None):
    has_absent = bool(clean_teacher_name_from_ui(abs_teacher))
    clean_period = extract_clean_period_number(period_value) if period_value else ""
    has_period = bool(str(clean_period or "").strip())
    sub_clean = str(substitute_teacher or "").strip()
    valid_substitute = bool(
        sub_clean
        and sub_clean not in ["ℹ️ اختر الحصة أولاً", "❌ لا يوجد بديل مناسب", "اختر الحصة أولاً"]
        and not sub_clean.startswith("❌")
    )
    has_substitute = bool(has_absent and has_period and valid_substitute)

    return (
        gr.update(interactive=has_substitute),              # تكليف احتياط رسمي
        gr.update(interactive=has_substitute),              # اعتماد كتبادل
        gr.update(interactive=has_absent and has_period),   # رصد تقصير في التكليف
        gr.update(interactive=has_absent),                  # التراجع عن غياب اليوم بالكامل
    )


def show_home_dashboard_js():
    return """() => {
        const home = document.getElementById('masar_home_dashboard');
        const tabsBox = document.getElementById('masar_tabs_container');
        if (home) home.style.setProperty('display', 'block', 'important');
        if (tabsBox) tabsBox.style.setProperty('display', 'none', 'important');

        const centerVisibleOrphanCards = () => {
            const grid = document.querySelector('#masar_home_dashboard .masar-card-grid');
            if (!grid) return;

            const cards = Array.from(grid.querySelectorAll('.masar-card'));
            cards.forEach(card => card.style.removeProperty('grid-column'));

            if (window.innerWidth < 769) return;

            const visibleCards = cards.filter(card => {
                const st = window.getComputedStyle(card);
                return st.display !== 'none' && st.visibility !== 'hidden' && card.offsetParent !== null;
            });

            if (visibleCards.length > 1 && visibleCards.length % 3 === 1) {
                visibleCards[visibleCards.length - 1].style.setProperty('grid-column', '2 / span 1', 'important');
            }
        };

        setTimeout(centerVisibleOrphanCards, 80);
        setTimeout(centerVisibleOrphanCards, 300);
        setTimeout(centerVisibleOrphanCards, 700);

        if (!window.__masarCenterCardsResizeReady) {
            window.__masarCenterCardsResizeReady = true;
            window.addEventListener('resize', centerVisibleOrphanCards, { passive: true });
        }

        window.scrollTo({ top: 0, behavior: 'smooth' });
    }"""

def return_home_dashboard_js():
    return """() => {
        const home = document.getElementById('masar_home_dashboard');
        const tabsBox = document.getElementById('masar_tabs_container');
        if (tabsBox) tabsBox.style.setProperty('display', 'none', 'important');
        if (home) home.style.setProperty('display', 'block', 'important');

        const centerVisibleOrphanCards = () => {
            const grid = document.querySelector('#masar_home_dashboard .masar-card-grid');
            if (!grid) return;

            const cards = Array.from(grid.querySelectorAll('.masar-card'));
            cards.forEach(card => card.style.removeProperty('grid-column'));

            if (window.innerWidth < 769) return;

            const visibleCards = cards.filter(card => {
                const st = window.getComputedStyle(card);
                return st.display !== 'none' && st.visibility !== 'hidden' && card.offsetParent !== null;
            });

            if (visibleCards.length > 1 && visibleCards.length % 3 === 1) {
                visibleCards[visibleCards.length - 1].style.setProperty('grid-column', '2 / span 1', 'important');
            }
        };

        setTimeout(centerVisibleOrphanCards, 80);
        setTimeout(centerVisibleOrphanCards, 300);
        setTimeout(centerVisibleOrphanCards, 700);

        if (!window.__masarCenterCardsResizeReady) {
            window.__masarCenterCardsResizeReady = true;
            window.addEventListener('resize', centerVisibleOrphanCards, { passive: true });
        }

        window.scrollTo({ top: 0, behavior: 'smooth' });
    }"""

def select_tab_js(label_text, tab_index):
    safe_label = str(label_text).replace("\\", "\\\\").replace("'", "\\'")
    try:
        safe_index = int(tab_index)
    except Exception:
        safe_index = 0

    return f"""() => {{
        const home = document.getElementById('masar_home_dashboard');
        const tabsBox = document.getElementById('masar_tabs_container');

        // أخفِ اللوحة الرئيسية فقط.
        if (home) home.style.setProperty('display', 'none', 'important');

        // أظهر حاوية التبويبات أولًا حتى تكون عناصرها موجودة للبحث والنقر.
        if (tabsBox) {{
            tabsBox.style.setProperty('display', 'block', 'important');
            tabsBox.style.setProperty('visibility', 'hidden', 'important');

            // مهم: لا تترك شريط التبويبات display:none، لأن JS لن يضغطه بثبات.
            const tabLists = tabsBox.querySelectorAll('.tab-nav, [role="tablist"]');
            tabLists.forEach(el => {{
                el.style.setProperty('display', 'flex', 'important');
                el.style.setProperty('position', 'absolute', 'important');
                el.style.setProperty('width', '1px', 'important');
                el.style.setProperty('height', '1px', 'important');
                el.style.setProperty('overflow', 'hidden', 'important');
                el.style.setProperty('opacity', '0', 'important');
                el.style.setProperty('pointer-events', 'none', 'important');
            }});
        }}

        setTimeout(() => {{
            const selectors = [
                '#masar_tabs_container [role="tab"]',
                '#masar_tabs_container .tab-nav button',
                '#masar_tabs_container button[aria-controls]',
                '#masar_tabs_container button',
                '.masar-tabs-container [role="tab"]',
                '.masar-tabs-container .tab-nav button',
                '.masar-tabs-container button[aria-controls]'
            ];

            let tabs = [];
            selectors.forEach(sel => {{
                document.querySelectorAll(sel).forEach(el => {{
                    const txt = (el.textContent || '').trim();
                    if (txt && !tabs.includes(el)) tabs.push(el);
                }});
            }});

            const wanted = '{safe_label}';
            let target = tabs.find(btn => (btn.textContent || '').includes(wanted));

            if (!target && tabs.length > {safe_index}) {{
                target = tabs[{safe_index}];
            }}

            if (target) {{
                target.click();
                target.dispatchEvent(new MouseEvent('click', {{ bubbles: true, cancelable: true, view: window }}));
            }}

            setTimeout(() => {{
                if (tabsBox) {{
                    tabsBox.style.setProperty('display', 'block', 'important');
                    tabsBox.style.setProperty('visibility', 'visible', 'important');
                }}
                window.scrollTo({{ top: 0, behavior: 'smooth' }});
            }}, 160);
        }}, 100);
    }}"""


def show_home_dashboard_after_login(dept_value, is_admin, is_owner, role=""):
    permissions = get_permissions(
        role=role,
        is_owner=is_owner,
        dept_value=dept_value,
        is_admin_flag=is_admin
    )

    return (
        gr.update(),               # home_dashboard controlled by JS/CSS
        gr.update(),               # tabs_container controlled by JS/CSS
        gr.update(visible=permissions["can_view_distribution"]),
        gr.update(visible=permissions["can_view_balances"]),
        gr.update(visible=permissions["can_manage_exemptions"]),
        gr.update(visible=permissions["can_view_swap"]),
        gr.update(visible=permissions["can_view_day_table"]),
        gr.update(visible=permissions["can_view_teacher_table"]),
        gr.update(visible=permissions["can_access_school_data"]),
    )

def open_home_section(tab_id=None):
    return (
        gr.update(),  # home_dashboard controlled by JS
        gr.update(),  # tabs_container controlled by JS
        gr.update(selected=tab_id or "distribution"),  # main_tabs selected by Gradio
    )

def return_to_home_dashboard():
    return (
        gr.update(),  # home_dashboard controlled by JS
        gr.update(),  # tabs_container controlled by JS
    )

def show_selected_tab_container_js():
    return """() => {
        const home = document.getElementById('masar_home_dashboard');
        const tabsBox = document.getElementById('masar_tabs_container');

        if (home) home.style.setProperty('display', 'none', 'important');

        if (tabsBox) {
            tabsBox.style.setProperty('display', 'block', 'important');
            tabsBox.style.setProperty('visibility', 'visible', 'important');

            const tabLists = tabsBox.querySelectorAll('.tab-nav, [role="tablist"]');
            tabLists.forEach(el => {
                el.style.setProperty('display', 'flex', 'important');
                el.style.setProperty('position', 'absolute', 'important');
                el.style.setProperty('width', '1px', 'important');
                el.style.setProperty('height', '1px', 'important');
                el.style.setProperty('overflow', 'hidden', 'important');
                el.style.setProperty('opacity', '0', 'important');
                el.style.setProperty('pointer-events', 'none', 'important');
            });
        }

        window.scrollTo({ top: 0, behavior: 'smooth' });
    }"""

# ================================================================
# واجهة Gradio الرئيسية — كل شيء داخل كتلة واحدة
# ================================================================
with gr.Blocks() as app:
    current_user_is_admin = gr.State(value=False)
    current_user_is_owner = gr.State(value=False)
    current_user_name = gr.State(value="")
    current_user_role = gr.State(value="")
    current_user_account_id = gr.State(value="")
    current_schedule_state = gr.State()
    reserve_generation_state = gr.State(value=get_empty_generation_state())

    with gr.Column(visible=True, elem_classes="login-box") as login_container:
        login_branding_html = gr.HTML(value=build_login_branding_html())

        pin_input = gr.Textbox(type="password", show_label=False, placeholder="Enter ثم اضغط (PIN) 🔑 أدخل رمز الدخول", text_align="center")
        login_btn = gr.Button("تسجيل الدخول", elem_classes="admin-btn")
        login_msg = gr.HTML()
        login_credits_html = gr.HTML(value=build_login_credits_html())

    with gr.Column(visible=False) as main_app_container:
        header_branding_html = gr.HTML(value=build_header_html())
        
        with gr.Row(elem_classes="top-user-row"):
            with gr.Column(scale=5, elem_classes="welcome-col"):
                welcome_html = gr.HTML(elem_classes="welcome-html-box")
            with gr.Column(scale=1, min_width=120, elem_classes="logout-col"):
                logout_btn = gr.Button("🚪 خروج و إقفال", elem_classes=["reset-btn", "logout-btn"])
        
        with gr.Accordion("🔑 تغيير رمز دخولي", open=False, elem_classes=["masar-accordion-arrow-fix", "self-pin-accordion-card"]) as self_pin_accordion:
            with gr.Column(elem_classes="self-pin-card"):
                gr.HTML(
                    "<div class='self-pin-card-head'>"
                    "<div class='self-pin-card-icon'>🔐</div>"
                    "<div>"
                    "<div class='self-pin-card-title'>تحديث رمز الدخول</div>"
                    "<div class='self-pin-card-desc'>اكتب رمزك الحالي ثم الرمز الجديد، ثم اضغط حفظ الرمز الجديد.</div>"
                    "</div>"
                    "</div>"
                    "<div class='self-pin-card-note'>رمز مالك النظام يُدار من Secret الاستضافة.</div>"
                )
                with gr.Row(elem_classes="self-pin-row-ltr"):
                    self_current_pin = gr.Textbox(
                        type="password",
                        label="الرمز الحالي",
                    )
                    self_new_pin = gr.Textbox(
                        type="password",
                        label="الرمز الجديد",
                    )
                    self_confirm_pin = gr.Textbox(
                        type="password",
                        label="تأكيد الرمز الجديد",
                    )
                self_change_pin_btn = gr.Button(
                    "حفظ الرمز الجديد",
                    elem_classes="admin-btn",
                )
                self_change_pin_status = gr.HTML()

        with gr.Column(visible=True, elem_id="masar_home_dashboard", elem_classes="masar-home-dashboard") as home_dashboard:
            home_hero_html = gr.HTML(value=build_home_hero_html())

            with gr.Row(elem_classes="masar-card-grid"):
                with gr.Column(visible=True, elem_classes="masar-card") as card_distribution:
                    gr.HTML("<div class='masar-card-icon'>📋</div><div class='masar-card-title'>التوزيع والاحتياط</div><div class='masar-card-desc'>إدارة الغياب وتوزيع حصص الاحتياط.</div>")
                    btn_open_distribution = gr.Button("دخول القسم ←", elem_classes="masar-card-btn")

                with gr.Column(visible=True, elem_classes="masar-card") as card_balances:
                    gr.HTML("<div class='masar-card-icon'>⚖️</div><div class='masar-card-title'>الأرصدة والتقارير</div><div class='masar-card-desc'>متابعة الأرصدة، الغياب، والتقارير.</div>")
                    btn_open_balances = gr.Button("دخول القسم ←", elem_classes="masar-card-btn")

                with gr.Column(visible=True, elem_classes="masar-card") as card_exemptions:
                    gr.HTML("<div class='masar-card-icon'>🛡️</div><div class='masar-card-title'>حالات الإعفاء</div><div class='masar-card-desc'>تثبيت ومراجعة إعفاءات المعلمين.</div>")
                    btn_open_exemptions = gr.Button("دخول القسم ←", elem_classes="masar-card-btn")

                with gr.Column(visible=True, elem_classes="masar-card") as card_swap:
                    gr.HTML("<div class='masar-card-icon'>🤝</div><div class='masar-card-title'>التبادل الودي الأسبوعي</div><div class='masar-card-desc'>تنظيم التبادلات بين المعلمين الحاضرين.</div>")
                    btn_open_swap = gr.Button("دخول القسم ←", elem_classes="masar-card-btn")

                with gr.Column(visible=True, elem_classes="masar-card") as card_day:
                    gr.HTML("<div class='masar-card-icon'>📅</div><div class='masar-card-title'>جدول اليوم</div><div class='masar-card-desc'>عرض جدول اليوم الدراسي حسب القسم.</div>")
                    btn_open_day = gr.Button("دخول القسم ←", elem_classes="masar-card-btn")

                with gr.Column(visible=True, elem_classes="masar-card") as card_teacher:
                    gr.HTML("<div class='masar-card-icon'>🔍</div><div class='masar-card-title'>جدول المعلم</div><div class='masar-card-desc'>عرض الجدول الأسبوعي للمعلم.</div>")
                    btn_open_teacher = gr.Button("دخول القسم ←", elem_classes="masar-card-btn")

                with gr.Column(visible=True, elem_classes="masar-card") as card_school_data:
                    gr.HTML("<div class='masar-card-icon'>🏫</div><div class='masar-card-title'>مركز البيانات المدرسية</div><div class='masar-card-desc'>إدارة الملفات المرجعية للمنظومة.</div>")
                    btn_open_school_data = gr.Button("دخول القسم ←", elem_classes="masar-card-btn")

        with gr.Column(visible=True, elem_id="masar_tabs_container", elem_classes="masar-tabs-container") as tabs_container:
            btn_back_home = gr.Button("🏠 العودة للوحة الرئيسية", elem_classes="home-back-btn")
            with gr.Row(elem_classes="yellow-box") as controls_row:
                dept_in = gr.Dropdown(["الكل"] + OFFICIAL_DEPTS, label="📂 مركز التحكم", value="الكل", scale=2, elem_classes="masar-arrow-fix")
                day_in = gr.Dropdown(SCHOOL_WEEK_DAYS, label="📅 اختر اليوم الدراسي", value=get_current_day_oman(), scale=2, elem_classes="masar-arrow-fix")
                refresh_btn = gr.Button("🔄 تحديث الشاشة والبيانات", elem_classes="refresh-btn", scale=1)

            with gr.Tabs(selected="distribution") as main_tabs:
                with gr.Tab("📋 التوزيع والاحتياط", id="distribution") as distribution_tab:
                    with gr.Column():
                        with gr.Column(elem_classes=["yellow-box", "distribution-controls-card"]):
                            gr.HTML("<div class='distribution-controls-title'>⚙️ ضوابط التوزيع اليومية</div>")
                            max_reserves_input = gr.Number(value=1, label="🛑 الحد الأقصى للاحتياط لكل معلم في اليوم", precision=0)
                    
                        radar_warning_html = gr.HTML()
                        gr.HTML("""
                        <div class='reserve-guide-box'>
                            💡 <b>طريقة الاستخدام:</b><br>
                            ⚫️ حدد المعلمين الغائبين، ثم اضغط <b>توليد وتوزيع الاحتياط</b>.<br>
                            ⚫️ عند إضافة معلم غائب جديد، اضغط الزر نفسه لإضافة حصصه فوق التوزيع السابق دون تغييره، أو استخدم <b>إعادة توليد من جديد</b> لمسح محتوى اليوم وإعادة بنائه من الصفر.<br>
                            ⚫️ استخدم <b>مقترح آخر</b> لعرض توزيع آلي بديل للغيابات الحالية.
                        </div>
                        """)
                        abs_in = gr.Dropdown([], label="حدد المعلمين الغائبين", multiselect=True, elem_classes=["absent-box", "masar-arrow-fix", "masar-field-label-right"])

                        with gr.Row():
                            btn = gr.Button("توليد وتوزيع الاحتياط", variant="primary", interactive=False, elem_classes="action-btn")
                            btn_regenerate = gr.Button("🔁 إعادة توليد من جديد", visible=False, interactive=False, elem_classes="regen-btn")
                            btn_alt = gr.Button("مقترح آخر", interactive=False, elem_classes="action-btn")
                            btn_img = gr.Button("🖼️ تحميل الجدول كصورة", interactive=False, elem_classes="export-btn")

                        date_display = gr.HTML(get_initial_header)
                        img_out = gr.Image(label="الصورة الجاهزة للنسخ", interactive=False, elem_classes="masar-image-label-right")
                        tbl_out = gr.HTML(value="")
                    
                        with gr.Column(elem_classes="whatsapp-box"):
                            gr.Markdown("## 📱 مركز التواصل الذكي ومهام الواتساب")
                            with gr.Row(): msg_summary = gr.Textbox(label="📊 تقرير الجروب الإداري", lines=4, interactive=True)
                            with gr.Row(): msg_individual_html = gr.HTML(label="💌 بطاقات التكليف الفردية")

                        gr.HTML("<div class='external-section-title leader-title'>⚙️ لوحة القائد: التعديل اليدوي والتبادل</div>")
                        with gr.Accordion("فتح / إغلاق لوحة القائد", open=False, elem_classes=["leader-accordion", "masar-accordion-arrow-fix", "leader-outer-card"]):
                            with gr.Column(elem_classes="admin-zone"):
                                admin_zone_title = gr.HTML("<h4 style='color:#004d40; text-align:center; margin-top:0;'>🛠️ غرفة العمليات والقيادة</h4>")
                                admin_zone_help = gr.HTML("<div style='color:#00695c; background:#e0f2f1; padding:15px; border-radius:8px; border-right: 4px solid #00897b;'>💡 <b>توضيح:</b> اختر المعلم الغائب ثم الحصة، وبعدها نفّذ الإجراء المناسب من نفس اللوحة حسب دورك وصلاحيتك.</div>")
                            
                                with gr.Row(elem_classes="leader-row-rtl"):
                                    edit_abs_t = gr.Dropdown([], label="‏1️⃣ المعلم الغائب", allow_custom_value=True, elem_classes=["masar-arrow-fix", "masar-field-label-right"])
                                    edit_period = gr.Dropdown([], label="‏2️⃣ اختر الحصة", allow_custom_value=False, elem_classes=["masar-arrow-fix", "masar-field-label-right"])
                                    edit_intervention_type = gr.Dropdown([], label="‏3️⃣ نطاق البحث عن بديل (تلقائي ذكي)", allow_custom_value=True, elem_classes=["masar-arrow-fix", "masar-field-label-right"])
                            
                                with gr.Row():
                                    cb_cross_dept = gr.Checkbox(label="🔓 تفعيل التعاون مع قسم آخر 🤝", visible=False)
                            
                                with gr.Row(elem_classes="leader-row-rtl"):
                                    edit_new_sub = gr.Dropdown([], label="‏4️⃣ البديل المنقذ", allow_custom_value=True, elem_classes=["masar-arrow-fix", "masar-field-label-right"])
                                with gr.Row():
                                    btn_apply_override = gr.Button("✍🏻 تكليف احتياط رسمي", elem_classes=["admin-btn", "leader-official-btn"], interactive=False)
                                    btn_apply_tabadul = gr.Button("🤝 اعتماد كـ تبادل", elem_classes=["tabadul-btn", "leader-swap-btn"], interactive=False)
                                    btn_apply_penalty = gr.Button("🚨 رصد تقصير في التكليف", elem_classes=["reset-btn", "leader-penalty-btn"], interactive=False)
                            
                                with gr.Row():
                                    btn_cancel_absence = gr.Button("⏪ إلغاء غياب اليوم بالكامل", elem_classes=["reset-btn", "leader-cancel-btn"], interactive=False)

                with gr.Tab("⚖️ الأرصدة والتقارير", id="balances") as balances_tab:
                    monthly_status = gr.HTML()

                    with gr.Row(elem_classes="yellow-box"):
                        export_btn = gr.Button("📥 تصدير تقرير المدرسة (Excel)", elem_classes="export-btn")
                        reset_month_btn = gr.Button("🔄 إقفال الشهر (تصفير الأرصدة فقط)", elem_classes="reset-btn", visible=False)
                
                    report_file = gr.File(label="📥 التقرير الجاهز للتحميل")

                    with gr.Row():
                        with gr.Column():
                            gr.HTML("<h3 style='text-align:center; color:#004d40; font-size: 1.3em; font-weight: 900; margin-bottom: 10px;'>🟢 رصيد الاحتياط</h3>")
                            tbl_bal = gr.HTML()
                        with gr.Column():
                            gr.HTML("<h3 style='text-align:center; color:#c62828; font-size: 1.3em; font-weight: 900; margin-bottom: 10px;'>🔴 حصر الغياب</h3>")
                            tbl_abs = gr.HTML()

                    with gr.Row():
                        with gr.Column():
                            gr.HTML("<h3 style='text-align:center; color:#b45309; font-size: 1.3em; font-weight: 900; margin-bottom: 10px;'>🟠 حالات التقصير</h3>")
                            tbl_short = gr.HTML()
                
                    gr.HTML("<div class='external-section-title vault-title'>🔒 الخزنة: تعديل يدوي للأرصدة والهواتف</div>")
                    with gr.Accordion("فتح / إغلاق الخزنة", open=False, elem_classes=["yellow-box", "vault-accordion", "masar-accordion-arrow-fix"]):
                        gr.HTML("""
                        <div class='vault-guide-box'>
                            💡 <b>توضيح:</b><br>
                            ⚫️ اختر المعلم من القائمة لعرض بياناته الحالية.<br>
                            ⚫️ يمكن تعديل رصيد الاحتياط، مرات الغياب، وحالات التقصير عند الحاجة، ثم اضغط <b>حفظ التعديلات</b>.
                        </div>
                        """)
                        with gr.Row():
                            t_name = gr.Dropdown(list(teachers_db.keys()), label="المعلم", elem_classes="masar-arrow-fix")
                            t_dept_edit = gr.Textbox(label="القسم / المادة (للعرض فقط)", interactive=False)
                            t_role_edit = gr.Dropdown(ALL_ROLES, label="المنصب الإشرافي", interactive=False, elem_classes="masar-arrow-fix")
                        with gr.Row():
                            t_phone_edit = gr.Textbox(label="رقم الهاتف (الواتساب)", interactive=False)
                            t_specialty_edit = gr.Dropdown(
                                choices=["فيزياء", "كيمياء", "أحياء", "تقنية المعلومات",
                                         "الفنون التشكيلية", "الرياضة المدرسية",
                                         "المهارات الحياتية", "المهارات الموسيقية"],
                                label="التخصص الدقيق",
                                visible=False,
                                interactive=False,
                                allow_custom_value=True
                            )
                        with gr.Row():
                            t_val = gr.Number(label="رصيد الاحتياط")
                            t_abs_val = gr.Number(label="مرات الغياب")
                            t_short_val = gr.Number(label="حالات التقصير")
                            t_btn = gr.Button("✅ حفظ التعديلات", elem_classes=["admin-btn", "vault-save-btn"])
                            t_del_btn = gr.Button("🗑️ حذف السجل", elem_classes=["reset-btn", "vault-delete-btn"], visible=False)
                        vault_status = gr.HTML()
                            
                with gr.Tab("🛡️ حالات الإعفاء", id="exemptions") as exemptions_tab:
                    gr.Markdown("### 🚫 تثبيت الإعفاءات الدائمة")
                    with gr.Column(elem_classes="shield-box"):
                        rule_teacher = gr.Dropdown(get_teacher_choices("الكل"), label="👨‍🏫 اختر المعلم المراد إعفاؤه", elem_classes="masar-arrow-fix")
                        with gr.Row():
                            rule_days = gr.CheckboxGroup(SCHOOL_WEEK_DAYS, label="📅 أيام معفى منها", elem_classes="exemption-rtl-group")
                            rule_periods = gr.CheckboxGroup(list(range(1, MAX_PERIODS + 1)), label="⏱️ حصص معفى منها", elem_classes="exemption-rtl-group exemption-periods-order")
                        rule_save_btn = gr.Button("✅ حفظ قوانين هذا المعلم", elem_classes="admin-btn")
                        rule_status = gr.HTML()
                        exemptions_log_html = gr.HTML(value=render_exemptions_log_html())

                with gr.Tab("🤝 التبادل الودي الأسبوعي", id="swap") as swap_tab:
                    swap_confirmed_state = gr.State({})

                    gr.HTML("""
                    <div class='swap-guide-box'>
                        💡 <b>توضيح:</b><br>
                        ⚫️ التبادل الودي مخصص للاتفاق بين <b>معلمين حاضرين</b>، وليس لمعالجة غياب معلم.<br>
                        ⚫️ اختر <b>اليوم</b>، ثم <b>القسم</b>، ثم اختر <b>المعلم الراغب في التبادل</b>.<br>
                        ⚫️ اختر <b>الحصة المطلوب تبديلها</b>، وسيعرض النظام البدائل المناسبة حسب جدول اليوم.<br>
                        ⚫️ عند اختيار بديل مناسب، ستظهر <b>رسالة واتساب جاهزة</b> للتنسيق.<br>
                        ⚫️ بعد الاتفاق، اضغط <b>اعتماد التبادل</b> ليتم حفظه في جدول التبادلات المعتمدة.<br>
                        ⚫️ بعد اعتماد التبادل، سيظهر في جدول التبادلات المعتمدة، وتظهر خيارات المشاركة أو التصدير حسب صلاحية المستخدم.
                    </div>
                    """)

                    with gr.Row(elem_classes="yellow-box"):
                        swap_day = gr.Dropdown(
                            SCHOOL_WEEK_DAYS,
                            label="1️⃣ اليوم",
                            value=get_current_day_oman(),
                            elem_classes="masar-arrow-fix"
                        )
                        swap_dept = gr.Dropdown(
                            ["الكل"] + [d for d in OFFICIAL_DEPTS if d != "الهيئة الإدارية"],
                            label="2️⃣ القسم",
                            value="الكل",
                            elem_classes="masar-arrow-fix"
                        )
                        swap_t1 = gr.Dropdown(
                            get_teacher_choices("الكل"),
                            label="3️⃣ المعلم الطالب للتبادل",
                            value=None,
                            allow_custom_value=False,
                            elem_classes="masar-arrow-fix"
                        )
                        swap_p1 = gr.Dropdown([], label="4️⃣ الحصة المراد مبادلتها", allow_custom_value=False, elem_classes="masar-arrow-fix")

                    btn_run_radar = gr.Button("تشغيل الرادار والبحث عن بدائل الآن", variant="primary", visible=False)

                    swap_options = gr.Radio(
                        label="5️⃣ الخيارات المتاحة (اختر المعلم الذي يناسبك لتوليد الرسالة 💬)",
                        choices=[],
                        visible=True,
                        elem_classes="swap-radio-square"
                    )

                    whatsapp_msg = gr.Textbox(
                        label="💬 معاينة رسالة الواتساب التلقائية (يمكنك التعديل عليها)",
                        lines=6,
                        interactive=True,
                        value=SWAP_EMPTY_MSG,
                        visible=True
                    )

                    wa_html_btn = gr.HTML(value="", visible=True)
                    btn_swap_confirm = gr.Button("✅ اعتماد", visible=True, interactive=False, elem_classes="tabadul-btn")

                    confirmed_tbl_html = gr.HTML(value=render_swap_table_html({}))
                    with gr.Row():
                        btn_swap_img = gr.Button("🖼️ إنشاء صورة جدول التبادلات", elem_classes="export-btn")
                    swap_img_out = gr.Image(label="صورة جدول التبادلات المعتمدة", interactive=False)
                    with gr.Column(visible=True) as swap_export_row:
                        btn_swap_excel = gr.Button("📥 تصدير التبادلات المعتمدة Excel", elem_classes="export-btn")
                        swap_excel_out = gr.File(label="ملف Excel للتبادلات المعتمدة", interactive=False)

                with gr.Tab("📅 جدول اليوم", id="day_table") as day_tab:
                    day_page_state = gr.State(value=0)
                    with gr.Group(elem_classes=["teacher-schedule-selector-card"]):
                        gr.HTML("""
<div class="teacher-schedule-selector-title">
  <div class="title-main">عرض جدول اليوم</div>
  <div class="title-sub">هذا الفلتر خاص بجدول اليوم فقط ولا يغيّر صلاحيات التوزيع أو الإعفاءات.</div>
</div>
""")
                        day_dept_filter = gr.Dropdown(["الكل"] + OFFICIAL_DEPTS, label="القسم المعروض", value="الكل", elem_classes=["masar-arrow-fix", "masar-field-label-right"])
                    tbl_day = gr.Dataframe(headers=["المعلم"] + [f"ح {p}" for p in range(1, MAX_PERIODS + 1)], interactive=False, visible=True)
                    day_table_html = gr.HTML(visible=False)
                    with gr.Row(visible=False) as day_pagination_row:
                        btn_prev_page = gr.Button("◀ السابق", elem_classes="admin-btn", scale=1, min_width=110)
                        page_info_html = gr.HTML(elem_classes="day-page-info")
                        btn_next_page = gr.Button("التالي ▶", elem_classes="admin-btn", scale=1, min_width=110)
                with gr.Tab("🔍 جدول المعلم", id="teacher_table") as teacher_tab:
                    with gr.Group(elem_classes=["teacher-schedule-selector-card"]):
                        gr.HTML("""
<div class="teacher-schedule-selector-title">
  <div class="title-main">اختيار جدول المعلم</div>
  <div class="title-sub">اختر اسم المعلم لعرض جدوله الأسبوعي.</div>
</div>
""")
                        check_teacher_in = gr.Dropdown(get_teacher_schedule_choices("الكل"), label="اختر المعلم", elem_classes=["masar-arrow-fix", "masar-field-label-right"])
                    check_tbl = gr.HTML("<div style='text-align:center; color:#64748b; padding:18px; background:#f8fafc; border:1px dashed #cbd5e1; border-radius:12px; direction:rtl;'>اختر المعلم لعرض جدوله الأسبوعي.</div>")
                    check_teacher_in.change(get_teacher_weekly_schedule_html, check_teacher_in, check_tbl)
                with gr.Tab("🗄️ مركز البيانات المدرسية", id="school_data") as school_data_tab:
                    gr.HTML("""
<div class="school-data-center-note">
    <b>🗄️ مركز البيانات المدرسية</b><br>
    اختر بطاقة من الأعلى. كل بطاقة تفتح أدواتها مباشرة دون قوائم داخلية زائدة.
</div>
""")
                    with gr.Row(elem_classes="school-data-nav-row"):
                        btn_school_data_references_panel = gr.Button("🗂️\nالملفات المرجعية", elem_id="school_data_btn_references", elem_classes=["school-data-nav-btn"])
                        btn_school_data_identity_panel = gr.Button("🎨\nهوية المدرسة", elem_id="school_data_btn_identity", elem_classes=["school-data-nav-btn"])
                        btn_school_data_periods_panel = gr.Button("⚙️\nإعدادات التشغيل", elem_id="school_data_btn_periods", elem_classes=["school-data-nav-btn"])
                        btn_school_data_accounts_panel = gr.Button("🔐\nحسابات الدخول", elem_id="school_data_btn_accounts", elem_classes=["school-data-nav-btn"])
                        btn_school_data_audit_panel = gr.Button("🛡️\nالسجل والنسخ", elem_id="school_data_btn_audit", elem_classes=["school-data-nav-btn"])

                    persistent_storage_status_html = gr.HTML(value=render_persistent_storage_status_html(), elem_id="school_data_overview_storage")
                    school_config_summary_html = gr.HTML(value=render_school_config_summary_html(), elem_id="school_data_overview_config")


                    school_data_section_status = gr.HTML(
                        elem_id="school_data_section_status",
                        value="<div class='school-data-panel-title'>اختر بطاقة من الأعلى لعرض أدواتها. حالة التخزين وإعدادات المدرسة تظهر هنا دائمًا.</div>"
                    )

                    with gr.Column(visible=False, elem_id="school_data_panel_references", elem_classes="school-data-panel-box") as school_data_references_panel:
                        gr.HTML("<div class='school-data-panel-title'>🗂️ الملفات المرجعية</div>")
                        gr.HTML("<div class='school-data-panel-lead'>رفع وتحديث جداول الأقسام، أرقام المعلمين، وملف الإداريين المرجعي.</div>")
                        with gr.Row(visible=False):
                            clear_noop = gr.Textbox(label="noop", value="", visible=False)
                            up_dept = gr.Dropdown([], label="noop", visible=False)

                        school_data_admin_html = gr.HTML(value=render_admin_reference_card())
                        with gr.Row():
                            admin_reference_upload = gr.File(label="رفع ملف الإداريين المرجعي", file_types=[".xlsx", ".xls", ".csv"], elem_classes="masar-file-upload-right")
                        with gr.Row():
                            save_admin_reference_btn = gr.Button("💾 اعتماد ملف الإداريين المرجعي", elem_classes="admin-btn")
                            refresh_admin_reference_btn = gr.Button("🔄 تحديث الإداريين من الملف المرجعي", elem_classes="admin-btn")
                        admin_reference_status_html = gr.HTML()

                        school_data_phones_html = gr.HTML(value=render_phones_reference_card())
                        with gr.Row():
                            phones_reference_upload = gr.File(label="رفع ملف أرقام المعلمين المرجعي", file_types=[".xlsx", ".xls", ".csv"], elem_classes="masar-file-upload-right")
                        with gr.Row():
                            save_phones_reference_btn = gr.Button("💾 اعتماد ملف أرقام المعلمين المرجعي", elem_classes="admin-btn")
                            refresh_phones_reference_btn = gr.Button("🔄 تحديث أرقام المعلمين من الملف المرجعي", elem_classes="admin-btn")
                        phones_reference_status_html = gr.HTML()

                        school_data_schedules_html = gr.HTML(value=render_schedule_reference_cards())
                        with gr.Row():
                            schedule_reference_dept = gr.Dropdown(
                                choices=list(SCHEDULE_FILES.keys()),
                                label="اختر القسم لملفه المرجعي",
                                value="التربية الإسلامية",
                                elem_classes="masar-arrow-fix"
                            )
                        with gr.Row():
                            schedule_reference_upload = gr.File(
                                label="رفع ملف الجدول المرجعي للقسم المختار",
                                file_types=[".xlsx", ".xls", ".csv"],
                                elem_classes="masar-file-upload-right"
                            )
                        with gr.Row():
                            save_schedule_reference_btn = gr.Button("💾 اعتماد الملف المرجعي للقسم", elem_classes="admin-btn")
                            refresh_schedule_reference_btn = gr.Button("🔄 تحديث القسم من الملف المرجعي", elem_classes="admin-btn")
                        schedule_reference_status_html = gr.HTML()


                        with gr.Column(visible=False, elem_classes=["school-data-inner-card", "manual-tools-direct-panel"]) as manual_entry_container:
                            gr.HTML("<div class='school-data-subsection-title'>👨‍💼 الإدخال اليدوي للطاقم الإداري</div>")
                            with gr.Row(elem_classes="yellow-box"):
                                manual_name = gr.Textbox(label="الاسم الثلاثي")
                                manual_dept = gr.Dropdown(["الهيئة الإدارية"], label="القسم", value="الهيئة الإدارية", interactive=False, elem_classes="fixed-dd")
                                manual_role = gr.Dropdown(ADMIN_ROLES, label="المنصب", value="أخصائي اجتماعي", elem_classes=["fixed-dd", "masar-arrow-fix"])
                                manual_phone = gr.Textbox(label="رقم الواتساب")
                            with gr.Row():
                                manual_add_btn = gr.Button("➕ حفظ وإضافة", elem_classes="admin-btn")
                        manual_status_html = gr.HTML()
                        clear_status_html = gr.HTML()
                        clear_btn = gr.Button("مسح وتصفير المنظومة", elem_classes="reset-btn", visible=False)


                    with gr.Column(visible=False, elem_id="school_data_panel_identity", elem_classes="school-data-panel-box") as school_data_identity_panel:
                        gr.HTML("<div class='school-data-panel-title'>🎨 هوية المدرسة</div>")
                        with gr.Column(elem_classes=["school-data-inner-card", "identity-direct-panel"]):
                            gr.HTML("<div class='school-data-subsection-title'>🎨 إعدادات الهوية</div>")
                            gr.HTML("<div style='background:#eef6f3;color:#004d40;padding:12px;border-radius:10px;border-right:5px solid #0f766e;margin-bottom:12px;font-weight:800;line-height:1.8;'>هذه اللوحة مخصصة لمالك النظام. يمكن تعديل الهوية البصرية دون تعديل app.py. العناوين والشعار تتحدث فورًا، أما الألوان العامة فتُطبق بالكامل بعد إعادة تشغيل التطبيق.</div>")

                            gr.HTML(
                                "<div style='background:#f8fafc;color:#334155;padding:12px;"
                                "border-radius:10px;border-right:5px solid #64748b;"
                                "margin-bottom:12px;font-weight:800;line-height:1.9;text-align:right;'>"
                                "العناصر الثابتة في الهوية: <b>وزارة التعليم</b>، "
                                "<b>منظومة مسار</b>، <b>للاحتياط والتبادل الودي</b>، "
                                "وعبارة <b>فكرة وتطوير</b>. يمكن تعديل اسم المدرسة، المحافظة، الشعار، والألوان فقط."
                                "</div>"
                            )
                            identity_system_name = gr.Textbox(value=SYSTEM_NAME, label="اسم المنظومة", visible=False)
                            identity_system_subtitle = gr.Textbox(value=SYSTEM_SUBTITLE, label="العنوان الفرعي", visible=False)
                            identity_developer_credit = gr.Textbox(value=DEVELOPER_CREDIT, label="عبارة الحقوق والتطوير", visible=False)
                            with gr.Row():
                                identity_school_name = gr.Textbox(value=SCHOOL_NAME, label="اسم المدرسة")
                                identity_directorate_region = gr.Textbox(value=DIRECTORATE_REGION, label="المحافظة في سطر المديرية", placeholder="مثال: جنوب الباطنة")
                            with gr.Row():
                                identity_logo_url = gr.Textbox(value=SCHOOL_LOGO_URL, label="رابط الشعار أو مساره المحلي")
                                identity_logo_upload = gr.File(
                                    label="رفع شعار بديل اختياري",
                                    file_types=[".png", ".jpg", ".jpeg", ".webp"],
                                    type="filepath",
                                    elem_classes="masar-file-upload-right",
                                )
                            with gr.Row():
                                identity_theme_color = gr.Textbox(value=THEME_COLOR, label="اللون الأساسي HEX")
                                identity_theme_color_2 = gr.Textbox(value=THEME_COLOR_2, label="اللون الثانوي HEX")
                                identity_accent_color = gr.Textbox(value=ACCENT_COLOR, label="اللون البارز HEX")

                            with gr.Row():
                                identity_preview_btn = gr.Button("معاينة هوية المدرسة", elem_classes="admin-btn")
                                identity_save_btn = gr.Button("حفظ هوية المدرسة", elem_classes="admin-btn")
                                identity_reset_btn = gr.Button("استعادة الهوية الافتراضية", elem_classes="reset-btn")

                            identity_status_html = gr.HTML()
                            gr.HTML(
                                "<div style='background:#fff7ed;color:#9a3412;padding:10px;"
                                "border-radius:10px;border-right:5px solid #f59e0b;"
                                "margin:12px 0;font-weight:900;line-height:1.8;text-align:right;'>"
                                "تنبيه: هذه معاينة هوية المدرسة داخل مركز البيانات، وليست الهيدر الفعلي بعد تسجيل الدخول."
                                "</div>"
                            )
                            gr.HTML(
                                "<div style='font-weight:900;color:#004d40;margin:8px 0 6px;text-align:right;'>"
                                "🖼️ معاينة هوية المدرسة"
                                "</div>"
                            )
                            identity_preview_html = gr.HTML(
                                value=render_school_identity_preview_html(
                                    SYSTEM_NAME,
                                    SYSTEM_SUBTITLE,
                                    SCHOOL_NAME,
                                    DEVELOPER_CREDIT,
                                    SCHOOL_LOGO_URL,
                                    THEME_COLOR,
                                    THEME_COLOR_2,
                                    ACCENT_COLOR,
                                    DIRECTORATE_REGION,
                                )
                            )


                    with gr.Column(visible=False, elem_id="school_data_panel_periods", elem_classes="school-data-panel-box") as school_data_periods_panel:
                        gr.HTML("<div class='school-data-panel-title'>⚙️ إعدادات التشغيل المدرسية</div>")
                        with gr.Column(elem_classes=["school-data-inner-card", "periods-direct-panel"]):
                            gr.HTML("<div class='school-data-subsection-title'>🕒 عدد الحصص اليومية</div>")
                            gr.HTML(
                                "<div style='background:#eef6f3;color:#004d40;padding:12px;"
                                "border-radius:10px;border-right:5px solid #0f766e;"
                                "margin-bottom:12px;font-weight:800;line-height:1.8;text-align:right;'>"
                                "اختر عدد الحصص اليومي للمدرسة. هذا الإعداد يُحفظ في ملف school_config.json الحقيقي. "
                                "بعد الحفظ يلزم إعادة تشغيل المنظومة حتى تُبنى القوائم والجداول على العدد الجديد."
                                "</div>"
                            )
                            operational_settings_current_html = gr.HTML(value=render_operational_settings_status_html())
                            with gr.Row():
                                operational_periods_dropdown = gr.Dropdown(
                                    choices=[7, 8],
                                    value=get_config_periods_per_day(),
                                    label="عدد الحصص اليومية",
                                    elem_classes="masar-arrow-fix",
                                )
                            with gr.Row():
                                operational_periods_save_btn = gr.Button("حفظ إعداد عدد الحصص", elem_classes="admin-btn")
                            operational_periods_status_html = gr.HTML()
                            gr.HTML(
                                "<div style='background:#fff7ed;color:#9a3412;padding:10px;"
                                "border-radius:10px;border-right:5px solid #f59e0b;"
                                "margin-top:12px;font-weight:900;line-height:1.8;text-align:right;'>"
                                "تنبيه: لا تحدّث ملفات جداول 7 حصص والمنظومة مضبوطة على 8، ولا العكس. "
                                "غيّر الإعداد ثم أعد التشغيل قبل رفع الجداول المناسبة."
                                "</div>"
                            )


                    with gr.Column(visible=False, elem_id="school_data_panel_accounts", elem_classes="school-data-panel-box") as school_data_accounts_panel:
                        gr.HTML("<div class='school-data-panel-title'>🔐 حسابات الدخول</div>")
                        with gr.Column(elem_classes=["school-data-inner-card", "accounts-direct-panel"]):
                            gr.HTML("<div class='school-data-subsection-title'>🔐 إدارة الحسابات والرموز</div>")
                            gr.HTML(
                                "<div style='background:#eef6f3;color:#004d40;padding:12px;"
                                "border-radius:10px;border-right:5px solid #0f766e;"
                                "margin-bottom:12px;font-weight:800;line-height:1.8;'>"
                                "لوحة المالك لإعادة تعيين الرموز وتعطيل الحسابات. "
                                "لا تعرض المنظومة أي رمز قديم. الرمز الجديد يظهر مرة واحدة فقط بعد إعادة التعيين."
                                "</div>"
                            )
                            owner_accounts_html = gr.HTML(
                                value=render_auth_accounts_html(False)
                            )
                            with gr.Row():
                                owner_account_selector = gr.Dropdown(
                                    choices=[],
                                    value=None,
                                    label="اختر الحساب",
                                    elem_classes="masar-arrow-fix",
                                )
                                owner_requested_pin = gr.Textbox(
                                    type="password",
                                    label="رمز جديد اختياري",
                                    placeholder="اتركه فارغًا لتوليد رمز من 6 أرقام",
                                )
                            with gr.Row():
                                owner_reset_pin_btn = gr.Button(
                                    "إعادة تعيين رمز الحساب",
                                    elem_classes="admin-btn",
                                )
                                owner_toggle_account_btn = gr.Button(
                                    "تفعيل / تعطيل الحساب",
                                    elem_classes="reset-btn",
                                )
                                owner_refresh_accounts_btn = gr.Button(
                                    "تحديث قائمة الحسابات",
                                    elem_classes="admin-btn",
                                )
                            owner_one_time_pin = gr.Textbox(
                                label="الرمز الجديد لمرة واحدة",
                                interactive=False,
                                value="",
                            )

                            with gr.Column(elem_classes=["school-data-subsection-card", "account-profile-direct-panel"]):
                                gr.HTML("<div class='school-data-subsection-title'>الترحيب والمسميات</div>")
                                gr.HTML(
                                    "<div style='background:#fff7ed;color:#7c2d12;padding:10px;"
                                    "border-radius:8px;border-right:4px solid #f59e0b;"
                                    "font-weight:800;line-height:1.8;'>"
                                    "اختر حسابًا من الأعلى، ثم اضبط اسم العرض والمسمى المختصر وعبارة الهيدر. "
                                    "هذه الحقول تخص عرض الحساب والترحيب، ولا تغيّر رسائل الاحتياط أو مادة الحصة. "
                                    "يمكن استخدام المتغيرات: {display_name}، {welcome_title}، {department_label}، {official_title}، {whatsapp_title}، {school_name}."
                                    "</div>"
                                )
                                with gr.Row():
                                    owner_profile_display_name = gr.Textbox(
                                        label="اسم العرض",
                                        placeholder="مثال: أ. سعود المعولي",
                                    )
                                    owner_profile_official_title = gr.Textbox(
                                        label="المسمى الرسمي",
                                        placeholder="مثال: منسق مادة اللغة العربية",
                                    )
                                    owner_profile_whatsapp_title = gr.Textbox(
                                        label="المسمى المختصر للحساب",
                                        placeholder="مثال: منسق اللغة العربية",
                                    )
                                with gr.Row():
                                    owner_profile_welcome_title = gr.Textbox(
                                        label="اللقب الجمالي",
                                        placeholder="مثال: مايسترو البيان",
                                    )
                                    owner_profile_department_label = gr.Textbox(
                                        label="القسم الظاهر",
                                        placeholder="مثال: قسم اللغة العربية",
                                    )
                                owner_profile_welcome_phrase = gr.Textbox(
                                    label="عبارة الترحيب",
                                    placeholder="مثال: نورتنا، وقسم اللغة العربية جاهز لك",
                                )
                                owner_profile_welcome_template = gr.Textbox(
                                    label="قالب الهيدر",
                                    value=ACCOUNT_WELCOME_DEFAULT_TEMPLATE,
                                    placeholder="مثال: {welcome_title} ({display_name}) {welcome_phrase}",
                                )
                                with gr.Row():
                                    owner_profile_preview_btn = gr.Button(
                                        "معاينة الترحيب",
                                        elem_classes="admin-btn",
                                    )
                                    owner_profile_save_btn = gr.Button(
                                        "حفظ تخصيص الحساب",
                                        elem_classes="admin-btn",
                                    )
                                owner_profile_preview_html = gr.HTML()

                            owner_accounts_status = gr.HTML()


                    with gr.Column(visible=False, elem_id="school_data_panel_audit", elem_classes="school-data-panel-box") as school_data_audit_panel:
                        gr.HTML("<div class='school-data-panel-title'>🛡️ السجل والنسخ</div>")
                        with gr.Column(elem_classes=["school-data-inner-card", "audit-direct-panel"]):
                            gr.HTML("<div style='background:#eef6f3;color:#004d40;padding:12px;border-radius:10px;border-right:5px solid #0f766e;margin-bottom:12px;font-weight:800;line-height:1.8;'>هذه أدوات رقابية لمالك النظام فقط. سجل العمليات لا ينفذ أي تعديل؛ بل يوضح من غيّر ماذا، وعلى أي معلم، وما القيمة القديمة والجديدة، ومتى حدث ذلك. اختر الفلاتر للعرض، ثم صدّر النتائج المطابقة إلى Excel عند الحاجة.</div>")

                            with gr.Column(elem_classes="school-data-subsection-card"):
                                gr.HTML("<div class='school-data-subsection-title'>📑 سجل العمليات الحساسة</div>")
                                with gr.Row():
                                    audit_action_filter = gr.Dropdown(["الكل"], value="الكل", label="نوع العملية", elem_classes="masar-arrow-fix")
                                    audit_actor_filter = gr.Dropdown(["الكل"], value="الكل", label="اسم المنفذ", elem_classes="masar-arrow-fix")
                                    audit_teacher_filter = gr.Dropdown(["الكل"], value="الكل", label="المعلم المتأثر", elem_classes="masar-arrow-fix")
                                with gr.Row():
                                    audit_date_from = gr.DateTime(
                                        label="تاريخ البداية",
                                        include_time=False,
                                        type="string",
                                        timezone="Asia/Muscat",
                                        elem_classes="masar-date-label-right",
                                    )
                                    audit_date_to = gr.DateTime(
                                        label="تاريخ النهاية",
                                        include_time=False,
                                        type="string",
                                        timezone="Asia/Muscat",
                                        elem_classes="masar-date-label-right",
                                    )
                                with gr.Row():
                                    audit_today_btn = gr.Button("اليوم", elem_classes="admin-btn")
                                    audit_last_7_days_btn = gr.Button("آخر 7 أيام", elem_classes="admin-btn")
                                    audit_this_month_btn = gr.Button("هذا الشهر", elem_classes="admin-btn")
                                    audit_clear_dates_btn = gr.Button("مسح التاريخ", elem_classes="reset-btn")
                                with gr.Row():
                                    audit_refresh_btn = gr.Button("عرض النتائج حسب الفلاتر", elem_classes="admin-btn")
                                    audit_export_btn = gr.Button("تصدير السجل إلى Excel", elem_classes="export-btn")
                                audit_action_status_html = gr.HTML()
                                audit_table_html = gr.HTML("<div style='text-align:center;color:#64748b;padding:16px;'>افتح القسم أو اضغط تحديث لعرض سجل العمليات.</div>")
                                audit_export_file = gr.File(label="ملف سجل العمليات", interactive=False)

                            with gr.Column(elem_classes="school-data-subsection-card"):
                                gr.HTML("<div class='school-data-subsection-title'>💾 حالة النسخ الاحتياطية</div>")
                                backup_status_html = gr.HTML("<div style='text-align:center;color:#64748b;padding:16px;'>اضغط تحديث لعرض حالة النسخ الاحتياطية.</div>")
                                with gr.Row():
                                    backup_refresh_btn = gr.Button("تحديث حالة النسخ الاحتياطية", elem_classes="admin-btn")
                                    backup_zip_btn = gr.Button("تحميل نسخة احتياطية كاملة ZIP", elem_classes="export-btn")
                                backup_action_status_html = gr.HTML()
                                backup_zip_file = gr.File(label="الحزمة الاحتياطية", interactive=False)

    # ── ربط الأحداث ──────────────────────────────────────────────
    update_outputs = [
        abs_in, tbl_bal, tbl_abs, tbl_short, tbl_day, day_table_html, day_pagination_row, btn_prev_page, btn_next_page, page_info_html, day_page_state, t_name, check_teacher_in, rule_teacher, 
        radar_warning_html, tbl_out, edit_abs_t, current_schedule_state, 
        msg_summary, msg_individual_html, date_display, admin_zone_title,
        admin_zone_help, edit_period, cb_cross_dept, btn_alt, btn_img
    ]
    app.load(sync_current_school_days, None, [day_in, swap_day])

    btn_school_data_references_panel.click(
        lambda: show_school_data_panel("references"),
        [],
        [school_data_section_status, persistent_storage_status_html, school_config_summary_html, school_data_references_panel, school_data_identity_panel, school_data_periods_panel, school_data_accounts_panel, school_data_audit_panel],
        queue=False,
    ).then(
        None,
        None,
        None,
        js=school_data_panel_js("references"),
    )
    btn_school_data_identity_panel.click(
        lambda: show_school_data_panel("identity"),
        [],
        [school_data_section_status, persistent_storage_status_html, school_config_summary_html, school_data_references_panel, school_data_identity_panel, school_data_periods_panel, school_data_accounts_panel, school_data_audit_panel],
        queue=False,
    ).then(
        None,
        None,
        None,
        js=school_data_panel_js("identity"),
    )
    btn_school_data_periods_panel.click(
        lambda: show_school_data_panel("periods"),
        [],
        [school_data_section_status, persistent_storage_status_html, school_config_summary_html, school_data_references_panel, school_data_identity_panel, school_data_periods_panel, school_data_accounts_panel, school_data_audit_panel],
        queue=False,
    ).then(
        None,
        None,
        None,
        js=school_data_panel_js("periods"),
    )
    btn_school_data_accounts_panel.click(
        lambda: show_school_data_panel("accounts"),
        [],
        [school_data_section_status, persistent_storage_status_html, school_config_summary_html, school_data_references_panel, school_data_identity_panel, school_data_periods_panel, school_data_accounts_panel, school_data_audit_panel],
        queue=False,
    ).then(
        None,
        None,
        None,
        js=school_data_panel_js("accounts"),
    ).then(
        refresh_owner_accounts_panel,
        [current_user_is_owner],
        [owner_accounts_html, owner_account_selector, owner_one_time_pin, owner_accounts_status],
        queue=False,
    )
    btn_school_data_audit_panel.click(
        lambda: show_school_data_panel("audit"),
        [],
        [school_data_section_status, persistent_storage_status_html, school_config_summary_html, school_data_references_panel, school_data_identity_panel, school_data_periods_panel, school_data_accounts_panel, school_data_audit_panel],
        queue=False,
    ).then(
        None,
        None,
        None,
        js=school_data_panel_js("audit"),
    )

    btn_open_distribution.click(lambda: open_home_section("distribution"), [], [home_dashboard, tabs_container, main_tabs], queue=False).then(None, None, None, js=show_selected_tab_container_js())
    btn_open_balances.click(lambda: open_home_section("balances"), [], [home_dashboard, tabs_container, main_tabs], queue=False).then(None, None, None, js=show_selected_tab_container_js())
    btn_open_exemptions.click(lambda: open_home_section("exemptions"), [], [home_dashboard, tabs_container, main_tabs], queue=False).then(None, None, None, js=show_selected_tab_container_js())
    btn_open_swap.click(lambda: open_home_section("swap"), [], [home_dashboard, tabs_container, main_tabs], queue=False).then(None, None, None, js=show_selected_tab_container_js())
    btn_open_day.click(lambda: open_home_section("day_table"), [], [home_dashboard, tabs_container, main_tabs], queue=False).then(None, None, None, js=show_selected_tab_container_js())
    btn_open_teacher.click(lambda: open_home_section("teacher_table"), [], [home_dashboard, tabs_container, main_tabs], queue=False).then(None, None, None, js=show_selected_tab_container_js())
    # v1.8.3 Fix 13b — direct school data panels and no select races
    btn_open_school_data.click(
        lambda: open_home_section("school_data"),
        [],
        [home_dashboard, tabs_container, main_tabs],
        queue=False,
    ).then(
        lambda: show_school_data_panel("overview"),
        [],
        [
            school_data_section_status,
            persistent_storage_status_html,
            school_config_summary_html,
            school_data_references_panel,
            school_data_identity_panel,
            school_data_periods_panel,
            school_data_accounts_panel,
            school_data_audit_panel,
        ],
        queue=False,
    ).then(
        None,
        None,
        None,
        js=show_selected_tab_container_js(),
    ).then(
        None,
        None,
        None,
        js=school_data_panel_js("overview"),
    ).then(
        refresh_school_data_center_cards,
        [current_user_is_owner],
        [
            persistent_storage_status_html,
            school_config_summary_html,
            school_data_admin_html,
            school_data_phones_html,
            school_data_schedules_html,
        ],
        queue=False,
    ).then(
        refresh_owner_accounts_panel,
        [current_user_is_owner],
        [
            owner_accounts_html,
            owner_account_selector,
            owner_one_time_pin,
            owner_accounts_status,
        ],
        queue=False,
    ).then(
        refresh_owner_tools_dashboard,
        [
            audit_action_filter,
            audit_actor_filter,
            audit_teacher_filter,
            audit_date_from,
            audit_date_to,
            current_user_is_owner,
        ],
        [
            audit_action_filter,
            audit_actor_filter,
            audit_teacher_filter,
            audit_table_html,
            backup_status_html,
        ],
        queue=False,
    )
    btn_back_home.click(return_to_home_dashboard, [], [home_dashboard, tabs_container], queue=False).then(None, None, None, js=return_home_dashboard_js())
    login_btn.click(
        attempt_login,
        inputs=[pin_input, day_in],
        outputs=[login_container, main_app_container, welcome_html, self_pin_accordion, dept_in, day_dept_filter, login_msg, up_dept, manual_entry_container, current_user_is_admin, current_user_is_owner, current_user_name, current_user_role, current_user_account_id] + update_outputs + [t_specialty_edit, clear_btn, school_data_tab, controls_row, exemptions_tab, distribution_tab, balances_tab, swap_tab, day_tab, teacher_tab, swap_export_row]
    ).then(
        show_home_dashboard_after_login,
        [dept_in, current_user_is_admin, current_user_is_owner, current_user_role],
        [home_dashboard, tabs_container, card_distribution, card_balances, card_exemptions, card_swap, card_day, card_teacher, card_school_data],
        queue=False
    ).then(
        update_home_hero_after_login,
        [
            current_user_account_id,
            current_user_name,
            current_user_role,
            current_user_is_owner,
        ],
        [home_hero_html],
        queue=False,
    ).then(
        None, None, None,
        js=show_home_dashboard_js()
    ).then(
        refresh_ui_on_change,
        [dept_in, day_in, current_user_is_admin],
        update_outputs
    ).then(
        lambda dy, day_dp: get_day_table_updates(dy, day_dp, 0),
        [day_in, day_dept_filter],
        [tbl_day, day_table_html, day_pagination_row, btn_prev_page, btn_next_page, page_info_html, day_page_state],
        queue=False
    ).then(
        get_generation_button_updates,
        [abs_in, day_in, dept_in, reserve_generation_state],
        [btn, btn_regenerate]
    ).then(
        lambda: gr.update(value=render_exemptions_log_html()),
        [],
        [exemptions_log_html],
        queue=False
    ).then(
        lambda adm, own: gr.update(visible=bool(adm or own)),
        [current_user_is_admin, current_user_is_owner],
        [reset_month_btn],
        queue=False
    )
    pin_input.submit(
        attempt_login,
        inputs=[pin_input, day_in],
        outputs=[login_container, main_app_container, welcome_html, self_pin_accordion, dept_in, day_dept_filter, login_msg, up_dept, manual_entry_container, current_user_is_admin, current_user_is_owner, current_user_name, current_user_role, current_user_account_id] + update_outputs + [t_specialty_edit, clear_btn, school_data_tab, controls_row, exemptions_tab, distribution_tab, balances_tab, swap_tab, day_tab, teacher_tab, swap_export_row]
    ).then(
        show_home_dashboard_after_login,
        [dept_in, current_user_is_admin, current_user_is_owner, current_user_role],
        [home_dashboard, tabs_container, card_distribution, card_balances, card_exemptions, card_swap, card_day, card_teacher, card_school_data],
        queue=False
    ).then(
        update_home_hero_after_login,
        [
            current_user_account_id,
            current_user_name,
            current_user_role,
            current_user_is_owner,
        ],
        [home_hero_html],
        queue=False,
    ).then(
        None, None, None,
        js=show_home_dashboard_js()
    ).then(
        refresh_ui_on_change,
        [dept_in, day_in, current_user_is_admin],
        update_outputs
    ).then(
        lambda dy, day_dp: get_day_table_updates(dy, day_dp, 0),
        [day_in, day_dept_filter],
        [tbl_day, day_table_html, day_pagination_row, btn_prev_page, btn_next_page, page_info_html, day_page_state],
        queue=False
    ).then(
        get_generation_button_updates,
        [abs_in, day_in, dept_in, reserve_generation_state],
        [btn, btn_regenerate]
    ).then(
        lambda: gr.update(value=render_exemptions_log_html()),
        [],
        [exemptions_log_html],
        queue=False
    ).then(
        lambda adm, own: gr.update(visible=bool(adm or own)),
        [current_user_is_admin, current_user_is_owner],
        [reset_month_btn],
        queue=False
    )
    logout_btn.click(do_logout, inputs=[], outputs=[login_container, main_app_container, welcome_html, self_pin_accordion, dept_in, day_dept_filter, current_user_is_admin, current_user_is_owner, current_user_name, current_user_role, current_user_account_id, current_schedule_state, img_out, cb_cross_dept, school_data_tab, controls_row, exemptions_tab, distribution_tab, balances_tab, swap_tab, day_tab, teacher_tab, swap_export_row, reserve_generation_state, swap_confirmed_state]).then(
        None, None, None,
        js="""() => {
            const style = document.createElement('style');
            style.textContent = '.toast-wrap, .toast-body { display: none !important; }';
            document.head.appendChild(style);
            setTimeout(() => { window.location.reload(); }, 200);
        }"""
    )
    
    update_trigger = [dept_in, day_in, current_user_is_admin]
    dept_in.change(
    lambda d, dy, adm: refresh_ui_on_change(d, dy, adm) + (gr.update(visible=d in ["العلوم", "المهارات الفردية"]),),
    update_trigger,
    update_outputs + [t_specialty_edit]
    )
    day_in.change(lambda d, dy, adm: refresh_ui_on_change(d, dy, adm), update_trigger, update_outputs).then(
        lambda dy, day_dp: get_day_table_updates(dy, day_dp, 0),
        [day_in, day_dept_filter],
        [tbl_day, day_table_html, day_pagination_row, btn_prev_page, btn_next_page, page_info_html, day_page_state],
        queue=False
    )
    day_dept_filter.change(
        lambda dy, day_dp: get_day_table_updates(dy, day_dp, 0),
        [day_in, day_dept_filter],
        [tbl_day, day_table_html, day_pagination_row, btn_prev_page, btn_next_page, page_info_html, day_page_state],
        queue=False
    )
    btn_prev_page.click(
        lambda dy, dp, pg: change_day_page(-1, dy, dp, pg),
        [day_in, day_dept_filter, day_page_state],
        [tbl_day, day_table_html, day_pagination_row, btn_prev_page, btn_next_page, page_info_html, day_page_state],
        queue=False
    )
    btn_next_page.click(
        lambda dy, dp, pg: change_day_page(1, dy, dp, pg),
        [day_in, day_dept_filter, day_page_state],
        [tbl_day, day_table_html, day_pagination_row, btn_prev_page, btn_next_page, page_info_html, day_page_state],
        queue=False
    )
    dept_in.change(clear_generated_image, None, [img_out], queue=False)
    day_in.change(clear_generated_image, None, [img_out], queue=False)
    refresh_btn.click(
        force_refresh_data,
        [dept_in, day_in, current_user_is_admin, abs_in],
        update_outputs
    ).then(
        lambda dy, day_dp: get_day_table_updates(dy, day_dp, 0),
        [day_in, day_dept_filter],
        [tbl_day, day_table_html, day_pagination_row, btn_prev_page, btn_next_page, page_info_html, day_page_state],
        queue=False
    ).then(
        get_generation_button_updates,
        [abs_in, day_in, dept_in, reserve_generation_state],
        [btn, btn_regenerate]
    )
    refresh_btn.click(sync_current_school_days, None, [day_in, swap_day])
    btn_img.click(
        generate_image_only,
        [dept_in, day_in],
        [img_out],
        queue=False
    )
    refresh_btn.click(clear_generated_image, None, [img_out], queue=False)
    btn.click(clear_generated_image, None, [img_out], queue=False)
    btn_regenerate.click(clear_generated_image, None, [img_out], queue=False)
    btn_alt.click(clear_generated_image, None, [img_out], queue=False)
    btn_apply_override.click(clear_generated_image, None, [img_out], queue=False)
    btn_apply_tabadul.click(clear_generated_image, None, [img_out], queue=False)
    btn_apply_penalty.click(clear_generated_image, None, [img_out], queue=False)
    btn_cancel_absence.click(clear_generated_image, None, [img_out], queue=False)
    
    manual_add_btn.click(add_manual_staff, [manual_name, manual_dept, manual_phone, manual_role, dept_in, current_user_is_owner], [manual_status_html, abs_in, check_teacher_in, rule_teacher, t_name, manual_name, manual_phone])
    save_admin_reference_btn.click(
        save_admin_reference_file,
        [admin_reference_upload, current_user_is_owner],
        [admin_reference_status_html, school_data_admin_html]
    )
    refresh_admin_reference_btn.click(
        refresh_admins_from_reference,
        [dept_in, current_user_is_owner],
        [admin_reference_status_html, abs_in, check_teacher_in, rule_teacher, t_name, tbl_bal, school_data_admin_html, admin_reference_upload]
    )
    save_phones_reference_btn.click(
        save_phones_reference_file,
        [phones_reference_upload, current_user_is_owner],
        [phones_reference_status_html, school_data_phones_html]
    )
    refresh_phones_reference_btn.click(
        refresh_phones_from_reference,
        [dept_in, current_user_is_owner],
        [phones_reference_status_html, tbl_bal, school_data_phones_html, phones_reference_upload]
    )
    save_schedule_reference_btn.click(
        save_schedule_reference_file,
        [schedule_reference_upload, schedule_reference_dept, current_user_is_owner],
        [schedule_reference_status_html, school_data_schedules_html]
    )
    refresh_schedule_reference_btn.click(
        refresh_schedule_from_reference,
        [schedule_reference_dept, day_in, current_user_is_owner],
        [schedule_reference_status_html, abs_in, check_teacher_in, rule_teacher, tbl_bal, tbl_abs, tbl_day, school_data_schedules_html, schedule_reference_upload]
    )

    self_change_pin_btn.click(
        change_own_account_pin,
        [
            current_user_account_id,
            self_current_pin,
            self_new_pin,
            self_confirm_pin,
            current_user_name,
            current_user_role,
            current_user_is_owner,
        ],
        [
            self_current_pin,
            self_new_pin,
            self_confirm_pin,
            self_change_pin_status,
        ],
        queue=False,
    )

    owner_refresh_accounts_btn.click(
        refresh_owner_accounts_panel,
        [current_user_is_owner],
        [
            owner_accounts_html,
            owner_account_selector,
            owner_one_time_pin,
            owner_accounts_status,
        ],
        queue=False,
    )

    owner_reset_pin_btn.click(
        owner_reset_account_pin,
        [
            owner_account_selector,
            owner_requested_pin,
            current_user_is_owner,
            current_user_name,
            current_user_role,
        ],
        [
            owner_accounts_html,
            owner_account_selector,
            owner_requested_pin,
            owner_one_time_pin,
            owner_accounts_status,
        ],
        queue=False,
    )

    owner_toggle_account_btn.click(
        owner_toggle_account_status,
        [
            owner_account_selector,
            current_user_is_owner,
            current_user_name,
            current_user_role,
        ],
        [
            owner_accounts_html,
            owner_account_selector,
            owner_one_time_pin,
            owner_accounts_status,
        ],
        queue=False,
    )

    owner_account_selector.change(
        load_auth_account_profile_for_editor,
        [owner_account_selector, current_user_is_owner],
        [
            owner_profile_display_name,
            owner_profile_official_title,
            owner_profile_welcome_title,
            owner_profile_department_label,
            owner_profile_welcome_phrase,
            owner_profile_welcome_template,
            owner_profile_whatsapp_title,
            owner_profile_preview_html,
            owner_accounts_status,
        ],
        queue=False,
    )

    owner_profile_preview_btn.click(
        preview_account_profile_settings,
        [
            owner_profile_display_name,
            owner_profile_official_title,
            owner_profile_welcome_title,
            owner_profile_department_label,
            owner_profile_welcome_phrase,
            owner_profile_welcome_template,
            owner_profile_whatsapp_title,
            current_user_is_owner,
        ],
        [owner_profile_preview_html, owner_accounts_status],
        queue=False,
    )

    owner_profile_save_btn.click(
        save_auth_account_profile,
        [
            owner_account_selector,
            owner_profile_display_name,
            owner_profile_official_title,
            owner_profile_welcome_title,
            owner_profile_department_label,
            owner_profile_welcome_phrase,
            owner_profile_welcome_template,
            owner_profile_whatsapp_title,
            current_user_is_owner,
            current_user_name,
            current_user_role,
        ],
        [
            owner_accounts_html,
            owner_account_selector,
            owner_profile_preview_html,
            owner_accounts_status,
        ],
        queue=False,
    )

    operational_periods_save_btn.click(
        save_school_operational_settings,
        [operational_periods_dropdown, current_user_is_owner, current_user_name, current_user_role],
        [operational_periods_dropdown, operational_periods_status_html, school_config_summary_html, operational_settings_current_html],
        queue=False,
    )

    identity_preview_btn.click(
        preview_school_identity_settings,
        [
            identity_system_name,
            identity_system_subtitle,
            identity_school_name,
            identity_developer_credit,
            identity_directorate_region,
            identity_logo_url,
            identity_theme_color,
            identity_theme_color_2,
            identity_accent_color,
            current_user_is_owner,
        ],
        [identity_preview_html, identity_status_html],
        queue=False,
    )

    identity_save_btn.click(
        save_school_identity_settings,
        [
            identity_system_name,
            identity_system_subtitle,
            identity_school_name,
            identity_developer_credit,
            identity_directorate_region,
            identity_logo_url,
            identity_logo_upload,
            identity_theme_color,
            identity_theme_color_2,
            identity_accent_color,
            current_user_is_owner,
        ],
        [
            identity_system_name,
            identity_system_subtitle,
            identity_school_name,
            identity_directorate_region,
            identity_developer_credit,
            identity_logo_url,
            identity_logo_upload,
            identity_theme_color,
            identity_theme_color_2,
            identity_accent_color,
            identity_status_html,
            identity_preview_html,
            login_branding_html,
            login_credits_html,
            header_branding_html,
            home_hero_html,
            school_config_summary_html,
        ],
        queue=False,
    )

    identity_reset_btn.click(
        reset_school_identity_settings,
        [current_user_is_owner],
        [
            identity_system_name,
            identity_system_subtitle,
            identity_school_name,
            identity_directorate_region,
            identity_developer_credit,
            identity_logo_url,
            identity_logo_upload,
            identity_theme_color,
            identity_theme_color_2,
            identity_accent_color,
            identity_status_html,
            identity_preview_html,
            login_branding_html,
            login_credits_html,
            header_branding_html,
            home_hero_html,
            school_config_summary_html,
        ],
        queue=False,
    )


    audit_today_btn.click(
        lambda: get_audit_date_range("today"),
        [],
        [audit_date_from, audit_date_to],
        queue=False,
    ).then(
        refresh_audit_dashboard,
        [audit_action_filter, audit_actor_filter, audit_teacher_filter, audit_date_from, audit_date_to, current_user_is_owner],
        [audit_action_filter, audit_actor_filter, audit_teacher_filter, audit_table_html],
        queue=False,
    ).then(
        lambda: (gr.update(value=None), gr.update(value="")),
        [],
        [audit_export_file, audit_action_status_html],
        queue=False,
    )

    audit_last_7_days_btn.click(
        lambda: get_audit_date_range("last_7_days"),
        [],
        [audit_date_from, audit_date_to],
        queue=False,
    ).then(
        refresh_audit_dashboard,
        [audit_action_filter, audit_actor_filter, audit_teacher_filter, audit_date_from, audit_date_to, current_user_is_owner],
        [audit_action_filter, audit_actor_filter, audit_teacher_filter, audit_table_html],
        queue=False,
    ).then(
        lambda: (gr.update(value=None), gr.update(value="")),
        [],
        [audit_export_file, audit_action_status_html],
        queue=False,
    )

    audit_this_month_btn.click(
        lambda: get_audit_date_range("this_month"),
        [],
        [audit_date_from, audit_date_to],
        queue=False,
    ).then(
        refresh_audit_dashboard,
        [audit_action_filter, audit_actor_filter, audit_teacher_filter, audit_date_from, audit_date_to, current_user_is_owner],
        [audit_action_filter, audit_actor_filter, audit_teacher_filter, audit_table_html],
        queue=False,
    ).then(
        lambda: (gr.update(value=None), gr.update(value="")),
        [],
        [audit_export_file, audit_action_status_html],
        queue=False,
    )

    audit_clear_dates_btn.click(
        lambda: get_audit_date_range("clear"),
        [],
        [audit_date_from, audit_date_to],
        queue=False,
    ).then(
        refresh_audit_dashboard,
        [audit_action_filter, audit_actor_filter, audit_teacher_filter, audit_date_from, audit_date_to, current_user_is_owner],
        [audit_action_filter, audit_actor_filter, audit_teacher_filter, audit_table_html],
        queue=False,
    ).then(
        lambda: (gr.update(value=None), gr.update(value="")),
        [],
        [audit_export_file, audit_action_status_html],
        queue=False,
    )

    audit_refresh_btn.click(
        refresh_audit_dashboard,
        [audit_action_filter, audit_actor_filter, audit_teacher_filter, audit_date_from, audit_date_to, current_user_is_owner],
        [audit_action_filter, audit_actor_filter, audit_teacher_filter, audit_table_html],
        queue=False
    ).then(
        lambda: (gr.update(value=None), gr.update(value="")),
        [],
        [audit_export_file, audit_action_status_html],
        queue=False,
    )
    audit_export_btn.click(
        export_audit_log_excel,
        [audit_action_filter, audit_actor_filter, audit_teacher_filter, audit_date_from, audit_date_to, current_user_is_owner],
        [audit_export_file, audit_action_status_html],
        queue=False
    )
    backup_refresh_btn.click(
        refresh_backup_status,
        [current_user_is_owner],
        [backup_status_html],
        queue=False
    )
    backup_zip_btn.click(
        create_backup_bundle,
        [current_user_is_owner],
        [backup_zip_file, backup_action_status_html],
        queue=False
    )
    clear_btn.click(
        clear_all_data,
        [current_user_is_owner],
        [
            dept_in,
            abs_in,
            check_teacher_in,
            rule_teacher,
            tbl_bal,
            tbl_abs,
            tbl_short,
            tbl_day,
            day_table_html,
            day_pagination_row,
            btn_prev_page,
            btn_next_page,
            page_info_html,
            day_page_state,
            clear_status_html,
            t_name,
            clear_noop,
            tbl_out,
            img_out,
            current_schedule_state,
            reserve_generation_state,
            msg_summary,
            msg_individual_html,
            date_display,
            radar_warning_html,
            edit_abs_t,
            edit_period,
            btn,
            btn_regenerate,
            btn_alt,
            btn_img,
        ]
    )
    
    rule_teacher.change(
        lambda selected_teacher: load_teacher_rules(selected_teacher) + (gr.update(value=""), gr.update(value=render_exemptions_log_html())),
        rule_teacher,
        [rule_days, rule_periods, rule_status, exemptions_log_html]
    )
    rule_save_btn.click(save_teacher_rules, [rule_teacher, rule_days, rule_periods, current_user_name, current_user_role, current_user_is_admin, current_user_is_owner], [rule_status, exemptions_log_html])
    exemptions_tab.select(
        lambda: gr.update(value=render_exemptions_log_html()),
        [],
        [exemptions_log_html],
        queue=False
    )
    
    btn.click(
        run_main_generation,
        [abs_in, day_in, dept_in, max_reserves_input, current_user_is_admin, reserve_generation_state, current_user_name, current_user_role],
        update_outputs + [btn, btn_regenerate, reserve_generation_state]
    )
    btn_regenerate.click(
        run_full_regeneration,
        [abs_in, day_in, dept_in, max_reserves_input, current_user_is_admin, reserve_generation_state, current_user_name, current_user_role],
        update_outputs + [btn, btn_regenerate, reserve_generation_state]
    )
    btn_alt.click(lambda a, d, dp, mr, adm, an, ar: assign_logic(a, d, dp, mr, True, adm, an, ar), [abs_in, day_in, dept_in, max_reserves_input, current_user_is_admin, current_user_name, current_user_role], update_outputs)
    abs_in.change(
        get_generation_button_updates,
        [abs_in, day_in, dept_in, reserve_generation_state],
        [btn, btn_regenerate]
    )

    day_in.change(
        get_generation_button_updates,
        [abs_in, day_in, dept_in, reserve_generation_state],
        [btn, btn_regenerate]
    )

    dept_in.change(
        get_generation_button_updates,
        [abs_in, day_in, dept_in, reserve_generation_state],
        [btn, btn_regenerate]
    )
    edit_abs_t.change(on_abs_t_change, [current_schedule_state, edit_abs_t, current_user_is_admin], [edit_period, edit_intervention_type, cb_cross_dept]).then(
        lambda: gr.update(choices=[], value=None),
        None,
        [edit_new_sub],
        queue=False
    ).then(
        lambda at: get_leader_action_button_updates(at, None, None),
        [edit_abs_t],
        [btn_apply_override, btn_apply_tabadul, btn_apply_penalty, btn_cancel_absence],
        queue=False
    )
    cb_cross_dept.change(toggle_cross_dept, [cb_cross_dept, edit_abs_t], [edit_intervention_type])
    edit_period.change(update_available_subs_smart, [edit_abs_t, edit_period, edit_intervention_type, day_in, current_schedule_state, current_user_is_admin], [edit_new_sub]).then(
        get_leader_action_button_updates,
        [edit_abs_t, edit_period, edit_new_sub],
        [btn_apply_override, btn_apply_tabadul, btn_apply_penalty, btn_cancel_absence],
        queue=False
    )
    edit_intervention_type.change(update_available_subs_smart, [edit_abs_t, edit_period, edit_intervention_type, day_in, current_schedule_state, current_user_is_admin], [edit_new_sub]).then(
        get_leader_action_button_updates,
        [edit_abs_t, edit_period, edit_new_sub],
        [btn_apply_override, btn_apply_tabadul, btn_apply_penalty, btn_cancel_absence],
        queue=False
    )
    edit_new_sub.change(
        get_leader_action_button_updates,
        [edit_abs_t, edit_period, edit_new_sub],
        [btn_apply_override, btn_apply_tabadul, btn_apply_penalty, btn_cancel_absence],
        queue=False
    )
    btn_apply_override.click(lambda dfs, at, p, ns, dn, dpt, adm, ca, an, ar: process_admin_action(dfs, at, p, ns, dn, dpt, adm, ca, "normal", an, ar), [current_schedule_state, edit_abs_t, edit_period, edit_new_sub, day_in, dept_in, current_user_is_admin, abs_in, current_user_name, current_user_role], update_outputs)
    btn_apply_tabadul.click(lambda dfs, at, p, ns, dn, dpt, adm, ca, an, ar: process_admin_action(dfs, at, p, ns, dn, dpt, adm, ca, "tabadul", an, ar), [current_schedule_state, edit_abs_t, edit_period, edit_new_sub, day_in, dept_in, current_user_is_admin, abs_in, current_user_name, current_user_role], update_outputs)
    btn_apply_penalty.click(lambda dfs, at, p, ns, dn, dpt, adm, ca, an, ar: process_admin_action(dfs, at, p, ns, dn, dpt, adm, ca, "penalty", an, ar), [current_schedule_state, edit_abs_t, edit_period, edit_new_sub, day_in, dept_in, current_user_is_admin, abs_in, current_user_name, current_user_role], update_outputs)
    btn_cancel_absence.click(
        cancel_teacher_absence_with_generation_state,
        [edit_abs_t, day_in, dept_in, current_user_is_admin, abs_in, current_user_name, current_user_role],
        update_outputs + [btn, btn_regenerate, reserve_generation_state]
    )
    # ── أحداث الخزنة والتقارير والتبادل ─────────────────────────
    t_name.change(
        lambda selected_teacher, is_admin, is_owner: load_teacher_data_for_edit(selected_teacher, is_admin, is_owner) + (gr.update(value=""),),
        [t_name, current_user_is_admin, current_user_is_owner],
        [t_dept_edit, t_val, t_abs_val, t_short_val, t_phone_edit, t_specialty_edit, t_role_edit, vault_status]
    )
    t_dept_edit.change(
    lambda td, d, own: gr.update(
        visible=td in ["العلوم", "المهارات الفردية"] or d in ["العلوم", "المهارات الفردية"],
        interactive=bool(own)
    ),
    [t_dept_edit, dept_in, current_user_is_owner],
    t_specialty_edit
    )
    t_btn.click(update_manual_count, [t_name, t_val, t_abs_val, t_short_val, t_phone_edit, t_specialty_edit, t_role_edit, dept_in, day_in, current_schedule_state, abs_in, current_user_is_admin, current_user_is_owner, current_user_name, current_user_role], [tbl_bal, tbl_abs, tbl_short, tbl_day, vault_status, abs_in, check_teacher_in, rule_teacher])
    t_del_btn.click(delete_single_teacher, [t_name, dept_in, day_in, current_user_is_owner], [tbl_bal, tbl_abs, tbl_short, tbl_day, vault_status, abs_in, check_teacher_in, rule_teacher, t_name])
    export_btn.click(export_excel_report, [dept_in], [report_file])
    reset_month_btn.click(reset_monthly_balances, [dept_in, day_in, current_user_is_admin, current_user_is_owner, current_user_name, current_user_role], [tbl_bal, tbl_abs, tbl_short, tbl_day, monthly_status])
    
    swap_dept.change(
        filter_swap_teachers_safe,
        [swap_dept],
        [swap_t1],
        queue=False
    ).then(
        lambda: gr.update(choices=[], value=None),
        None,
        [swap_p1],
        queue=False
    ).then(
        lambda: ({}, gr.update(value=render_swap_table_html({}))),
        None,
        [swap_confirmed_state, confirmed_tbl_html],
        queue=False
    ).then(
        clear_swap_detail_ui,
        None,
        [swap_options, whatsapp_msg, wa_html_btn, btn_swap_confirm],
        queue=False
    )

    swap_day.change(
        load_confirmed_swaps_for_context,
        [swap_t1, swap_day],
        [swap_confirmed_state, confirmed_tbl_html],
        queue=False
    ).then(
        get_teacher_periods_marked,
        [swap_t1, swap_day, swap_confirmed_state, swap_p1],
        [swap_p1],
        queue=False
    ).then(
        clear_swap_detail_ui,
        None,
        [swap_options, whatsapp_msg, wa_html_btn, btn_swap_confirm],
        queue=False
    )

    swap_t1.change(
        load_confirmed_swaps_for_context,
        [swap_t1, swap_day],
        [swap_confirmed_state, confirmed_tbl_html],
        queue=False
    ).then(
        get_teacher_periods_marked,
        [swap_t1, swap_day, swap_confirmed_state, swap_p1],
        [swap_p1],
        queue=False
    ).then(
        clear_swap_detail_ui,
        None,
        [swap_options, whatsapp_msg, wa_html_btn, btn_swap_confirm],
        queue=False
    )
    swap_p1.change(
        get_swap_candidates_for_period,
        [swap_t1, swap_p1, swap_day, swap_confirmed_state],
        [swap_options, whatsapp_msg, wa_html_btn, btn_swap_confirm],
        queue=False
    )

    swap_options.change(
        on_swap_option_selected,
        [swap_options, swap_t1, swap_p1, swap_day],
        [whatsapp_msg, wa_html_btn, btn_swap_confirm],
        queue=False
    )

    swap_options.input(
        on_swap_option_selected,
        [swap_options, swap_t1, swap_p1, swap_day],
        [whatsapp_msg, wa_html_btn, btn_swap_confirm],
        queue=False
    )

    swap_options.select(
        on_swap_option_selected_from_event,
        [swap_options, swap_t1, swap_p1, swap_day],
        [whatsapp_msg, wa_html_btn, btn_swap_confirm],
        queue=True
    )

    btn_swap_confirm.click(
        confirm_swap,
        [swap_t1, swap_p1, swap_options, swap_day, whatsapp_msg, swap_confirmed_state, current_user_name, current_user_role],
        [swap_confirmed_state, confirmed_tbl_html],
        queue=False
    ).then(
        get_teacher_periods_marked,
        [swap_t1, swap_day, swap_confirmed_state, swap_p1],
        [swap_p1],
        queue=False
    )
    btn_swap_img.click(
        generate_swap_table_image,
        [swap_confirmed_state, swap_t1, swap_day],
        [swap_img_out],
        queue=False
    )
    btn_swap_excel.click(
        export_confirmed_swaps_excel,
        outputs=[swap_excel_out],
        queue=False
    )

app.launch(
    css=css,
    js=js_code,
    server_name="0.0.0.0",
    server_port=7860,
    ssr_mode=False,
    allowed_paths=[
        EXPORTS_DIR,
        IMG_DIR,
        SWAP_IMG_DIR,
    ],
)
